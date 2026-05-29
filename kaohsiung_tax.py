# -*- coding: utf-8 -*-
"""
高雄地政查詢 自動化（多筆標的 × 多位所有權人）
- 支援：解析 JSON、多筆地段地號處理、驗證碼人為節奏重試、PDF 儲存、自動命名與資料夾
- 視窗：--show 會開啟 GUI，左上角(0,0)，1024x1024；
用法：
    python kaohsiung_tax.py --json output\\transcript_data_?????.json --show
或是只給資料夾（會自動挑最新 JSON）：
    python kaohsiung_tax.py --json-dir output --show
"""
import sys
import os

# 🔥 強制使用無緩衝輸出（解決打包環境下輸出延遲問題）
if hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(line_buffering=True)
        sys.stderr.reconfigure(line_buffering=True)
    except:
        pass

# 🔥 打包環境修正：將 _internal 加入 sys.path
if getattr(sys, 'frozen', False):
    # 如果是從打包的 EXE 執行
    internal_dir = os.path.dirname(os.path.abspath(__file__))
    if internal_dir not in sys.path:
        sys.path.insert(0, internal_dir)
elif '__file__' in globals():
    # 開發環境
    script_dir = os.path.dirname(os.path.abspath(__file__))
    if script_dir not in sys.path:
        sys.path.insert(0, script_dir)


import sys
print("\n" + "=" * 78, flush=True)
print("        高雄市土地增值稅自動估算程式", flush=True)
print("=" * 78, flush=True)
print("正在載入程式模組...", flush=True)
sys.stdout.flush()

# 抑制 PyTorch 和其他模組的警告訊息
import warnings
warnings.filterwarnings("ignore", category=UserWarning)
warnings.filterwarnings("ignore", message=".*pin_memory.*")

import io
import os
import re

import json
import time
import math
import ctypes
import keyboard  # 用於系統級鍵盤事件
# 延遲載入重量級模組（只在需要時才載入）
# import numpy as np
# import cv2
# import easyocr
import base64
import random
import argparse
import traceback
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from PyPDF2 import PdfMerger
import tempfile
from PIL import Image
# ---------- Selenium ----------
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import (
    UnexpectedAlertPresentException,
    NoAlertPresentException,
    TimeoutException,
    JavascriptException,
)
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from webdriver_helper import create_chrome_driver, verify_and_fix_chrome_window
from base_dir_helper import get_data_json_path, BASE_DIR, get_work_folder

# 🔥 自動偵測 DPI 縮放比例
def get_dpi_scale():
    """偵測系統 DPI 縮放比例"""
    try:
        hdc = ctypes.windll.user32.GetDC(0)
        dpi = ctypes.windll.gdi32.GetDeviceCaps(hdc, 88)  # LOGPIXELSX
        ctypes.windll.user32.ReleaseDC(0, hdc)
        dpi_scale = dpi / 96.0  # 96 DPI = 100%
        return dpi_scale
    except Exception:
        return 1.0

# 🔥 視窗大小設定：從 main.py 生成的設定檔讀取，自動填滿剩餘空間
kaohsiung_window_width = 1024
kaohsiung_window_height = 1024
kaohsiung_window_x = 0
kaohsiung_window_y = 0

try:
    config_path = os.path.join(BASE_DIR, 'window_config.json')
    if os.path.exists(config_path):
        with open(config_path, 'r', encoding='utf-8') as f:
            window_config = json.load(f)

        work_area = window_config.get('work_area', {})
        main_window = window_config.get('main_window', {})

        if work_area and main_window:
            # 取得 DPI 縮放比例
            dpi_scale = window_config.get('dpi_scale')
            if dpi_scale is None:
                dpi_scale = get_dpi_scale()

            # 取得主程式視窗資訊
            main_x = main_window.get('x', 1280)
            main_y = main_window.get('y', 0)
            main_w = main_window.get('width', 640)
            main_h = main_window.get('height', 1020)

            # 根據主程式位置，智能計算 Chrome 視窗的位置和大小
            work_left = work_area.get('left', 0)
            work_top = work_area.get('top', 0)
            work_right = work_area.get('right', 1920)
            work_bottom = work_area.get('bottom', 1080)

            # 🔥 計算 Chrome 視窗大小：填滿剩餘空間（與其他子程式一致）
            screen_center_x = work_left + (work_right - work_left) // 2

            if main_x < screen_center_x:
                # 主程式在左側，Chrome 放右側，填滿剩餘空間
                chrome_x = main_x + main_w
                chrome_width = work_right - chrome_x
            else:
                # 主程式在右側，Chrome 放左側，填滿剩餘空間
                chrome_x = work_left
                chrome_width = main_x - work_left

            # 高度與主程式相同
            chrome_y = main_y
            chrome_height = main_h

            # 確保視窗大小合理
            if chrome_width < 800:
                chrome_width = 800
            if chrome_height < 600:
                chrome_height = 600

            # 轉換為 Chrome 的邏輯像素
            chrome_width_logical = int(chrome_width / dpi_scale)
            chrome_x_logical = int(chrome_x / dpi_scale)
            chrome_y_logical = int(chrome_y / dpi_scale)

            # 🔥 Chrome 視窗高度需要加上標題欄高度（約 32px 邏輯像素）
            # 這樣整個視窗（包含標題欄）才會和主程式一樣高
            CHROME_TITLEBAR_HEIGHT = 32
            chrome_height_logical = int(chrome_height / dpi_scale) + CHROME_TITLEBAR_HEIGHT

            kaohsiung_window_width = chrome_width_logical
            kaohsiung_window_height = chrome_height_logical
            kaohsiung_window_x = chrome_x_logical
            kaohsiung_window_y = chrome_y_logical
except Exception as e:
    pass

print("✓ 模組載入完成！", flush=True)
sys.stdout.flush()

PDF_SEQUENCE_NUMBER = 0
EXCEL_ROW_COUNTER = 0
OWNER_TOTAL_TRACKER = {}
SECTION_MAPPING_CACHE = {}  # 儲存使用者的地段選擇記憶

KAOHSIUNG_URL = "https://eqs-landp.kcg.gov.tw/KCGQuery/lcaap060f.jsp"

SELECTORS = {
    "district":       'select[name="AA46"]',     # 行政區
    "section_in":     'input[name="AA48"]',      # 段小段（4碼）
    "section_sel":    'select[name="T2"]',       # 段小段下拉
    "lot":            'input[name="AA49"]',      # 地號
    "uid":            'input[name="AA09"]',      # 統一編號/身分證
    "birthday":       'input#LBIR_2',            # 出生日期（民國7碼）
    "captcha_img":    'img#captchaImage',
    "captcha_in":     'input#open_captcha',
    "start_js":       'javascript:submit_str();',
    "refresh_link":   'a[onclick^="myReload("]',
}

# 行政區名稱清單（供文字拆解使用）
KAO_DISTRICTS = [
    "鹽埕區","鼓山區","左營區","楠梓區","三民區","新興區","前金區","苓雅區",
    "前鎮區","旗津區","小港區","鳳山區","林園區","大寮區","大樹區","大社區",
    "仁武區","鳥松區","岡山區","橋頭區","燕巢區","田寮區","阿蓮區","路竹區",
    "湖內區","茄萣區","永安區","彌陀區","梓官區","旗山區","美濃區","六龜區",
    "甲仙區","杉林區","內門區","茂林區","桃源區","那瑪夏區"
]

# 行政區 <option value=""> 對照表（用 value 代碼選取最穩）
KAO_DISTRICT_MAP = {
    "鹽埕區":"EA01","鼓山區":"EA02","左營區":"EE03","楠梓區":"EE04","三民區":"ED05",
    "新興區":"EB06","前金區":"EA07","苓雅區":"EB08","前鎮區":"EC09","旗津區":"EA10",
    "小港區":"EC11","鳳山區":"EG12","林園區":"EL13","大寮區":"EL14","大樹區":"EG15",
    "大社區":"EI16","仁武區":"EI17","鳥松區":"EI18","岡山區":"EF19","橋頭區":"EF20",
    "燕巢區":"EF21","田寮區":"EJ22","阿蓮區":"EJ23","路竹區":"EJ24","湖內區":"EJ25",
    "茄萣區":"EJ26","永安區":"EF27","彌陀區":"EF28","梓官區":"EF29","旗山區":"EH30",
    "美濃區":"EK31","六龜區":"EK32","甲仙區":"EH33","杉林區":"EH34","內門區":"EH35",
    "茂林區":"EK36","桃源區":"EK37","那瑪夏區":"EH38",
}

# ========= 小工具 =========

def now_ymd():
    return datetime.now().strftime("%Y%m%d")

def now_hms():
    return datetime.now().strftime("%H%M%S")

def _hs(a=0.55, b=0.95):
    time.sleep(random.uniform(a, b))

def _norm(s: str) -> str:
    return re.sub(r"\s+", "", (s or "").strip())

def zfill8_digits(s: str) -> str:
    return re.sub(r"\D", "", s or "").zfill(8)[:8]

def to_roc7(s: str) -> str:
    s = (s or "").strip()
    if not s:
        return ""
    
    # 處理 "民國040年01月11日" 格式
    m = re.search(r"民國(\d{3})年(\d{1,2})月(\d{1,2})日", s)
    if m:
        y, mm, dd = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return f"{y:03d}{mm:02d}{dd:02d}"
    

def latest_json_in(folder: str):
    if not os.path.isdir(folder):
        return None
    cands = []
    for fn in os.listdir(folder):
        if not fn.lower().endswith(".json"): 
            continue
        p = os.path.join(folder, fn)
        try:
            cands.append((os.path.getmtime(p), p))
        except Exception:
            pass
    if not cands:
        return None
    return sorted(cands)[-1][1]

def label_for_lot_dir(district: str, section: str, lot8: str) -> str:
    """回傳資料夾名稱：{行政區}{段名}-{前四去零}{可選-後四去零}"""
    lot8 = zfill8_digits(lot8)
    a4 = lot8[:4].lstrip("0") or "0"
    b4 = lot8[4:].lstrip("0")
    label = f"{district}{section}-{a4}"
    if b4:
        label += f"-{b4}"
    return label

def get_next_pdf_sequence():
    """取得下一個PDF序號"""
    global PDF_SEQUENCE_NUMBER
    PDF_SEQUENCE_NUMBER += 1
    return PDF_SEQUENCE_NUMBER

def file_name_for_pdf_with_sequence(district: str, section: str, lot8: str, owner: str) -> str:
    """
    帶序號的PDF檔名生成函數
    格式：01_土增稅001-{行政區}{段名}-{前四去零}{可選-後四去零}_{所有權人}_{YYYYMMDD}
    """
    lot8 = zfill8_digits(lot8)
    a4 = lot8[:4].lstrip("0") or "0"
    b4 = lot8[4:].lstrip("0")
    
    # 建構基本檔名部分
    head = f"{district}{section}-{a4}" if a4 else f"{district}{section}"
    if b4:
        head += f"-{b4}"
    
    # 取得序號
    sequence = get_next_pdf_sequence()
    
    # 組合完整檔名：01_土增稅 + 序號 + 原檔名
    filename = f"01_土增稅{sequence:03d}-{head}_{owner}_{now_ymd()}"
    
    return filename

# ========= 解析 JSON（修正版） =========

def parse_geo_from_title(title: str):
    """
    正確分離：
    林園區王公廟段苦苓腳小段 1183-0000地號
    → 行政區：林園區
    → 段小段：王公廟段苦苓腳小段  
    → 地號：1183-0000 (前四碼1183，後四碼0000)
    """
    title = (title or "").strip()
    district = ""
    section = ""
    lot_front = ""
    lot_tail = ""
    
    # 主要解析模式：行政區 + 段名 + 地號
    pattern = r"(?P<district>..區)(?P<section>.+?段(?:.+?小段)?)\s*(?P<lot_front>\d{3,4})(?:-(?P<lot_tail>\d{4}))?地?號?"
    match = re.search(pattern, title)
    
    if match:
        district = match.group("district")
        section = match.group("section")
        lot_front = match.group("lot_front")
        lot_tail = match.group("lot_tail") or "0000"
        
        return {
            "district": district,
            "section_hint": section,
            "lot_front": lot_front,
            "lot_tail": lot_tail,
            "lot_number": (lot_front + lot_tail).zfill(8)[:8]
        }
    
    # 備用解析：只有地號部分
    lot_match = re.search(r"(?:地號|)\s*(?P<front>\d{3,4})-(?P<tail>\d{4})", title)
    if lot_match:
        lot_front = lot_match.group("front")
        lot_tail = lot_match.group("tail")
        
        # 再嘗試找行政區和段名
        district_match = re.search(r"(?P<district>..區)", title)
        section_match = re.search(r"(?P<section>.+?段(?:.+?小段)?)", title)
        
        if district_match:
            district = district_match.group("district")
        if section_match:
            section = section_match.group("section")
            
        return {
            "district": district,
            "section_hint": section,
            "lot_front": lot_front,
            "lot_tail": lot_tail,
            "lot_number": (lot_front + lot_tail).zfill(8)[:8]
        }

    return {"district": "", "section_hint": "", "lot_front": "", "lot_tail": "", "lot_number": ""}

def validate_transcript_type(block):
    """
    驗證謄本類型：接受土地謄本 - 第一類，或土地謄本 - 第二類（限法人）
    """
    transcript_type = (block or {}).get("謄本類型", "").strip()

    # 檢查是否為土地謄本 - 第一類
    if "土地謄本" in transcript_type and "第一類" in transcript_type:
        return True, transcript_type, "第一類"

    # 🔥 檢查是否為土地謄本 - 第二類（需要額外檢查是否為法人）
    if "土地謄本" in transcript_type and "第二類" in transcript_type:
        # 檢查是否有所有權人資訊
        owners_raw = (block or {}).get("所有權部", [])
        if not owners_raw:
            return False, transcript_type, "第二類（無所有權人資訊）"

        # 檢查是否為法人（有統一編號，8個數字）
        has_legal_entity = False
        for ow in owners_raw:
            if not isinstance(ow, dict):
                continue
            uid = (ow.get("統一編號") or "").strip()
            # 檢查統一編號是否為8個數字
            if uid and uid.isdigit() and len(uid) == 8:
                has_legal_entity = True
                break

        if has_legal_entity:
            return True, transcript_type, "第二類（法人）"
        else:
            return False, transcript_type, "第二類（非法人）"

    return False, transcript_type, "不支援的類型"

def validate_kaohsiung_district(district_text):
    """
    驗證是否為高雄市行政區
    """
    if not district_text:
        return False, "行政區為空"
    
    # 高雄市所有行政區列表
    KAOHSIUNG_DISTRICTS = [
        "鹽埕區", "鼓山區", "左營區", "楠梓區", "三民區", "新興區", "前金區", "苓雅區",
        "前鎮區", "旗津區", "小港區", "鳳山區", "林園區", "大寮區", "大樹區", "大社區",
        "仁武區", "鳥松區", "岡山區", "橋頭區", "燕巢區", "田寮區", "阿蓮區", "路竹區",
        "湖內區", "茄萣區", "永安區", "彌陀區", "梓官區", "旗山區", "美濃區", "六龜區",
        "甲仙區", "杉林區", "內門區", "茂林區", "桃源區", "那瑪夏區"
    ]
    
    district = district_text.strip()
    
    if district in KAOHSIUNG_DISTRICTS:
        return True, f"高雄市{district}"
    
    return False, f"非高雄市行政區：{district}"

def extract_payloads_with_validation(json_path: str):
    """
    帶驗證的 JSON 解析函數
    """
    with open(json_path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    blocks = raw if isinstance(raw, list) else [raw]
    payloads = []
    parse_logs = []
    validation_logs = []

    print(f"開始解析 JSON，共 {len(blocks)} 個區塊...", flush=True)

    for idx, block in enumerate(blocks, 1):
        print(f"\n驗證第 {idx} 筆資料...", flush=True)

        # 驗證1：謄本類型
        is_valid_type, transcript_type, type_detail = validate_transcript_type(block)
        if not is_valid_type:
            validation_logs.append(f"第 {idx} 筆：謄本類型不符 - {type_detail}")
            print(f"X 第 {idx} 筆跳過：謄本類型 '{transcript_type}' - {type_detail}")
            continue

        print(f"✓ 謄本類型符合：{transcript_type} ({type_detail})", flush=True)
        
        # 解析基本資訊
        basic = (block or {}).get("基本資訊") or {}
        title = (basic.get("地號建號") or basic.get("地號") or "").strip()

        geo = parse_geo_from_title(title)
        district_text = geo.get("district") or basic.get("行政區") or ""
        
        # 從段名中補充行政區
        section_hint = (geo.get("section_hint") or basic.get("段名線索") or basic.get("地段線索") or "")
        if not district_text:
            for d in KAO_DISTRICTS:
                if d in section_hint:
                    district_text = d
                    section_hint = section_hint.replace(d, "").strip()
                    break
        
        # 驗證2：高雄市行政區
        is_valid_district, district_msg = validate_kaohsiung_district(district_text)
        if not is_valid_district:
            validation_logs.append(f"第 {idx} 筆：{district_msg}")
            print(f"X 第 {idx} 筆跳過：{district_msg}")
            continue
        
        print(f"✓ 行政區符合：{district_msg}", flush=True)
        
        # 繼續原有的解析邏輯
        lot_number = geo.get("lot_number") or ""
        lot_front = geo.get("lot_front") or ""
        lot_tail = geo.get("lot_tail") or ""

        # 解析所有權人
        owners_raw = (block or {}).get("所有權部") or []
        owners = []
        for ow in owners_raw:
            if not isinstance(ow, dict):
                continue
            name = (ow.get("所有權人") or ow.get("姓名") or "").strip() or "（未命名）"
            uid  = (ow.get("統一編號") or ow.get("身分證字號") or ow.get("身分證號") or "").strip().upper()
            bday = (ow.get("出生日期") or ow.get("生日") or "").strip()

            # 🔥 判斷是否為法人（公司、機構等）
            is_legal_entity = False
            legal_entity_keywords = ['公司', '有限', '股份', '企業', '行號', '商號', '事務所', '協會', '基金會', '財團法人', '社團法人']
            for keyword in legal_entity_keywords:
                if keyword in name:
                    is_legal_entity = True
                    break

            # 🔥 法人不需要出生日期，只處理統一編號
            if is_legal_entity:
                b7 = ""  # 法人不需要出生日期
            else:
                # 自然人才處理出生日期
                b7 = to_roc7(bday) if bday else ""

            # 🔥 修正：只要有名稱就保留（即使沒有統一編號或出生日期）
            # 因為法人可能沒有提供統一編號，但至少有公司名稱
            if name == "（未命名）":
                # 如果連名稱都沒有，檢查是否有統一編號或出生日期
                if not uid and not b7:
                    continue
            owners.append({"name": name, "id_no": uid, "birthday": b7})

        if not owners:
            owners = [{"name": "（未提供）", "id_no": "", "birthday": ""}]

        # 建立顯示用地號
        lot_display = ""
        if lot_front and lot_tail:
            front_clean = lot_front.lstrip("0") or "0"
            tail_clean = lot_tail.lstrip("0")
            if tail_clean:
                lot_display = f"{front_clean}-{tail_clean}"
            else:
                lot_display = front_clean

        payloads.append({
            "district": district_text,
            "section_hint": section_hint,
            "lot": lot_number,
            "lot_display": lot_display,
            "owners": owners,
            "transcript_type": transcript_type  # 保留謄本類型資訊
        })

        # 解析提示
        parse_logs.append(f"✓ 解析 {idx}: 行政區={district_text}；段={section_hint or '？'}；地號={lot_display or lot_number or '？'} ← {title}")
        print(f"✓ 第 {idx} 筆通過驗證，包含 {len(owners)} 位所有權人", flush=True)

    # 顯示驗證結果摘要
    print(f"\n{'='*39}", flush=True)
    print(f"資料驗證結果摘要", flush=True)
    print(f"{'='*39}", flush=True)
    print(f"總筆數：{len(blocks)}", flush=True)
    print(f"通過驗證：{len(payloads)} 筆", flush=True)
    print(f"被拒絕：{len(blocks) - len(payloads)} 筆", flush=True)

    if validation_logs:
        print(f"\n被拒絕的資料：", flush=True)
        for log in validation_logs:
            print(f"  {log}", flush=True)

    print(f"{'='*39}", flush=True)
    
    return payloads, parse_logs
# ========= 開瀏覽器 =========
def start_driver(show_gui: bool = True):
    """永遠優先顯示 GUI；使用智能視窗大小機制，自動填滿剩餘空間。"""
    driver = create_chrome_driver()

    # 🔥 啟動後立即設定視窗大小和位置
    try:
        driver.set_window_size(kaohsiung_window_width, kaohsiung_window_height)
        driver.set_window_position(kaohsiung_window_x, kaohsiung_window_y)
    except Exception as e:
        print(f"[WARNING] 視窗設定失敗: {e}，繼續執行")

    return driver

# ========= DOM 操作 =========

def accept_any_alert(driver):
    try:
        a = driver.switch_to.alert
        txt = ""
        try:
            txt = a.text or ""
        except Exception:
            pass
        a.accept()
        return txt
    except Exception:
        return ""

def _human_pause(a=0.6, b=1.2):
    time.sleep(random.uniform(a, b))

# ========= DOM：選行政區 + 段小段（修正版） =========
def select_district_and_section_with_memory(driver, district_text: str, section_hint: str, wait_sec: int = 15):
    """
    具備記憶功能的段小段選擇：自動匹配失敗時讓使用者選擇，並記住選擇
    """
    global SECTION_MAPPING_CACHE
    
    print("=" * 38)
    print("=" * 38)
    
    # 檢查是否已有快取的選擇
    cache_key = f"{district_text}_{section_hint}"
    if cache_key in SECTION_MAPPING_CACHE:
        cached_choice = SECTION_MAPPING_CACHE[cache_key]
        print(f"使用快取的選擇: {section_hint} -> {cached_choice['text']}")
        
        # 執行行政區選擇
        sel_district = WebDriverWait(driver, wait_sec).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, SELECTORS["district"]))
        )
        
        WebDriverWait(driver, wait_sec).until(
            lambda d: len(sel_district.find_elements(By.TAG_NAME, "option")) >= 5
        )
        
        dist_code = KAO_DISTRICT_MAP.get(district_text.strip(), "")
        if dist_code:
            for opt in sel_district.find_elements(By.TAG_NAME, "option"):
                if (opt.get_attribute("value") or "").strip() == dist_code:
                    opt.click()
                    break
        
        # 載入段小段選項
        sel_section = WebDriverWait(driver, wait_sec).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, SELECTORS["section_sel"]))
        )
        sel_section.click()
        time.sleep(0.5)
        driver.execute_script("SelectChange(arguments[0]);", sel_district)
        time.sleep(0.5)
        
        # 使用快取的選擇
        try:
            driver.execute_script(f"arguments[0].value = '{cached_choice['value']}';", sel_section)
            driver.execute_script("arguments[0].dispatchEvent(new Event('change'));", sel_section)
            print(f"SUCCESS: 已使用快取選擇段小段 - '{cached_choice['text']}'")
            return True
        except Exception as e:
            print(f"快取選擇失敗: {e}")
            # 繼續執行正常流程
    
    # 1) 行政區選擇
    print(f"STEP 1: 選擇行政區...")
    sel_district = WebDriverWait(driver, wait_sec).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, SELECTORS["district"]))
    )
    
    WebDriverWait(driver, wait_sec).until(
        lambda d: len(sel_district.find_elements(By.TAG_NAME, "option")) >= 5
    )
    
    opts = sel_district.find_elements(By.TAG_NAME, "option")

    dist_code = KAO_DISTRICT_MAP.get(district_text.strip(), "")
    picked = False
    
    if dist_code:
        for opt in opts:
            if (opt.get_attribute("value") or "").strip() == dist_code:
                opt.click()
                picked = True
                print(f"SUCCESS: 已選擇行政區: {district_text}")
                break

    if not picked:
        print(f"ERROR: 無法找到行政區 '{district_text}'")
        return False

    # 2) 點擊段小段下拉並觸發載入
    print("\n" + "=" * 40)
    print("STEP 2: 載入段小段選項...")
    print("=" * 40)
    
    sel_section = WebDriverWait(driver, wait_sec).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, SELECTORS["section_sel"]))
    )
    
    sel_section.click()
    time.sleep(0.5)
    driver.execute_script("SelectChange(arguments[0]);", sel_district)
    time.sleep(0.5)
    
    # 3) 獲取所有段小段選項
    print("STEP 3: 獲取所有段小段選項...")
    
    opts_all = sel_section.find_elements(By.TAG_NAME, "option")

    if opts_all and "請先選擇行政區" in (opts_all[0].text or ""):
        sel_section.click()
        time.sleep(0.5)
        driver.execute_script("SelectChange(arguments[0]);", sel_district)
        time.sleep(0.5)
        opts_all = sel_section.find_elements(By.TAG_NAME, "option")
    
    # 提取有效選項
    valid_options = []
    for i, opt in enumerate(opts_all):
        value = opt.get_attribute("value") or ""
        text = (opt.text or "").strip()
        if text and "請先選擇" not in text and "請選擇" not in text:
            valid_options.append((opt, text, value))
    
    
    if not valid_options:
        print("ERROR: 沒有找到有效的段小段選項")
        return False
    
    # 4) 自動匹配段小段
    print("\n" + "=" * 20)
    print("STEP 4: 自動匹配段小段...")
    print("=" * 20)
    
    want = section_hint.replace("　", "").replace(" ", "").strip()
    
    chosen = None
    match_info = ""
    
    # 策略1: 完全匹配
    for opt, text, value in valid_options:
        clean_text = text.replace("　", "").replace(" ", "")
        if want == clean_text:
            chosen = opt
            match_info = f"完全匹配: '{text}'"
            print(f"SUCCESS: {match_info}")
            break
    
    # 策略2: 包含匹配
    if not chosen:
        print("策略2 - 包含匹配...")
        for opt, text, value in valid_options:
            clean_text = text.replace("　", "").replace(" ", "")
            if want in clean_text or clean_text in want:
                chosen = opt
                match_info = f"包含匹配: '{text}'"
                print(f"SUCCESS: {match_info}")
                break
    
    # 策略3: 關鍵字匹配
    if not chosen:
        print("策略3 - 關鍵字匹配...")
        keywords = re.findall(r'[\u4e00-\u9fff]{2,}', section_hint)

        for keyword in keywords:
            for opt, text, value in valid_options:
                if keyword in text:
                    chosen = opt
                    match_info = f"關鍵字匹配: '{text}' (關鍵字: {keyword})"
                    print(f"SUCCESS: {match_info}")
                    break
            if chosen:
                break
    
    # 策略4: 模糊匹配
    if not chosen:
        print("策略4 - 模糊匹配...")
        want_fuzzy = want.replace("段", "").replace("小", "")
        
        for opt, text, value in valid_options:
            text_fuzzy = text.replace("段", "").replace("小", "").replace("　", "").replace(" ", "")
            if want_fuzzy and want_fuzzy in text_fuzzy:
                chosen = opt
                match_info = f"模糊匹配: '{text}'"
                print(f"SUCCESS: {match_info}")
                break
    
    # 5) 如果自動匹配失敗，讓使用者手動選擇
    if not chosen:
        print("\n" + "=" * 50)
        print("警告: 自動匹配失敗 - 請手動選擇")
        print("=" * 50)
        print(f"無法自動匹配段名: '{section_hint}'")
        print(f"行政區: {district_text}")
        print("\n可用的段小段選項：")
        
        for i, (opt, text, value) in enumerate(valid_options, 1):
            print(f"  {i:2d}. {text}")
        
        # 修改 select_district_and_section_with_memory 函數中的互動部分：

        # 先顯示所有訊息，確保都能看到
        print(f"\n請選擇正確的段小段 (1-{len(valid_options)})，或輸入 0 跳過:", flush=True)
        sys.stdout.flush()
        time.sleep(0.3)  # 確保訊息顯示

        while True:
            try:
                # 純輸入，不帶提示訊息
                choice = input().strip()
                
                if choice == '0':
                    print("用戶選擇跳過此筆資料", flush=True)
                    sys.stdout.flush()
                    return False
                
                choice_num = int(choice)
                if 1 <= choice_num <= len(valid_options):
                    chosen = valid_options[choice_num - 1][0]
                    chosen_text = valid_options[choice_num - 1][1]
                    chosen_value = valid_options[choice_num - 1][2]
                    
                    # 先顯示選擇結果
                    print(f"您選擇了: 使用者選擇: '{chosen_text}'", flush=True)
                    sys.stdout.flush()
                    time.sleep(0.3)
                    
                    # 然後顯示確認提示
                    print(f"確認選擇 '{chosen_text}' ? (y/n):", flush=True)
                    sys.stdout.flush()
                    time.sleep(0.3)  # 確保提示訊息顯示
                    
                    # 確認選擇的輸入循環
                    while True:
                        try:
                            # 純輸入，不帶提示
                            confirm = input().strip().lower()
                            
                            if confirm in ['y', 'yes', '是']:
                                # 顯示記憶詢問訊息
                                print("之後遇到相同情形時，是否自動選擇相同地段? (y/n):", flush=True)
                                sys.stdout.flush()
                                time.sleep(0.3)
                                
                                # 記憶選擇的輸入循環
                                while True:
                                    try:
                                        remember = input().strip().lower()
                                        
                                        if remember in ['y', 'yes', '是']:
                                            SECTION_MAPPING_CACHE[cache_key] = {
                                                'text': chosen_text,
                                                'value': chosen_value
                                            }
                                            print(f"已記住選擇: {section_hint} -> {chosen_text}", flush=True)
                                            sys.stdout.flush()
                                            break
                                        elif remember in ['n', 'no', '否']:
                                            print("不記住此次選擇", flush=True)
                                            sys.stdout.flush()
                                            break
                                        else:
                                            print("請輸入 y 或 n:", flush=True)
                                            sys.stdout.flush()
                                            time.sleep(0.3)
                                    except KeyboardInterrupt:
                                        break
                                break
                                
                            elif confirm in ['n', 'no', '否']:
                                print("請重新選擇...", flush=True)
                                sys.stdout.flush()
                                print(f"請選擇正確的段小段 (1-{len(valid_options)})，或輸入 0 跳過:", flush=True)
                                sys.stdout.flush()
                                time.sleep(0.3)
                                chosen = None
                                break
                            else:
                                print("請輸入 y 或 n:", flush=True)
                                sys.stdout.flush()
                                time.sleep(0.3)
                        except KeyboardInterrupt:
                            print("\n用戶中止操作", flush=True)
                            sys.stdout.flush()
                            return False
                    
                    if chosen:
                        break
                else:
                    print(f"請輸入 1-{len(valid_options)} 之間的數字:", flush=True)
                    sys.stdout.flush()
                    time.sleep(0.3)
                    
            except ValueError:
                print("請輸入有效的數字:", flush=True)
                sys.stdout.flush()
                time.sleep(0.3)
            except KeyboardInterrupt:
                print("\n用戶中止操作", flush=True)
                sys.stdout.flush()
                return False
    
    # 6) 選擇找到的選項
    print(f"\nSTEP 5: 選擇段小段...")
    
    try:
        driver.execute_script("arguments[0].scrollIntoView(true);", chosen)
        time.sleep(0.3)
        chosen.click()
        print(f"SUCCESS: 已選擇段小段 - '{chosen.text}'")
        
        # 添加確認選擇是否成功的檢查
        time.sleep(0.3)
        current_value = sel_section.get_attribute("value")
        if current_value:
            print(f"確認選擇成功，當前值: {current_value}")
            return True
        else:
            print("選擇可能未生效，嘗試備用方法...")
            raise Exception("選擇未生效")
            
    except Exception as e:
        print(f"ERROR: 點擊段小段選項失敗: {e}")
        try:
            print("嘗試使用 JavaScript 選擇...")
            value = chosen.get_attribute("value")
            driver.execute_script(f"arguments[0].value = '{value}';", sel_section)
            driver.execute_script("arguments[0].dispatchEvent(new Event('change'));", sel_section)
            time.sleep(0.3)
            print(f"SUCCESS: 使用 JavaScript 選擇成功")
            return True
        except Exception as e2:
            print(f"ERROR: JavaScript 選擇也失敗: {e2}")
            return False

def ensure_form_filled(driver, lot8: str, uid: str, bday7: str):
    """確保表單欄位正確填入"""
    
    # 地號
    if lot8 and re.fullmatch(r"\d{8}", lot8):
        el = driver.find_element(By.CSS_SELECTOR, SELECTORS["lot"])
        val = el.get_attribute("value") or ""
        if val.strip().replace(" ", "") != lot8:
            el.clear()
            time.sleep(0.3)
            el.send_keys(lot8)
            time.sleep(0.3)
            print(f"📝 已填地號：{lot8}")

    # 統編/身分證（可選）
    if uid:
        el = driver.find_element(By.CSS_SELECTOR, SELECTORS["uid"])
        cur = (el.get_attribute("value") or "").strip().upper()
        if cur != uid.strip().upper():
            el.clear()
            time.sleep(0.3)
            el.send_keys(uid)
            time.sleep(0.3)
            print(f"📝 已填統編/身分證：{uid}")

    # 生日（民國 7 碼，可選）
    if bday7 and re.fullmatch(r"\d{7}", bday7):
        el = driver.find_element(By.CSS_SELECTOR, SELECTORS["birthday"])
        cur = (el.get_attribute("value") or "").strip()
        if cur != bday7:
            el.clear()
            time.sleep(0.3)
            el.send_keys(bday7)
            time.sleep(0.5)  # 等待網頁可能觸發的驗證碼更新
            print(f"📝 已填生日：{bday7}")

# ========= 驗證碼與提交 =========
def safe_input(prompt="", check_interval=1.0):
    """
    安全的 input 函數，會定期檢查瀏覽器是否還存活
    如果瀏覽器被關閉，會拋出異常以終止等待
    支援輸入 'ESC' 來退出程式

    Args:
        prompt: 提示訊息
        check_interval: 檢查瀏覽器狀態的間隔（秒）
    """
    import threading
    import queue

    # 創建一個 queue 來傳遞輸入結果
    result_queue = queue.Queue()
    input_received = threading.Event()

    def input_thread():
        """執行緒中執行 input"""
        try:
            user_input = input(prompt)
            result_queue.put(('success', user_input))
        except (KeyboardInterrupt, EOFError) as e:
            result_queue.put(('interrupt', str(e)))
        finally:
            input_received.set()

    # 啟動輸入執行緒
    thread = threading.Thread(target=input_thread, daemon=True)
    thread.start()

    # 定期檢查瀏覽器狀態，直到收到輸入或發現瀏覽器關閉
    while not input_received.is_set():
        # 嘗試獲取當前 URL 來檢查連接是否還存在
        try:
            # 使用全域的 driver 變數
            import sys
            current_module = sys.modules[__name__]
            if hasattr(current_module, 'driver'):
                _ = current_module.driver.current_url
        except Exception:
            # driver 連接已斷開（瀏覽器被關閉）
            print("\n偵測到瀏覽器已關閉，程式將退出", flush=True)
            raise SystemExit("瀏覽器已關閉")

        # 等待一小段時間後再次檢查
        input_received.wait(timeout=check_interval)

    # 獲取輸入結果
    status, value = result_queue.get()
    if status == 'interrupt':
        print("\n程式被中斷", flush=True)
        raise SystemExit("使用者中斷")

    return value

def _read_captcha_digits(driver) -> str:
    """
    自動辨識驗證碼數字
    """
    try:
        # 延遲載入重量級模組（只在需要時才載入）
        import numpy as np
        import cv2

        # 尋找驗證碼圖片
        captcha_img = driver.find_element(By.CSS_SELECTOR, SELECTORS["captcha_img"])

        # 截取驗證碼圖片 - 使用溫和的方式避免閃爍
        # 🔥 使用 JavaScript 直接取得圖片的 base64，避免 screenshot 造成的閃爍
        img_base64 = driver.execute_script("""
            var img = arguments[0];
            var canvas = document.createElement('canvas');
            canvas.width = img.naturalWidth || img.width;
            canvas.height = img.naturalHeight || img.height;
            var ctx = canvas.getContext('2d');
            ctx.drawImage(img, 0, 0);
            return canvas.toDataURL('image/png').split(',')[1];
        """, captcha_img)

        import base64
        png_data = base64.b64decode(img_base64)
        nparr = np.frombuffer(png_data, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_GRAYSCALE)

        if img is None:
            print("無法讀取驗證碼圖片")
            return ""

        # 圖片預處理
        # 1. 二值化處理
        _, img_binary = cv2.threshold(img, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

        # 2. 放大圖片提高辨識率
        img_resized = cv2.resize(img_binary, None, fx=3, fy=3, interpolation=cv2.INTER_CUBIC)

        # 3. 降噪處理
        kernel = np.ones((2,2), np.uint8)
        img_cleaned = cv2.morphologyEx(img_resized, cv2.MORPH_CLOSE, kernel)
        
        # OCR 辨識
        print("開始 OCR 辨識驗證碼...")
        result_text = ""

        # 方法1: 嘗試 RapidOCR
        try:
            from rapidocr_helper import create_rapidocr_reader, rapidocr_readtext
            reader = create_rapidocr_reader(
                use_angle_cls=False,
                use_text_score=True,
                text_score=0.3,
                det_db_thresh=0.2,
                det_db_box_thresh=0.3,
                det_db_unclip_ratio=1.8,
                max_side_len=1280
            )
            # 將 OpenCV 圖片轉換為 PIL Image 格式
            from PIL import Image
            img_pil = Image.fromarray(img_cleaned)
            results = rapidocr_readtext(reader, np.array(img_pil), detail=0)
            if results:
                # 只保留數字
                result_text = ''.join(filter(str.isdigit, ''.join(results)))
                print(f" RapidOCR 結果: '{result_text}'")
        except ImportError:
            print(" RapidOCR 未安裝")
        except Exception as e:
            print(f" RapidOCR 失敗: {e}")

        # 清理結果：只保留數字
        clean_result = ''.join(filter(str.isdigit, result_text))

        # 驗證結果長度（通常驗證碼是4-6位數字）
        if len(clean_result) >= 4:
            # 🔥 如果超過 6 位，取前 6 位（可能是 OCR 誤判多出的字元）
            if len(clean_result) > 6:
                print(f"⚠️ OCR 辨識長度超過 6 位: '{clean_result}' (長度: {len(clean_result)})")
                clean_result = clean_result[:6]
                print(f"   截取前 6 位: '{clean_result}'")

            print(f"SUCCESS: OCR 辨識成功: '{clean_result}'")
            return clean_result
        else:
            print(f"❌ OCR 結果長度不足: '{clean_result}' (長度: {len(clean_result)})")
            return ""

    except Exception as e:
        print(f" OCR 處理失敗: {e}")
        return ""

def check_and_refill_form(driver, district, section, lot8, uid, bday7):
    """
    檢查並重新填寫表單（包含行政區和段小段）
    """
    try:
        print("檢查表單狀態並重新填寫...")

        # 檢查是否還在正確的頁面
        current_url = driver.current_url
        if "lcaap060f.jsp" not in current_url:
            print("頁面已跳轉，需要重新進入表單頁")
            try:
                driver.get(KAOHSIUNG_URL)
            except Exception as conn_error:
                error_msg = str(conn_error)
                if "ERR_CONNECTION_TIMED_OUT" in error_msg or "ERR_NAME_NOT_RESOLVED" in error_msg or "ERR_CONNECTION_REFUSED" in error_msg:
                    print("\033[91m" + "=" * 70, flush=True)
                    print("⚠️  高雄市土地增值稅網站連線失敗", flush=True)
                    print("=" * 70, flush=True)
                    print(f"網址: {KAOHSIUNG_URL}", flush=True)
                    print("\n可能原因:", flush=True)
                    print("  1. 網站伺服器暫時無法連線或維護中", flush=True)
                    print("  2. 網路連線不穩定", flush=True)
                    print("  3. 網站網址已變更", flush=True)
                    print("\n建議處理:", flush=True)
                    print("  1. 請稍後再試", flush=True)
                    print("  2. 檢查網路連線狀態", flush=True)
                    print("  3. 手動開啟瀏覽器確認網站是否正常運作", flush=True)
                    print("=" * 70 + "\033[0m", flush=True)
                    raise
                else:
                    raise
            WebDriverWait(driver, 12).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, SELECTORS["district"]))
            )
            time.sleep(0.3)

        # 🔥 修正：無論頁面是否跳轉，都重新選擇行政區和段小段
        # 因為驗證碼錯誤後，下拉選單會被重置
        print("重新選擇行政區和段小段...")
        success = select_district_and_section_with_memory(driver, district, section)
        if not success:
            print("重新選擇行政區/段小段失敗")
            return False

        # 重新填寫表單欄位
        ensure_form_filled(driver, lot8, uid, bday7)
        print("表單重新填寫完成")
        return True

    except Exception as e:
        print(f"檢查表單失敗: {e}")
        return False

def solve_captcha_and_submit_fixed_v2(driver, district, section, lot8, uid, bday7, wait=None, max_try=3, submit_js="submit_str()"):
    """
    改進版：前 3 次自動辨識，第 4 次開始手動確認並預填到 GUI
    """

    # 🔥 前 3 次自動嘗試
    for i in range(1, max_try + 1):
        print(f"\n\033[93m--- 驗證碼第 {i} 次嘗試 ---\033[0m", flush=True)

        # 從第二次開始，檢查並重新填寫表單
        if i > 1:
            success = check_and_refill_form(driver, district, section, lot8, uid, bday7)
            if not success:
                print("重新填寫表單失敗", flush=True)
                continue

        # 嘗試自動辨識
        code = _read_captcha_digits(driver)

        if not code:
            print("OCR辨識失敗", flush=True)
            code = ""

        # 填入驗證碼
        try:
            inp = driver.find_element(By.CSS_SELECTOR, SELECTORS["captcha_in"])
            inp.clear()
            time.sleep(0.3)
            inp.send_keys(code)
            time.sleep(0.3)
            print(f"\033[93m驗證碼第 {i} 次嘗試：{code}\033[0m", flush=True)
        except Exception as e:
            print(f"驗證碼輸入失敗: {e}", flush=True)
            continue

        # 提交表單
        try:
            driver.execute_script(submit_js + ";")
            print("表單已提交", flush=True)
        except Exception:
            try:
                submit_btn = driver.find_element(By.CSS_SELECTOR, 'input[type="submit"],button[type="submit"]')
                submit_btn.click()
                print("使用按鈕提交", flush=True)
            except Exception:
                print("提交失敗", flush=True)

        time.sleep(0.5)

        # 檢查警示
        try:
            WebDriverWait(driver, 3.0).until(EC.alert_is_present())
            alert = driver.switch_to.alert
            txt = (alert.text or "").strip()
            alert.accept()
            print(f"系統警示: {txt}", flush=True)

            if "驗證碼" in txt:
                print(f"驗證碼錯誤，第 {i} 次失敗。", flush=True)
                continue
            elif "查詢條件不完整" in txt:
                print("查詢條件不完整，將在下次迴圈重新填寫表單", flush=True)
                continue
            else:
                print("其他錯誤，重試...", flush=True)
                continue

        except TimeoutException:
            # 沒有警示，檢查是否成功
            pass

        # 檢查是否進入結果頁
        try:
            WebDriverWait(driver, 8).until(lambda d: "lcaap061f.jsp" in d.current_url)
            print("✅ 查詢成功，進入結果頁面！", flush=True)
            return True
        except TimeoutException:
            print("等待結果頁面超時", flush=True)
            continue

    # 🔥 前 3 次失敗後，進入手動確認模式（最多10次重試）
    print("所有自動驗證嘗試均失敗。", flush=True)
    print("進入手動驗證模式...", flush=True)

    max_manual_attempts = 10
    for manual_attempt in range(1, max_manual_attempts + 1):
        print(f"\n=== 手動驗證第 {manual_attempt}/{max_manual_attempts} 次 ===", flush=True)

        # 檢查並重新填寫表單
        success = check_and_refill_form(driver, district, section, lot8, uid, bday7)
        if not success:
            print("重新填寫表單失敗", flush=True)
            if manual_attempt == max_manual_attempts:
                print("已達到最大重試次數，停止嘗試", flush=True)
                return False
            continue

        # 🔥 在填入驗證碼前，先驗證表單欄位是否已填寫
        try:
            lot_input = driver.find_element(By.CSS_SELECTOR, SELECTORS["lot"])
            uid_input = driver.find_element(By.CSS_SELECTOR, SELECTORS["uid"])

            if not lot_input.get_attribute("value") or not uid_input.get_attribute("value"):
                print("⚠️ 表單欄位未正確填寫，重新填寫...", flush=True)
                ensure_form_filled(driver, lot8, uid, bday7)
        except Exception as e:
            print(f"檢查表單欄位時出錯: {e}", flush=True)

        # 🔧 自動辨識驗證碼並帶入主程式輸入框供使用者確認
        print("正在自動辨識驗證碼...", flush=True)
        auto_code = _read_captcha_digits(driver)

        if auto_code:
            print(f"\033[93m自動辨識到驗證碼: {auto_code}\033[0m", flush=True)
            print("請檢查驗證碼是否正確：", flush=True)
            print("  - 若正確，請直接按【ENTER】送出", flush=True)
            print("  - 若需修改，請輸入正確的驗證碼後按【ENTER】", flush=True)
            print("  - 輸入 ESC 可退出程式", flush=True)

            # 🔥 發送特殊標記給主程式，通知預填驗證碼到 GUI 輸入框
            print(f"__GUI_PREFILL__:{auto_code}", flush=True)

            # 🔥 讓使用者在主程式輸入框中確認或修改
            user_input = safe_input(f"驗證碼 [{auto_code}]: ").strip()

            if user_input.upper() == "ESC":
                print("使用者選擇退出", flush=True)
                return False
            elif user_input:  # 使用者輸入了修改內容
                code = user_input
            else:  # 使用者直接按ENTER，使用自動辨識的驗證碼
                code = auto_code
        else:
            # 自動辨識失敗，請求手動輸入
            print("自動辨識失敗，請手動輸入驗證碼（或輸入 ESC 退出）：", flush=True)
            user_input = safe_input("驗證碼: ").strip()

            if user_input.upper() == "ESC":
                print("使用者選擇退出", flush=True)
                return False
            code = user_input

        # 將最終確認的驗證碼填入網頁
        try:
            inp = driver.find_element(By.CSS_SELECTOR, SELECTORS["captcha_in"])
            inp.clear()
            time.sleep(0.3)
            inp.send_keys(code)
            time.sleep(0.3)
            print(f"驗證碼 [{code}]: 已填入驗證碼", flush=True)
        except Exception as e:
            print(f"驗證碼輸入失敗: {e}", flush=True)
            if manual_attempt == max_manual_attempts:
                print("已達到最大重試次數，停止嘗試", flush=True)
                return False
            continue

        # 提交表單
        try:
            driver.execute_script(submit_js + ";")
            print("表單已提交", flush=True)
        except Exception:
            try:
                submit_btn = driver.find_element(By.CSS_SELECTOR, 'input[type="submit"],button[type="submit"]')
                submit_btn.click()
                print("使用按鈕提交", flush=True)
            except Exception:
                print("提交失敗", flush=True)

        time.sleep(0.5)

        # 檢查警示
        try:
            WebDriverWait(driver, 3.0).until(EC.alert_is_present())
            alert = driver.switch_to.alert
            txt = (alert.text or "").strip()
            alert.accept()
            print(f"系統警示: {txt}", flush=True)

            if "驗證碼" in txt:
                print(f"驗證碼錯誤，第 {manual_attempt} 次手動嘗試失敗。", flush=True)
                if manual_attempt == max_manual_attempts:
                    print("已達到最大重試次數，停止嘗試", flush=True)
                    return False
                continue  # 🔥 重新循環，再次嘗試
            elif "查詢條件不完整" in txt:
                print("查詢條件不完整，將在下次迴圈重新填寫表單", flush=True)
                if manual_attempt == max_manual_attempts:
                    print("已達到最大重試次數，停止嘗試", flush=True)
                    return False
                continue
            else:
                print("其他錯誤，重試...", flush=True)
                if manual_attempt == max_manual_attempts:
                    print("已達到最大重試次數，停止嘗試", flush=True)
                    return False
                continue

        except TimeoutException:
            # 沒有警示，檢查是否成功
            pass

        # 檢查是否進入結果頁
        try:
            WebDriverWait(driver, 8).until(lambda d: "lcaap061f.jsp" in d.current_url)
            print(f"✅ 驗證成功！（手動第 {manual_attempt} 次）", flush=True)
            return True
        except TimeoutException:
            print("等待結果頁面超時", flush=True)
            manual_attempt += 1
            continue  # 🔥 重新循環，再次嘗試


def _refresh_captcha_like_human(driver):
    """刷新驗證碼圖片"""
    try:
        # 方法1: 點擊刷新連結
        refresh_link = driver.find_element(By.CSS_SELECTOR, SELECTORS["refresh_link"])
        refresh_link.click()
        print("🔄 點擊刷新連結")
        time.sleep(random.uniform(1.0, 2.0))
        return
    except:
        pass
    
    try:
        # 方法2: 點擊驗證碼圖片
        captcha_img = driver.find_element(By.CSS_SELECTOR, SELECTORS["captcha_img"])
        captcha_img.click()
        print("🔄 點擊驗證碼圖片")
        time.sleep(random.uniform(1.0, 2.0))
        return
    except:
        pass
    
    try:
        # 方法3: JavaScript 刷新
        driver.execute_script("""
            var img = document.getElementById('captchaImage');
            if(img) {
                var src = img.src.split('&time=')[0];
                img.src = src + (src.includes('?') ? '&' : '?') + 'time=' + Date.now();
            }
        """)
        print("🔄 JavaScript 刷新驗證碼")
        time.sleep(random.uniform(1.0, 2.0))
    except Exception as e:
        print(f"⚠️ 驗證碼刷新失敗: {e}")
        time.sleep(0.3)

# ========= PDF =========
def extract_result_page_info(driver):
    """
    從結果頁面提取土地增值稅資訊
    """
    try:
        print("📄 開始解析結果頁面...")
        
        # 等待頁面完全載入
        time.sleep(0.5)
        
        result_info = {}
        
        # 嘗試找到主要內容區域
        try:
            # 查找標題
            title_elements = driver.find_elements(By.CSS_SELECTOR, "h1, h2, h3, .title, [class*='title']")
            for elem in title_elements:
                text = elem.text.strip()
                if text and ("土地增值稅" in text or "試算" in text):
                    result_info["標題"] = text
                    print(f"📋 找到標題：{text}")
                    break
            
            # 查找表格或列表形式的資料
            tables = driver.find_elements(By.CSS_SELECTOR, "table")
            if tables:
                print(f"📊 找到 {len(tables)} 個表格")
                for i, table in enumerate(tables):
                    table_text = table.text.strip()
                    if table_text:
                        result_info[f"表格{i+1}"] = table_text
                        print(f"表格 {i+1} 內容預覽：{table_text[:100]}...")
            
            # 查找所有包含資料的 div
            content_divs = driver.find_elements(By.CSS_SELECTOR, "div")
            important_content = []
            
            for div in content_divs:
                text = div.text.strip()
                # 過濾出包含重要關鍵字的內容
                if text and any(keyword in text for keyword in 
                              ["增值稅", "稅額", "現值", "地價", "面積", "所有權人", "地號", "持分"]):
                    if len(text) > 10 and len(text) < 500:  # 避免過短或過長的內容
                        important_content.append(text)
            
            if important_content:
                result_info["重要內容"] = important_content
                print(f"📝 找到 {len(important_content)} 項重要內容")
            
            # 查找警告或錯誤訊息
            warning_elements = driver.find_elements(By.CSS_SELECTOR, 
                "[class*='warning'], [class*='error'], [class*='alert'], .message")
            for elem in warning_elements:
                text = elem.text.strip()
                if text:
                    result_info["系統訊息"] = text
                    print(f"⚠️ 系統訊息：{text}")
            
            # 如果沒找到結構化內容，就取整個頁面的文字
            if not result_info or len(result_info) <= 1:
                body_text = driver.find_element(By.TAG_NAME, "body").text.strip()
                if body_text:
                    # 只取前1000字符，避免太長
                    result_info["頁面內容"] = body_text[:1000]
                    print("📄 提取頁面完整內容")
            
        except Exception as e:
            print(f"⚠️ 內容解析部分失敗：{e}")
            
        return result_info
        
    except Exception as e:
        print(f"❌ 結果頁面解析失敗：{e}")
        return {}

def parse_tax_calculation_data_fixed(page_text):
    """
    修正版：從頁面文字中解析土地增值稅試算數據
    """
    data = {}
    
    try:
        print(f" 開始解析頁面文字，總長度: {len(page_text)}")
        print(f" 前200字符: {page_text[:200]}")
        
        # 基本資訊 - 修正正則表達式
        # 匹配格式如: 左營區(0227)0181-0000
        lot_patterns = [
            r'([^(]+)\((\d+)\)(\d{4}-\d{4})',  # 原格式
            r'([^(]+)\((\d+)\)(\d+)',          # 簡化格式
            r'([^(]+區)[^0-9]*(\d+)[^0-9]*(\d{4}-\d{4})',  # 更寬鬆匹配
        ]
        
        for pattern in lot_patterns:
            lot_match = re.search(pattern, page_text)
            if lot_match:
                data['行政區'] = lot_match.group(1).strip()
                if len(lot_match.groups()) >= 3:
                    data['區碼'] = lot_match.group(2) if lot_match.group(2) else ''
                    data['地號'] = lot_match.group(3)
                print(f" 找到地號資訊: {data}")
                break
        
        # 土地面積 - 改進匹配
        area_patterns = [
            r'土地面積[：:\s]*([0-9,]+\.?\d*)',
            r'面積[：:\s]*([0-9,]+\.?\d*)',
        ]
        
        for pattern in area_patterns:
            area_match = re.search(pattern, page_text)
            if area_match:
                data['土地面積'] = area_match.group(1)
                print(f" 找到土地面積: {data['土地面積']}")
                break
        
        # 當期土地公告現值
        current_value_patterns = [
            r'當期土地公告現值[：:\s]*([0-9,]+)元[/／]平方公尺',
            r'當期.*?公告現值[：:\s]*([0-9,]+)元',
            r'公告現值[：:\s]*([0-9,]+)元',
        ]
        
        for pattern in current_value_patterns:
            current_value_match = re.search(pattern, page_text)
            if current_value_match:
                data['當期公告現值'] = current_value_match.group(1)
                print(f" 找到當期公告現值: {data['當期公告現值']}")
                break
        
        # 前次移轉現值 - 改進匹配，處理多筆資料（包含小數點）
        transfer_patterns = [
            r'前次移轉現值[：:\s]*([0-9,]+\.?[0-9]*)元[/／]平方公尺',
            r'前次移轉現值[：:\s]*([0-9,]+\.?[0-9]*)元',
            r'移轉現值[：:\s]*([0-9,]+\.?[0-9]*)元',
        ]
        
        transfer_values = []
        for pattern in transfer_patterns:
            matches = re.findall(pattern, page_text)
            if matches:
                transfer_values.extend(matches)
                break
        
        if transfer_values:
            for i, value in enumerate(transfer_values):
                key = f'第{i+1}次前次移轉現值' if len(transfer_values) > 1 else '前次移轉現值'
                data[key] = value
                print(f" 找到前次移轉現值 {i+1}: {value}")
        
        # 原地價年月
        date_patterns = [
            r'原地價年月[：:\s]*(\d+年\d+月)',
            r'年月[：:\s]*(\d+年\d+月)',
        ]
        
        dates = []
        for pattern in date_patterns:
            matches = re.findall(pattern, page_text)
            if matches:
                dates.extend(matches)
                break
        
        if dates:
            for i, date in enumerate(dates):
                key = f'第{i+1}次原地價年月' if len(dates) > 1 else '原地價年月'
                data[key] = date
                print(f" 找到原地價年月 {i+1}: {date}")
        
        # 權利範圍
        ratio_patterns = [
            r'歷次權利範圍[：:\s]*(\d+分之\d+)',
            r'權利範圍[：:\s]*(\d+分之\d+)',
            r'(\d+分之\d+)',
        ]
        
        ratios = []
        for pattern in ratio_patterns:
            matches = re.findall(pattern, page_text)
            if matches:
                ratios.extend(matches)
                break
        
        if ratios:
            for i, ratio in enumerate(ratios):
                key = f'第{i+1}次權利範圍' if len(ratios) > 1 else '權利範圍'
                data[key] = ratio
                print(f" 找到權利範圍 {i+1}: {ratio}")
        
        # 物價指數
        price_index_patterns = [
            r'物價指數[：:\s]*([0-9.]+)',
            r'指數[：:\s]*([0-9.]+)',
        ]
        
        indices = []
        for pattern in price_index_patterns:
            matches = re.findall(pattern, page_text)
            if matches:
                indices.extend(matches)
                break
        
        if indices:
            for i, index in enumerate(indices):
                key = f'第{i+1}次物價指數' if len(indices) > 1 else '物價指數'
                data[key] = index
                print(f" 找到物價指數 {i+1}: {index}")
        
        # 自用住宅用地應納稅額
        residential_tax_patterns = [
            r'自用住宅用地應納稅額[：:\s]*([0-9,]+)元',
            r'自用住宅.*?稅額[：:\s]*([0-9,]+)元',
            r'住宅.*?稅額[：:\s]*([0-9,]+)元',
        ]
        
        residential_taxes = []
        for pattern in residential_tax_patterns:
            matches = re.findall(pattern, page_text)
            if matches:
                residential_taxes.extend(matches)
                break
        
        if residential_taxes:
            for i, tax in enumerate(residential_taxes):
                key = f'第{i+1}次自用住宅稅額' if len(residential_taxes) > 1 else '自用住宅稅額'
                data[key] = tax
                print(f" 找到自用住宅稅額 {i+1}: {tax}")
        
        # 一般土地應納稅額
        general_tax_patterns = [
            r'一般土地應納稅額[：:\s]*([0-9,]+)元',
            r'一般.*?稅額[：:\s]*([0-9,]+)元',
            r'土地.*?稅額[：:\s]*([0-9,]+)元',
        ]
        
        general_taxes = []
        for pattern in general_tax_patterns:
            matches = re.findall(pattern, page_text)
            if matches:
                general_taxes.extend(matches)
                break
        
        if general_taxes:
            for i, tax in enumerate(general_taxes):
                key = f'第{i+1}次一般土地稅額' if len(general_taxes) > 1 else '一般土地稅額'
                data[key] = tax
                print(f" 找到一般土地稅額 {i+1}: {tax}")
        
        # 添加查詢時間
        data['查詢日期'] = datetime.now().strftime("%Y-%m-%d")
        data['查詢時間'] = datetime.now().strftime("%H:%M:%S")
        
        print(f" 最終解析結果: {data}")
        
    except Exception as e:
        print(f"數據解析錯誤: {e}")
        data['解析錯誤'] = str(e)
    
    return data

#=======================================================================
def create_vertical_excel_workbook_final():
    """
    最終版：創建垂直展開格式的Excel工作簿
    """
    wb = Workbook()
    sheet = wb.active
    sheet.title = '土地增值稅查詢結果'

    # 第一列：標題欄（A1~O1 跨欄合併）
    sheet.merge_cells('A1:O1')
    title_cell = sheet['A1']
    title_cell.value = "高 雄 市 自 動 估 算 土 地 增 值 稅 列 表"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    sheet.row_dimensions[1].height = 25  # 設定標題列高度

    # 🔥 第二列：保留給20年提示（稍後可能添加）
    # 暫時保留空白，如果需要提示會在後面填入

    # 第三列：欄位標題
    headers = [
        '編號', '行政區', '地段', '地號', '所有權人', '統一編號', '生日',
        '土地面積', '當期公告現值',
        '前次移轉現值', '原地價年月', '權利範圍', '物價指數',
        '自用住宅稅額', '一般土地稅額'
    ]

    # 設定標題行（第三列）
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 設定列寬
    column_widths = [8, 10, 15, 12, 15, 15, 12, 12, 15, 15, 12, 15, 10, 15, 15]
    for i, width in enumerate(column_widths, 1):
        sheet.column_dimensions[sheet.cell(row=2, column=i).column_letter].width = width

    # 設定頁面為 A4 橫式列印，所有欄位在同一頁寬中
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE  # 橫式
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4  # A4
    sheet.page_setup.fitToPage = True  # 縮放以適應頁面
    sheet.page_setup.fitToWidth = 1  # 所有欄位適應一頁寬
    sheet.page_setup.fitToHeight = 0  # 高度不限制（自動分頁）

    # 設定列印邊界（單位：英吋）
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.75
    sheet.page_margins.bottom = 0.75
    sheet.page_margins.header = 0.3
    sheet.page_margins.footer = 0.3

    return wb, sheet

def parse_multiple_transfer_data_final(page_text, fallback_lot_number=""):
    """
    最終版：解析多筆前次移轉資料，加入無權利人檢測
    """
    data = {}
    transfer_records = []
    
    try:
        print(f" 開始解析多筆移轉資料，頁面長度: {len(page_text)}")
        
        # 檢查是否為"無此權利人"情況
        no_rights_patterns = [
            r'該地號無此權利人',
            r'無此權利人',
            r'查無權利人',
            r'無相關權利人'
        ]
        
        is_no_rights = False
        for pattern in no_rights_patterns:
            if re.search(pattern, page_text):
                is_no_rights = True
                print(f" 檢測到無權利人情況: {pattern}")
                break
        
        if is_no_rights:
            # 無權利人的情況，使用fallback地號
            data['地號'] = fallback_lot_number
            data['無權利人'] = True
            data['移轉記錄'] = []
            return data
        
        # 基本資訊解析
        lot_patterns = [
            r'([^(]+)\((\d+)\)(\d{4}-\d{4})',
            r'([^(]+)\((\d+)\)(\d+)',
            r'([^(]+區)[^0-9]*(\d+)[^0-9]*(\d{4}-\d{4})',
            # r'(左營區|小港區|鳳山區|[^(]+區)[^0-9]*(\d+)[^0-9]*(\d{4}-\d{4})',
        ]
        
        for pattern in lot_patterns:
            lot_match = re.search(pattern, page_text)
            if lot_match:
                data['行政區'] = lot_match.group(1).strip()
                if len(lot_match.groups()) >= 3:
                    data['區碼'] = lot_match.group(2) if lot_match.group(2) else ''
                    data['地號'] = lot_match.group(3)
                print(f" 解析地號資訊: {data}")
                break
        
        # 如果無法從頁面解析地號，使用fallback
        if not data.get('地號') and fallback_lot_number:
            data['地號'] = fallback_lot_number
            print(f" 使用fallback地號: {fallback_lot_number}")
        
        # 土地面積
        area_match = re.search(r'土地面積[：:\s]*([0-9,]+\.?\d*)', page_text)
        if area_match:
            data['土地面積'] = area_match.group(1)
            print(f" 土地面積: {data['土地面積']}")
        
        # 當期公告現值
        current_value_match = re.search(r'當期土地公告現值[：:\s]*([0-9,]+)元[/／]平方公尺', page_text)
        if current_value_match:
            data['當期公告現值'] = current_value_match.group(1)
            print(f" 當期公告現值: {data['當期公告現值']}")
        
        # 改進的多筆資料解析
        print(" 開始解析移轉記錄...")

        transfer_values = re.findall(r'前次移轉現值[：:\s]*([0-9,]+\.?[0-9]*)元', page_text)
        print(f" 找到 {len(transfer_values)} 個前次移轉現值: {transfer_values}")
        
        if transfer_values:
            # 分割文本，以每個「前次移轉現值」為起點
            transfer_sections = []
            current_pos = 0
            
            for i, value in enumerate(transfer_values):
                pattern = rf'前次移轉現值[：:\s]*{re.escape(value)}元'
                match = re.search(pattern, page_text[current_pos:])
                if match:
                    start_pos = current_pos + match.start()
                    
                    if i < len(transfer_values) - 1:
                        next_value = transfer_values[i + 1]
                        next_pattern = rf'前次移轉現值[：:\s]*{re.escape(next_value)}元'
                        next_match = re.search(next_pattern, page_text[start_pos + 50:])
                        if next_match:
                            end_pos = start_pos + 50 + next_match.start()
                        else:
                            end_pos = len(page_text)
                    else:
                        end_pos = len(page_text)
                    
                    section_text = page_text[start_pos:end_pos]
                    transfer_sections.append((value, section_text))
                    current_pos = end_pos
            
            print(f" 分割出 {len(transfer_sections)} 個移轉區段")
            
            # 解析每個區段
            for i, (transfer_value, section) in enumerate(transfer_sections):
                record = {'前次移轉現值': transfer_value}
                
                date_match = re.search(r'原地價年月[：:\s]*(\d+年\d+月)', section)
                if date_match:
                    record['原地價年月'] = date_match.group(1)
                
                ratio_match = re.search(r'歷次權利範圍[：:\s]*(\d+分之\d+)', section)
                if ratio_match:
                    record['權利範圍'] = ratio_match.group(1)
                
                price_index_match = re.search(r'物價指數[：:\s]*([0-9.]+)', section)
                if price_index_match:
                    record['物價指數'] = price_index_match.group(1)
                
                residential_tax_match = re.search(r'自用住宅用地應納稅額[：:\s]*([0-9,]+)元', section)
                if residential_tax_match:
                    record['自用住宅稅額'] = residential_tax_match.group(1)
                
                general_tax_match = re.search(r'一般土地應納稅額[：:\s]*([0-9,]+)元', section)
                if general_tax_match:
                    record['一般土地稅額'] = general_tax_match.group(1)
                
                transfer_records.append(record)
                print(f" 第 {i+1} 筆移轉記錄: {record}")
        
        # 備用方法
        if not transfer_records:
            print(" 使用備用解析方法...")
            dates = re.findall(r'原地價年月[：:\s]*(\d+年\d+月)', page_text)
            ratios = re.findall(r'歷次權利範圍[：:\s]*(\d+分之\d+)', page_text)
            indices = re.findall(r'物價指數[：:\s]*([0-9.]+)', page_text)
            residential_taxes = re.findall(r'自用住宅用地應納稅額[：:\s]*([0-9,]+)元', page_text)
            general_taxes = re.findall(r'一般土地應納稅額[：:\s]*([0-9,]+)元', page_text)
            
            max_records = max(len(transfer_values), len(residential_taxes), len(general_taxes))
            
            for i in range(max_records):
                record = {
                    '前次移轉現值': transfer_values[i] if i < len(transfer_values) else '',
                    '原地價年月': dates[i] if i < len(dates) else '',
                    '權利範圍': ratios[i] if i < len(ratios) else '',
                    '物價指數': indices[i] if i < len(indices) else '',
                    '自用住宅稅額': residential_taxes[i] if i < len(residential_taxes) else '',
                    '一般土地稅額': general_taxes[i] if i < len(general_taxes) else ''
                }
                transfer_records.append(record)
        
        data['移轉記錄'] = transfer_records
        data['查詢日期'] = datetime.now().strftime("%Y-%m-%d")
        data['查詢時間'] = datetime.now().strftime("%H:%M:%S")
        
        print(f" 總共解析到 {len(transfer_records)} 筆移轉記錄")
        
    except Exception as e:
        print(f"多筆資料解析錯誤: {e}")
        data['解析錯誤'] = str(e)
        data['移轉記錄'] = []
    
    return data

# 全域變量追蹤編號
EXCEL_ROW_COUNTER = 0

# 需要在程式中添加一個追蹤同一所有權人的機制
OWNER_DATA_TRACKER = {}  # 追蹤每個所有權人的累計資料

def add_vertical_rows_to_excel_with_owner_total(wb, sheet, query_data, owner_data, query_status, district_name, section_name, is_last_record_for_owner=False, error_msg=None):
    """
    支援同一所有權人總加總的Excel寫入函數
    """
    global EXCEL_ROW_COUNTER, OWNER_TOTAL_TRACKER
    
    try:
        transfer_records = query_data.get('移轉記錄', [])
        is_no_rights = query_data.get('無權利人', False)
        
        if not transfer_records and not is_no_rights:
            transfer_records = [{}]
        
        start_row = sheet.max_row + 1
        owner_name = owner_data.get('name', '')
        owner_id = owner_data.get('id_no', '')
        
        owner_key = f"{owner_name}_{owner_id}"

        if owner_key not in OWNER_TOTAL_TRACKER:
            # 🔥 取得第一筆移轉記錄的原地價年月
            first_transfer_date = ''
            if transfer_records and len(transfer_records) > 0:
                first_transfer_date = transfer_records[0].get('原地價年月', '')

            OWNER_TOTAL_TRACKER[owner_key] = {
                'total_residential_tax': 0,
                'total_general_tax': 0,
                'record_count': 0,
                'owner_display_name': owner_name,
                'owner_id': owner_id,
                '原地價年月': first_transfer_date  # 🔥 儲存原地價年月
            }

        if is_no_rights:
            owner_name += '-無此權利人'
            transfer_records = [{}]
        
        
        EXCEL_ROW_COUNTER += 1
        
        base_data = [
            EXCEL_ROW_COUNTER,
            district_name,
            section_name,
            query_data.get('地號', ''),
            owner_name,
            owner_data.get('id_no', ''),
            owner_data.get('birthday', ''),
            query_data.get('土地面積', ''),
            query_data.get('當期公告現值', '')
        ]
        
        if is_no_rights:
            row_data = base_data + ['', '', '', '', '', '']
            
            for col, value in enumerate(row_data, 1):
                cell = sheet.cell(row=start_row, column=col, value=value)
                cell.fill = PatternFill(start_color="FFF8E8", end_color="FFF8E8", fill_type="solid")
            
            print(f"已將無權利人記錄寫入Excel第 {start_row} 行")
            
            if is_last_record_for_owner:
                add_owner_total_row(sheet, owner_key)
            
            return
        
        residential_total = 0
        general_total = 0
        
        for i, record in enumerate(transfer_records):
            current_row = start_row + i
            
            row_data = base_data.copy()
            
            if i > 0:
                for j in range(1, 9):
                    row_data[j] = ''
            
            row_data.extend([
                record.get('前次移轉現值', ''),
                record.get('原地價年月', ''),
                record.get('權利範圍', ''),
                record.get('物價指數', ''),
                record.get('自用住宅稅額', ''),
                record.get('一般土地稅額', '')
            ])
            
            try:
                residential_tax = record.get('自用住宅稅額', '')
                if residential_tax:
                    residential_num = int(str(residential_tax).replace(',', '').replace('元', '').strip())
                    residential_total += residential_num
                    OWNER_TOTAL_TRACKER[owner_key]['total_residential_tax'] += residential_num
            except (ValueError, TypeError) as e:
                pass

            try:
                general_tax = record.get('一般土地稅額', '')
                if general_tax:
                    general_num = int(str(general_tax).replace(',', '').replace('元', '').strip())
                    general_total += general_num
                    OWNER_TOTAL_TRACKER[owner_key]['total_general_tax'] += general_num
            except (ValueError, TypeError) as e:
                pass
            
            for col, value in enumerate(row_data, 1):
                cell = sheet.cell(row=current_row, column=col, value=value)
                
                if query_status == '成功':
                    cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                elif query_status in ['失敗', '錯誤']:
                    cell.fill = PatternFill(start_color="FFE8E8", end_color="FFE8E8", fill_type="solid")
        
        OWNER_TOTAL_TRACKER[owner_key]['record_count'] += 1
        
        if len(transfer_records) >= 1 and (residential_total > 0 or general_total > 0):
            total_row = start_row + len(transfer_records)
            
            total_data = [''] * 15
            total_data[9] = '小計'
            total_data[13] = f"{residential_total:,}" if residential_total > 0 else ''
            total_data[14] = f"{general_total:,}" if general_total > 0 else ''
            
            for col, value in enumerate(total_data, 1):
                cell = sheet.cell(row=total_row, column=col, value=value)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        
        if is_last_record_for_owner:
            add_owner_total_row(sheet, owner_key)
        
        print(f"已將 {owner_name} 的 {len(transfer_records)} 筆記錄寫入Excel")
        
    except Exception as e:
        print(f"垂直Excel寫入錯誤: {e}")
        traceback.print_exc()

def add_owner_total_row(sheet, owner_key):
    """
    添加所有權人總計行
    """
    global OWNER_TOTAL_TRACKER
    
    try:
        owner_totals = OWNER_TOTAL_TRACKER[owner_key]
        
        if owner_totals['total_residential_tax'] > 0 or owner_totals['total_general_tax'] > 0:
            total_row = sheet.max_row + 1
            
            total_data = [''] * 15
            total_data[9] = f"【{owner_totals['owner_display_name']} 總計】"
            total_data[13] = f"{owner_totals['total_residential_tax']:,}" if owner_totals['total_residential_tax'] > 0 else ''
            total_data[14] = f"{owner_totals['total_general_tax']:,}" if owner_totals['total_general_tax'] > 0 else ''
            
            for col, value in enumerate(total_data, 1):
                cell = sheet.cell(row=total_row, column=col, value=value)
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            
            print(f"已添加 {owner_totals['owner_display_name']} 的總計行")
            
            sheet.cell(row=total_row + 1, column=1, value='')
        
    except Exception as e:
        print(f"添加所有權人總計行失敗: {e}")

def fill_one_owner_with_complete_features(driver, payload, owner, wait, outdir_root, excel_wb, excel_sheet, is_last_record_for_owner=False):
    """
    完整功能版本的所有權人處理流程
    """
    
    district = payload.get("district", "").strip()
    section = payload.get("section_hint", "").strip()
    lot8 = zfill8_digits(payload.get("lot", ""))
    lot_display = payload.get("lot_display", "")
    uid = owner.get("id_no", "").strip().upper()
    bday7 = owner.get("birthday", "").strip()
    name = owner.get("name") or "（未命名）"

    print(f"處理資料：{district}{section} 地號 {lot_display} - {name}")

    if not district or not (uid or bday7):
        print("資料不足，跳過。")
        error_data = {'行政區': district, '地號': lot_display, '移轉記錄': []}
        add_vertical_rows_to_excel_with_owner_total(excel_wb, excel_sheet, error_data, owner, '資料不足', district, section, is_last_record_for_owner, '缺少行政區或身分資料')
        return None, False

    zoom_applied = False  # 🔥 標記是否已經執行過縮放，避免重複縮放

    for attempt in range(3):
        try:
            print(f"第 {attempt + 1} 次嘗試連線至高雄市土地增值稅網站...", flush=True)
            sys.stdout.flush()

            # 進入表單頁面
            try:
                driver.get(KAOHSIUNG_URL)
                sys.stdout.flush()  # 確保後續訊息立即顯示
            except Exception as conn_error:
                error_msg = str(conn_error)
                if "ERR_CONNECTION_TIMED_OUT" in error_msg or "ERR_NAME_NOT_RESOLVED" in error_msg or "ERR_CONNECTION_REFUSED" in error_msg:
                    print("\033[91m" + "=" * 70, flush=True)
                    print("⚠️  高雄市土地增值稅網站連線失敗", flush=True)
                    print("=" * 70, flush=True)
                    print(f"網址: {KAOHSIUNG_URL}", flush=True)
                    print("\n可能原因:", flush=True)
                    print("  1. 網站伺服器暫時無法連線或維護中", flush=True)
                    print("  2. 網路連線不穩定", flush=True)
                    print("  3. 網站網址已變更", flush=True)
                    print("\n建議處理:", flush=True)
                    print("  1. 請稍後再試", flush=True)
                    print("  2. 檢查網路連線狀態", flush=True)
                    print("  3. 手動開啟瀏覽器確認網站是否正常運作", flush=True)
                    print("=" * 70 + "\033[0m", flush=True)

                    if attempt < 2:
                        print(f"\n等待 5 秒後進行第 {attempt + 2} 次重試...", flush=True)
                        time.sleep(5)
                        continue
                    else:
                        print("\n已達最大重試次數，程式結束。", flush=True)
                        raise
                else:
                    raise

            # 網頁載入後重新設定視窗大小（避免被網站重置）
            try:
                driver.set_window_size(kaohsiung_window_width, kaohsiung_window_height)
                driver.set_window_position(kaohsiung_window_x, kaohsiung_window_y)
            except:
                pass

            # 🔥 DPI 自動修正（只在 DPI 不同步時動作，正常 PC 無影響）
            verify_and_fix_chrome_window(driver)

            # 🔥 自動判斷是否需要縮放頁面（解決高 DPI 或小螢幕下的排版問題）
            # 根據 DPI 決定縮放次數：150% → 2次(80%)，175% → 4次(67%)
            # 🔥 只在第一次嘗試時執行縮放，避免重試時重複縮放
            if not zoom_applied:
                zoom_count = 0
                try:
                    config_path = os.path.join(BASE_DIR, 'window_config.json')
                    if os.path.exists(config_path):
                        with open(config_path, 'r', encoding='utf-8') as f:
                            config = json.load(f)

                        current_dpi = config.get('dpi_scale', 1.0)
                        screen_width = config.get('screen_width', 1920)
                        logical_width = screen_width / current_dpi

                        # 🔥 根據 DPI 和螢幕尺寸決定縮放次數
                        if current_dpi >= 1.75:
                            zoom_count = 4  # 175% 或更高：縮放到 67%（按 4 次）
                            target_zoom = "67%"
                        elif current_dpi >= 1.5:
                            zoom_count = 2  # 150%：縮放到 80%
                            target_zoom = "80%"
                        elif current_dpi >= 1.25 and logical_width < 1200:
                            zoom_count = 2  # 小螢幕 + 125%：縮放到 80%
                            target_zoom = "80%"
                        else:
                            zoom_count = 0  # 100%、大螢幕125%：不縮放
                            target_zoom = "100%"

                        if zoom_count > 0:
                            print(f"[頁面縮放] 偵測到 DPI={current_dpi}x (即 {int(current_dpi * 100)}%), 邏輯寬度={logical_width:.0f}px → 需要縮放至 {target_zoom}")
                        else:
                            print(f"[頁面縮放] 偵測到 DPI={current_dpi}x (即 {int(current_dpi * 100)}%), 邏輯寬度={logical_width:.0f}px → 不需要縮放")
                except Exception as e:
                    print(f"[頁面縮放] 偵測配置時發生錯誤: {e}，使用預設不縮放")

                if zoom_count > 0:
                    try:
                        # 🔥 使用 keyboard 庫發送系統級的 Ctrl + - 按鍵（真正的瀏覽器縮放）
                        # 先點擊頁面確保 Chrome 視窗有焦點
                        body = driver.find_element(By.TAG_NAME, 'body')
                        body.click()
                        time.sleep(0.5)

                        # 🔥 根據 DPI 決定按幾次 Ctrl + -
                        print(f"[頁面縮放] 正在使用系統鍵盤縮放（按 {zoom_count} 次）...")
                        for i in range(zoom_count):
                            keyboard.press_and_release('ctrl+-')  # 系統級按鍵
                            time.sleep(0.5)

                        print(f"[頁面縮放] ✓ 已使用系統鍵盤縮放至 {target_zoom}")
                    except Exception as e:
                        print(f"[頁面縮放] 設定縮放時發生錯誤: {e}")
                else:
                    print("[頁面縮放] 螢幕配置正常，不需要縮放")

                zoom_applied = True  # 🔥 標記已執行過縮放

            WebDriverWait(driver, 12).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, SELECTORS["district"]))
            )
            time.sleep(0.5)
            print("表單頁面已載入")

            # 選擇地段
            print("開始選擇行政區和段小段...")
            success = select_district_and_section_with_memory(driver, district, section)
            
            if not success:
                print("地段選擇失敗，準備重試...")
                if attempt < 2:
                    print(f"將進行第 {attempt + 2} 次重試")
                    continue
                else:
                    print("已達到最大重試次數，地段選擇失敗")
                    error_data = {'行政區': district, '地號': lot_display, '移轉記錄': []}
                    add_vertical_rows_to_excel_with_owner_total(excel_wb, excel_sheet, error_data, owner, '選區失敗', district, section, is_last_record_for_owner, '無法選擇行政區或段小段')
                    return None, False
            else:
                print("地段選擇成功，準備填寫其他欄位...")
                # 添加短暫延遲確保頁面更新
                time.sleep(0.5)

            # 填寫表單欄位
            print("填寫地號和所有權人資料...")
            ensure_form_filled(driver, lot8, uid, bday7)
            print("表單欄位填寫完成")

            # 處理驗證碼並提交
            print("開始處理驗證碼...")
            ok = solve_captcha_and_submit_fixed_v2(driver, district, section, lot8, uid, bday7, wait=None, max_try=3)
            
            if ok:
                print("查詢成功！收集帶標識的PDF數據...")
                
                try:
                    # 收集PDF數據
                    pdf_data = collect_page_data_with_owner_label(driver, name)
                    
                    # 解析頁面內容
                    print("解析查詢結果...")
                    page_text = driver.find_element(By.TAG_NAME, "body").text
                    query_data = parse_multiple_transfer_data_final(page_text, lot_display)
                    
                    # 記錄到Excel
                    print("將結果記錄到Excel...")
                    add_vertical_rows_to_excel_with_owner_total(excel_wb, excel_sheet, query_data, owner, '成功', district, section, is_last_record_for_owner)
                    
                    print("帶標識的資料收集完成")
                    return pdf_data, True
                    
                except Exception as e:
                    print(f"結果處理失敗: {e}")
                    traceback.print_exc()
                    error_data = {'行政區': district, '地號': lot_display, '移轉記錄': []}
                    add_vertical_rows_to_excel_with_owner_total(excel_wb, excel_sheet, error_data, owner, '解析失敗', district, section, is_last_record_for_owner, str(e))
                    return None, False
            else:
                print("驗證碼處理失敗")
                if attempt < 2:
                    print(f"將進行第 {attempt + 2} 次重試")
                    continue
                else:
                    print("已達到最大重試次數，驗證碼處理失敗")
                    error_data = {'行政區': district, '地號': lot_display, '移轉記錄': []}
                    add_vertical_rows_to_excel_with_owner_total(excel_wb, excel_sheet, error_data, owner, '查詢失敗', district, section, is_last_record_for_owner, '驗證碼處理失敗')
                    return None, False

        except Exception as e:
            error_msg = str(e)
            # 如果是連線錯誤，不顯示技術性錯誤訊息（因為前面已經顯示友善訊息了）
            is_connection_error = any(err in error_msg for err in ["ERR_CONNECTION_TIMED_OUT", "ERR_NAME_NOT_RESOLVED", "ERR_CONNECTION_REFUSED"])

            if not is_connection_error:
                print(f"第 {attempt + 1} 次嘗試發生錯誤：{e}", flush=True)
                traceback.print_exc()

            if attempt >= 2:
                if not is_connection_error:
                    print("已達到最大重試次數，處理失敗", flush=True)
                error_data = {'行政區': district, '地號': lot_display, '移轉記錄': []}
                add_vertical_rows_to_excel_with_owner_total(excel_wb, excel_sheet, error_data, owner, '系統錯誤', district, section, is_last_record_for_owner, str(e))
            else:
                print(f"將進行第 {attempt + 2} 次重試")
            continue

    print("所有嘗試都失敗，返回失敗結果")
    return None, False

def add_vertical_rows_to_excel_final(wb, sheet, query_data, owner_data, query_status, district_name, section_name, error_msg=None):
    """
    最終版：將查詢結果以垂直展開方式添加到Excel，修正編號跳號問題
    """
    global EXCEL_ROW_COUNTER
    
    try:
        transfer_records = query_data.get('移轉記錄', [])
        is_no_rights = query_data.get('無權利人', False)
        
        if not transfer_records and not is_no_rights:
            transfer_records = [{}]
        
        start_row = sheet.max_row + 1
        owner_name = owner_data.get('name', '')
        
        # 處理無權利人情況
        if is_no_rights:
            owner_name += '-無此權利人'
            transfer_records = [{}]  # 至少要有一行來顯示基本資料
        
        print(f" 準備寫入Excel，共 {len(transfer_records)} 筆記錄，無權利人: {is_no_rights}")
        
        # 編號遞增（只在寫入第一行時）
        EXCEL_ROW_COUNTER += 1
        
        # 基本資料（修正：加入行政區欄位）
        base_data = [
            EXCEL_ROW_COUNTER,  # 編號（使用全域計數器）
            district_name,      # 行政區
            section_name,       # 地段
            query_data.get('地號', ''),
            owner_name,         # 所有權人（可能包含-無此權利人）
            owner_data.get('id_no', ''),
            owner_data.get('birthday', ''),
            query_data.get('土地面積', ''),
            query_data.get('當期公告現值', '')
        ]
        
        # 如果是無權利人，只寫一行基本資料
        if is_no_rights:
            row_data = base_data + ['', '', '', '', '', '']  # 移轉資料部分留空
            
            for col, value in enumerate(row_data, 1):
                cell = sheet.cell(row=start_row, column=col, value=value)
                cell.fill = PatternFill(start_color="FFF8E8", end_color="FFF8E8", fill_type="solid")  # 淺黃色
            
            print(f"已將無權利人記錄寫入Excel第 {start_row} 行")
            return
        
        # 正常情況：寫入每筆移轉記錄
        residential_total = 0
        general_total = 0
        
        for i, record in enumerate(transfer_records):
            current_row = start_row + i
            
            # 組合完整的行資料
            row_data = base_data.copy()
            
            # 如果不是第一行，基本資料部分留空（除了編號）
            if i > 0:
                for j in range(1, 9):  # 行政區到當期公告現值
                    row_data[j] = ''
            
            # 添加移轉記錄資料
            row_data.extend([
                record.get('前次移轉現值', ''),
                record.get('原地價年月', ''),
                record.get('權利範圍', ''),
                record.get('物價指數', ''),
                record.get('自用住宅稅額', ''),
                record.get('一般土地稅額', '')
            ])
            
            # 計算稅額總計
            try:
                residential_tax = record.get('自用住宅稅額', '')
                if residential_tax:
                    residential_num = int(str(residential_tax).replace(',', '').replace('元', '').strip())
                    residential_total += residential_num
            except (ValueError, TypeError) as e:
                print(f" 住宅稅額轉換失敗: {residential_tax}, 錯誤: {e}")
            
            try:
                general_tax = record.get('一般土地稅額', '')
                if general_tax:
                    general_num = int(str(general_tax).replace(',', '').replace('元', '').strip())
                    general_total += general_num
            except (ValueError, TypeError) as e:
                print(f" 一般稅額轉換失敗: {general_tax}, 錯誤: {e}")
            
            # 寫入Excel
            for col, value in enumerate(row_data, 1):
                cell = sheet.cell(row=current_row, column=col, value=value)
                
                # 設定樣式
                if query_status == '成功':
                    cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                elif query_status in ['失敗', '錯誤']:
                    cell.fill = PatternFill(start_color="FFE8E8", end_color="FFE8E8", fill_type="solid")
        
        # 添加合計行（只有在有稅額時才加）
        if len(transfer_records) >= 1 and (residential_total > 0 or general_total > 0):
            total_row = start_row + len(transfer_records)
            
            # 合計行資料（15個欄位）
            total_data = [''] * 15
            
            # 在前次移轉現值欄位寫"合計"（第10欄，索引9）
            total_data[9] = '合計'
            
            # 稅額合計（第14、15欄，索引13、14）
            total_data[13] = f"{residential_total:,}" if residential_total > 0 else ''
            total_data[14] = f"{general_total:,}" if general_total > 0 else ''
            
            # 寫入合計行
            for col, value in enumerate(total_data, 1):
                cell = sheet.cell(row=total_row, column=col, value=value)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        
        print(f"已將 {owner_name} 的 {len(transfer_records)} 筆記錄寫入Excel")
        
    except Exception as e:
        print(f"垂直Excel寫入錯誤: {e}")
        traceback.print_exc()


def collect_page_data(driver):
    """
    收集當前頁面資料，返回頁面截圖或PDF數據
    """
    try:
        # 等待頁面完全載入
        time.sleep(2)

        # 🔥 修正 PDF 中文字型問題：注入 CSS 確保使用正確的中文字型
        try:
            font_fix_css = """
            <style id="pdf-font-fix">
                * {
                    font-family: "Microsoft YaHei", "微軟正黑體", "Arial", "sans-serif" !important;
                    -webkit-font-smoothing: antialiased !important;
                    -moz-osx-font-smoothing: grayscale !important;
                }
                body, table, td, th, div, span, p {
                    font-family: "Microsoft YaHei", "微軟正黑體", "Arial", "sans-serif" !important;
                }
            </style>
            """
            driver.execute_script(f"""
                var style = document.createElement('style');
                style.id = 'pdf-font-fix';
                style.innerHTML = `
                    * {{
                        font-family: "Microsoft YaHei", "微軟正黑體", "Arial", "sans-serif" !important;
                        -webkit-font-smoothing: antialiased !important;
                        -moz-osx-font-smoothing: grayscale !important;
                    }}
                    body, table, td, th, div, span, p {{
                        font-family: "Microsoft YaHei", "微軟正黑體", "Arial", "sans-serif" !important;
                    }}
                `;
                document.head.appendChild(style);
            """)
            time.sleep(0.5)  # 等待 CSS 生效
        except Exception as e:
            print(f"注入字型 CSS 失敗（不影響功能）: {e}")

        # 方法1：使用 Chrome DevTools Protocol 獲取PDF數據
        try:
            result = driver.execute_cdp_cmd("Page.printToPDF", {
                "landscape": False,
                "printBackground": True,
                "scale": 0.8,
                "paperWidth": 8.27,   # A4寬度
                "paperHeight": 11.69,  # A4高度
                "marginTop": 0.39,
                "marginBottom": 0.39,
                "marginLeft": 0.39,
                "marginRight": 0.39
            })
            
            if 'data' in result:
                pdf_data = base64.b64decode(result['data'])
                print(f"成功獲取PDF數據，大小: {len(pdf_data)} bytes")
                return pdf_data
                
        except Exception as e:
            print(f"PDF數據獲取失敗: {e}")
        
        # 方法2：截圖備用
        try:
            total_height = driver.execute_script("return document.body.scrollHeight")
            driver.set_window_size(1024, total_height)
            time.sleep(0.5)
            
            screenshot = driver.get_screenshot_as_png()
            print(f"成功獲取截圖數據，大小: {len(screenshot)} bytes")
            return screenshot
            
        except Exception as e:
            print(f"截圖獲取失敗: {e}")
            return None
            
    except Exception as e:
        print(f"頁面數據收集失敗: {e}")
        return None

def add_owner_label_to_page(driver, owner_name):
    """
    在網頁上添加所有權人標識
    """
    try:
        js_code = f"""
        var ownerLabel = document.createElement('div');
        ownerLabel.innerHTML = '所有權人：{owner_name}';
        ownerLabel.style.cssText = `
            position: fixed;
            top: 10px;
            right: 20px;
            background-color: #f0f0f0;
            border: 2px solid #333;
            padding: 8px 15px;
            font-size: 16px;
            font-weight: bold;
            z-index: 9999;
            border-radius: 5px;
            box-shadow: 0 2px 5px rgba(0,0,0,0.3);
        `;
        
        document.body.appendChild(ownerLabel);
        
        var contentLabel = document.createElement('div');
        contentLabel.innerHTML = '<h2 style="background: #e8f4fd; padding: 10px; margin: 10px 0; border-left: 4px solid #2196F3;">查詢對象：{owner_name}</h2>';
        
        var firstTable = document.querySelector('table');
        var firstDiv = document.querySelector('div');
        var targetElement = firstTable || firstDiv || document.body;
        
        if (targetElement === document.body) {{
            targetElement.insertBefore(contentLabel, targetElement.firstChild);
        }} else {{
            targetElement.parentNode.insertBefore(contentLabel, targetElement);
        }}
        """
        
        driver.execute_script(js_code)
        print(f"已在頁面添加所有權人標識: {owner_name}")
        time.sleep(0.5)
        
    except Exception as e:
        print(f"添加所有權人標識失敗: {e}")

def collect_page_data_with_owner_label(driver, owner_name):
    """
    添加所有權人標識後收集頁面資料，保持原始PDF格式
    """
    try:
        add_owner_label_to_page(driver, owner_name)
        time.sleep(2)
        
        # 使用原本的PDF生成方式
        try:
            print("嘗試方法1：Chrome DevTools Protocol...")
            result = driver.execute_cdp_cmd("Page.printToPDF", {
                "landscape": False,
                "printBackground": True,
                "scale": 0.8,
                "paperWidth": 8.27,
                "paperHeight": 11.7,
                "marginTop": 0.4,
                "marginBottom": 0.4,
                "marginLeft": 0.4,
                "marginRight": 0.4
            })
            
            if 'data' in result:
                pdf_data = base64.b64decode(result['data'])
                print(f"成功獲取帶標識的PDF數據，大小: {len(pdf_data)} bytes")
                return pdf_data
            else:
                print("方法1 沒有返回數據")
                
        except Exception as e:
            print(f"方法1 失敗：{e}")
        
        try:
            print("嘗試方法2：截圖轉PDF...")
            total_height = driver.execute_script("return document.body.scrollHeight")
            driver.set_window_size(1200, total_height)
            time.sleep(0.5)
            
            screenshot = driver.get_screenshot_as_png()
            
            try:                
                image = Image.open(io.BytesIO(screenshot))
                print(f"成功獲取帶標識的截圖數據，大小: {len(screenshot)} bytes")
                return screenshot
                
            except ImportError:
                print(f"成功獲取帶標識的截圖數據，大小: {len(screenshot)} bytes")
                return screenshot
                
        except Exception as e:
            print(f"截圖獲取失敗: {e}")
            return None
            
    except Exception as e:
        print(f"頁面數據收集失敗: {e}")
        return None

def merge_pdfs_to_single_file(pdf_data_list, output_path, owner_names):
    """
    將多個PDF數據合併成一個PDF檔案
    """
    try:
        if not pdf_data_list:
            print("沒有PDF數據可合併")
            return False
        
        merger = PdfMerger()
        temp_files = []
        
        print(f"開始合併 {len(pdf_data_list)} 個PDF...")
        
        for i, (pdf_data, owner_name) in enumerate(zip(pdf_data_list, owner_names)):
            if pdf_data:
                temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
                temp_files.append(temp_file.name)
                
                if pdf_data.startswith(b'\x89PNG'):
                    try:
                        image = Image.open(io.BytesIO(pdf_data))
                        image.save(temp_file.name, "PDF", resolution=100.0)
                        print(f"第 {i+1} 筆({owner_name}): PNG轉PDF成功")
                    except Exception as e:
                        print(f"第 {i+1} 筆PNG轉PDF失敗: {e}")
                        continue
                else:
                    temp_file.write(pdf_data)
                    temp_file.close()
                    print(f"第 {i+1} 筆({owner_name}): PDF數據寫入成功")
                
                try:
                    merger.append(temp_file.name)
                except Exception as e:
                    print(f"第 {i+1} 筆PDF合併失敗: {e}")
        
        with open(output_path, 'wb') as output_file:
            merger.write(output_file)
        
        merger.close()
        
        for temp_file in temp_files:
            try:
                os.remove(temp_file)
            except:
                pass
        
        # print(f"PDF合併完成: {output_path}")  # 註解掉，避免重複顯示
        return True
        
    except Exception as e:
        print(f"PDF合併過程發生錯誤: {e}")
        return False

def fill_one_owner_with_vertical_excel_merged_pdf(driver, payload, owner, wait, outdir_root, excel_wb, excel_sheet):
    """
    修改版：收集PDF數據而不立即存檔
    """
    
    # 取用資料
    district = payload.get("district", "").strip()
    section = payload.get("section_hint", "").strip()
    lot8 = zfill8_digits(payload.get("lot", ""))
    lot_display = payload.get("lot_display", "")
    uid = owner.get("id_no", "").strip().upper()
    bday7 = owner.get("birthday", "").strip()
    name = owner.get("name") or "（未命名）"

    print(f"處理資料：{district}{section} 地號 {lot_display} - {name}")

    # 基本檢查
    if not district or not (uid or bday7):
        print("資料不足，跳過。")
        error_data = {'行政區': district, '地號': lot_display, '移轉記錄': []}
        add_vertical_rows_to_excel_final(excel_wb, excel_sheet, error_data, owner, '資料不足', district, section, '缺少行政區或身分資料')
        return None, False  # 返回 (pdf_data, success)

    # 執行查詢流程
    for attempt in range(3):
        try:
            print(f"第 {attempt + 1} 次嘗試...")

            try:
                driver.get(KAOHSIUNG_URL)
            except Exception as conn_error:
                error_msg = str(conn_error)
                if "ERR_CONNECTION_TIMED_OUT" in error_msg or "ERR_NAME_NOT_RESOLVED" in error_msg or "ERR_CONNECTION_REFUSED" in error_msg:
                    print("\033[91m" + "=" * 70, flush=True)
                    print("⚠️  高雄市土地增值稅網站連線失敗", flush=True)
                    print("=" * 70, flush=True)
                    print(f"網址: {KAOHSIUNG_URL}", flush=True)
                    print("\n可能原因:", flush=True)
                    print("  1. 網站伺服器暫時無法連線或維護中", flush=True)
                    print("  2. 網路連線不穩定", flush=True)
                    print("  3. 網站網址已變更", flush=True)
                    print("\n建議處理:", flush=True)
                    print("  1. 請稍後再試", flush=True)
                    print("  2. 檢查網路連線狀態", flush=True)
                    print("  3. 手動開啟瀏覽器確認網站是否正常運作", flush=True)
                    print("=" * 70 + "\033[0m", flush=True)
                    raise
                else:
                    raise
            WebDriverWait(driver, 12).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, SELECTORS["district"]))
            )
            time.sleep(0.5)

            success = select_district_and_section_with_memory(driver, district, section)
            if not success:
                if attempt < 2:
                    continue
                else:
                    error_data = {'行政區': district, '地號': lot_display, '移轉記錄': []}
                    add_vertical_rows_to_excel_final(excel_wb, excel_sheet, error_data, owner, '選區失敗', district, section, '無法選擇行政區或段小段')
                    return None, False

            ensure_form_filled(driver, lot8, uid, bday7)
            ok = solve_captcha_and_submit_fixed_v2(driver, district, section, lot8, uid, bday7, wait=None, max_try=3)
            
            if ok:
                print("查詢成功！收集PDF數據和Excel資料...")
                
                try:
                    # 收集PDF數據
                    # 收集帶所有權人標識的PDF數據
                    pdf_data = collect_page_data_with_owner_label(driver, name)
                    
                    # 解析Excel數據
                    page_text = driver.find_element(By.TAG_NAME, "body").text
                    query_data = parse_multiple_transfer_data_final(page_text, lot_display)
                    
                    # 記錄到Excel
                    add_vertical_rows_to_excel_final(excel_wb, excel_sheet, query_data, owner, '成功', district, section)
                    
                    print("資料收集完成")
                    return pdf_data, True
                    
                except Exception as e:
                    print(f"結果處理失敗: {e}")
                    traceback.print_exc()
                    error_data = {'行政區': district, '地號': lot_display, '移轉記錄': []}
                    add_vertical_rows_to_excel_final(excel_wb, excel_sheet, error_data, owner, '解析失敗', district, section, str(e))
                    return None, False
            else:
                if attempt < 2:
                    continue
                else:
                    error_data = {'行政區': district, '地號': lot_display, '移轉記錄': []}
                    add_vertical_rows_to_excel_final(excel_wb, excel_sheet, error_data, owner, '查詢失敗', district, section, '驗證碼處理失敗')
                    return None, False

        except Exception as e:
            error_msg = str(e)
            # 如果是連線錯誤，不顯示技術性錯誤訊息（因為前面已經顯示友善訊息了）
            is_connection_error = any(err in error_msg for err in ["ERR_CONNECTION_TIMED_OUT", "ERR_NAME_NOT_RESOLVED", "ERR_CONNECTION_REFUSED"])

            if not is_connection_error:
                print(f"第 {attempt + 1} 次嘗試發生錯誤：{e}", flush=True)

            if attempt >= 2:
                error_data = {'行政區': district, '地號': lot_display, '移轉記錄': []}
                add_vertical_rows_to_excel_final(excel_wb, excel_sheet, error_data, owner, '系統錯誤', district, section, str(e))
            continue

    return None, False

def run_with_complete_features(args):
    """
    完整功能的主程式：包含所有權人總計、PDF合併、智能地段選擇
    """
    global EXCEL_ROW_COUNTER, OWNER_TOTAL_TRACKER, SECTION_MAPPING_CACHE
    EXCEL_ROW_COUNTER = 0
    OWNER_TOTAL_TRACKER = {}
    SECTION_MAPPING_CACHE = {}

    # 🔥 在程式開始時檢查是否為高雄市
    try:
        with open(get_data_json_path(), 'r', encoding='utf-8') as file:
            data_list = json.load(file)
            if data_list and len(data_list) > 0:
                first_entry = data_list[0]
                city = first_entry.get('city', '')
                if city != '高雄市':
                    print(f"\033[96m目前查詢是「{city}」，非【高雄市】！\033[0m", flush=True)
                    print(f"\033[96m高雄市土地增值稅估算系統僅適用於高雄市。\033[0m", flush=True)
                    return
    except Exception as e:
        print(f"讀取 data.json 時發生錯誤: {e}", flush=True)

    json_path = args.json
    if not json_path:
        base = args.json_dir or "output"
        print(f"\n【JSON 來源】：自動搜尋目錄")
        print(f"  目錄路徑：{os.path.abspath(base)}")
        json_path = latest_json_in(base)
        if not json_path:
            print("\n❌ 錯誤：找不到 JSON 檔案")
            print(f"  請確認 {base} 資料夾中有 .json 檔案")
            return
        print(f"  找到最新檔案：{os.path.basename(json_path)}")
        print(f"  完整路徑：{json_path}")
    else:
        print(f"\n【JSON 來源】：指定檔案")
        print(f"  檔案名稱：{os.path.basename(json_path)}")
        print(f"  完整路徑：{json_path}")
        if not os.path.exists(json_path):
            print(f"\n❌ 錯誤：檔案不存在")
            print(f"  路徑：{json_path}")
            return
    
    print("="*60 + "\n")

    payloads, parse_logs = extract_payloads_with_validation(json_path)
    
    for line in parse_logs:
        print(line)

    if not payloads:
        print("\n沒有符合條件的資料可處理。")
        print("請確認：")
        print("1. 謄本類型是否為「土地謄本 - 第一類」或「土地謄本 - 第二類（法人）」")
        print("2. 行政區是否為高雄市轄區")
        return

    print(f"\n符合條件的標的筆數：{len(payloads)}")
    print("開始執行自動化查詢...")

    p0 = payloads[0]
    district0 = p0.get("district", "").strip()
    section0 = p0.get("section_hint", "").strip()
    lot80 = zfill8_digits(p0.get("lot", ""))

    # 🔥 優先用 data.json 第一筆當資料夾（避免謄本與 data.json 順序不同／過濾後導致檔案散落）
    # 規則：只要任一謄本地號在 data.json 中存在 → 用 data.json[0] 建資料夾
    # 完全無交集（不同地段）→ 退回謄本第一筆
    def _norm_lot(lot):
        s = str(lot or '').strip().replace('地號', '')
        try:
            if '-' in s:
                a, b = s.split('-', 1)
                return f"{int(a)}-{int(b)}"
            if len(s) >= 8:
                return f"{int(s[:4])}-{int(s[4:])}"
            return f"{int(s)}-0"
        except (ValueError, IndexError):
            return s

    folder_name = None
    try:
        data_json_path = get_data_json_path()
        if os.path.exists(data_json_path):
            with open(data_json_path, 'r', encoding='utf-8') as _f:
                _dl = json.load(_f)
            if _dl:
                data_set = {(
                    d.get('area', '').strip(),
                    d.get('section', '').strip(),
                    _norm_lot(d.get('lot_number', ''))
                ) for d in _dl}
                has_overlap = any(
                    (p.get('district', '').strip(),
                     p.get('section_hint', '').strip(),
                     _norm_lot(p.get('lot', '') or p.get('lot_number', ''))) in data_set
                    for p in payloads
                )
                if has_overlap:
                    first = _dl[0]
                    first_lot80 = zfill8_digits(first.get('lot_number', ''))
                    folder_name = label_for_lot_dir(
                        first.get('area', '').strip(),
                        first.get('section', '').strip(),
                        first_lot80
                    )
                    print(f"✓ 偵測到地號與 data.json 重疊，依 data.json 第一筆建立資料夾：{folder_name}", flush=True)
    except Exception as _e:
        print(f"[警告] 讀取 data.json 比對失敗：{_e}", flush=True)

    if not folder_name:
        folder_name = label_for_lot_dir(district0, section0, lot80)
        print(f"✓ 與 data.json 無交集，依謄本第一筆建立資料夾：{folder_name}", flush=True)

    # 🔥 使用 get_work_folder 確保在主程式目錄下建立資料夾
    outdir_root = get_work_folder(folder_name)

    transcript_dir = os.path.join(outdir_root, "1.基本資料")
    os.makedirs(transcript_dir, exist_ok=True)

    excel_path = os.path.join(transcript_dir, f"01-1.1_土地增值稅_高雄市版_{now_ymd()}.xlsx")
    excel_wb, excel_sheet = create_vertical_excel_workbook_final()

    show_gui = bool(args.show)
    driver = start_driver(show_gui=show_gui)
    
    pdf_data_list = []
    owner_names = []
    
    owner_record_counts = {}
    owner_processed_counts = {}
    
    for payload in payloads:
        owners = payload.get("owners") or []
        for owner in owners:
            owner_name = owner.get("name") or "未命名"
            owner_id = owner.get("id_no", "")
            owner_key = f"{owner_name}_{owner_id}"
            
            if owner_key not in owner_record_counts:
                owner_record_counts[owner_key] = 0
                owner_processed_counts[owner_key] = 0
            owner_record_counts[owner_key] += 1
    
    try:
        wait = WebDriverWait(driver, 12)
        success_count = 0
        total_count = 0

        for pi, payload in enumerate(payloads, 1):
            owners = payload.get("owners") or []
            lot_display = payload.get("lot_display", "")
            
            print(f"\n標的 {pi}/{len(payloads)}：{payload.get('district', '')}{payload.get('section_hint', '')} 地號 {lot_display}；所有權人 {len(owners)} 位")

            for oi, owner in enumerate(owners, 1):
                name = owner.get("name") or f"所有權人{oi}"
                owner_id = owner.get("id_no", "")
                owner_key = f"{name}_{owner_id}"
                total_count += 1
                print(f"第 {oi} 位：{name}")
                
                owner_processed_counts[owner_key] += 1
                is_last_record = (owner_processed_counts[owner_key] >= owner_record_counts[owner_key])
                
                try:
                    pdf_data, success = fill_one_owner_with_complete_features(driver, payload, owner, wait, outdir_root, excel_wb, excel_sheet, is_last_record)
                    
                    if success and pdf_data:
                        pdf_data_list.append(pdf_data)
                        owner_names.append(name)
                        success_count += 1
                        print(f"第 {oi} 位完成，PDF數據已收集")
                    else:
                        print(f"第 {oi} 位處理未完成")
                        
                except Exception as e:
                    print(f"第 {oi} 位處理發生例外：{e}")

        # 🔥 在儲存Excel前，檢查是否需要添加20年提示
        first_transfer_date = ''
        for owner_data in OWNER_TOTAL_TRACKER.values():
            date = owner_data.get('原地價年月', '')
            if date:
                first_transfer_date = date
                break

        if first_transfer_date:
            import re
            match = re.search(r'(\d+)年(\d+)月', first_transfer_date)
            if match:
                year_roc = int(match.group(1))
                month = int(match.group(2))
                year_ad = year_roc + 1911
                from datetime import datetime
                transfer_date = datetime(year_ad, month, 1)
                current_date = datetime.now()
                years_diff = (current_date - transfer_date).days / 365.25

                if years_diff > 20:
                    # 🔥 添加20年提示到Excel第二列
                    excel_sheet = excel_wb.active
                    excel_sheet.merge_cells('A2:O2')
                    warning_cell = excel_sheet['A2']
                    warning_cell.value = f"⚠️ 前次移轉日期距今超過20年，已達 {int(years_diff)} 年（原地價年月：{first_transfer_date}）建議以財政部估算之稅額為主"
                    warning_cell.font = Font(size=12, bold=True, color="FF0000")
                    warning_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    warning_cell.alignment = Alignment(horizontal='center', vertical='center')
                    excel_sheet.row_dimensions[2].height = 20

        save_excel_file(excel_wb, excel_path)

        # 將稅額結果寫入 JSON 供主程式使用
        try:
            # 建立 "4.其他相關" 資料夾
            other_dir = os.path.join(outdir_root, "4.其他相關")
            os.makedirs(other_dir, exist_ok=True)

            tax_result_file = os.path.join(other_dir, "高雄市_tax_result.json")

            # 計算總計
            total_once = sum(data['total_residential_tax'] for data in OWNER_TOTAL_TRACKER.values())
            total_general = sum(data['total_general_tax'] for data in OWNER_TOTAL_TRACKER.values())

            section_name = section0
            lot_number = p0.get("lot_display", "") or lot80[:4].lstrip("0")

            # 🔥 取得第一筆原地價年月（用於判斷是否超過20年）
            first_transfer_date = ''
            for owner_data in OWNER_TOTAL_TRACKER.values():
                date = owner_data.get('原地價年月', '')
                if date:
                    first_transfer_date = date
                    break

            tax_summary = {
                "自用住宅稅額": total_once,
                "一般土地稅額": total_general,
                "計算日期": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "地段": section_name,
                "地號": lot_number,
                "原地價年月": first_transfer_date,  # 🔥 加入原地價年月
                "詳細資料": []
            }

            # 加入每位所有權人的明細（與內政部格式保持一致）
            for owner_key, owner_data in OWNER_TOTAL_TRACKER.items():
                if owner_data['total_residential_tax'] > 0 or owner_data['total_general_tax'] > 0:
                    owner_detail = {
                        "所有權人": owner_data['owner_display_name'],
                        "統一編號": owner_data.get('owner_id', ''),
                        "自用稅額": owner_data['total_residential_tax'],
                        "一般稅額": owner_data['total_general_tax']
                    }
                    tax_summary["詳細資料"].append(owner_detail)

            with open(tax_result_file, 'w', encoding='utf-8') as f:
                json.dump(tax_summary, f, ensure_ascii=False, indent=2)

            print(f"\n✓ 稅額結果已儲存：{tax_result_file}")

            # 🔥 檢查是否超過20年並顯示提示
            if first_transfer_date:
                import re
                match = re.search(r'(\d+)年(\d+)月', first_transfer_date)
                if match:
                    year_roc = int(match.group(1))
                    month = int(match.group(2))
                    year_ad = year_roc + 1911
                    from datetime import datetime
                    transfer_date = datetime(year_ad, month, 1)
                    current_date = datetime.now()
                    years_diff = (current_date - transfer_date).days / 365.25

                    if years_diff > 20:
                        print(f"\033[93m" + "=" * 70, flush=True)
                        print(f"⚠️  前次移轉日期距今超過20年，已達 {int(years_diff)} 年（原地價年月：{first_transfer_date}）", flush=True)
                        print(f"⚠️  建議以財政部估算之稅額為主", flush=True)
                        print("=" * 70 + "\033[0m", flush=True)

        except Exception as e:
            print(f"儲存稅額結果失敗：{e}")

        if pdf_data_list:
            merged_pdf_path = os.path.join(transcript_dir, f"01-1_土地增值稅_高雄市版_{now_ymd()}.pdf")
            merge_success = merge_pdfs_to_single_file(pdf_data_list, merged_pdf_path, owner_names)
            
            if merge_success:
                print(f"PDF合併成功：{merged_pdf_path}")
            else:
                print("PDF合併失敗")
        else:
            print("沒有成功的PDF數據可合併")

        print(f"\n程式執行完成", flush=True)
        print(f"成功處理：{success_count}/{total_count}", flush=True)
        print(f"\033[93mExcel結果檔案：{excel_path}\033[0m", flush=True)  # 亮黃色
        if pdf_data_list:
            print(f"\033[93m合併PDF檔案：{merged_pdf_path}\033[0m", flush=True)  # 亮黃色

        # 🔥 最後再次顯示20年提示（如果需要）
        if first_transfer_date:
            import re
            match = re.search(r'(\d+)年(\d+)月', first_transfer_date)
            if match:
                year_roc = int(match.group(1))
                month = int(match.group(2))
                year_ad = year_roc + 1911
                from datetime import datetime
                transfer_date = datetime(year_ad, month, 1)
                current_date = datetime.now()
                years_diff = (current_date - transfer_date).days / 365.25

                if years_diff > 20:
                    print(f"", flush=True)
                    print(f"\033[93m" + "=" * 70, flush=True)
                    print(f"⚠️  重要提醒：前次移轉日期距今超過20年，已達 {int(years_diff)} 年（原地價年月：{first_transfer_date}）", flush=True)
                    print(f"⚠️  請注意：建議以財政部估算之稅額為主", flush=True)
                    print(f"⚠️  提示已標註在 Excel 及 PDF 中", flush=True)
                    print("=" * 70 + "\033[0m", flush=True)
        
    finally:
        try:
            driver.quit()
        except Exception:
            pass

#==============================================================================

def save_excel_file(wb, output_path):
    """
    儲存Excel檔案
    """
    try:
        wb.save(output_path)
        print(f"Excel檔案已儲存: {output_path}")
        return True
    except Exception as e:
        print(f"Excel儲存失敗: {e}")
        return False

def save_current_page_as_pdf_with_sequence(driver, outdir: str, district: str, section: str, lot8: str, owner: str) -> str:
    """
    使用序號化檔名的PDF存檔函數
    """
    try:
        # 建立"1.基本資料"子目錄
        transcript_dir = os.path.join(outdir, "1.基本資料")
        os.makedirs(transcript_dir, exist_ok=True)
        print(f"確保目錄存在：{transcript_dir}")
        
        # 生成序號化檔名
        filename_base = file_name_for_pdf_with_sequence(district, section, lot8, owner)
        safe_name = re.sub(r'[\\/:*?"<>|]+', "_", filename_base).strip("_")
        pdf_path = os.path.join(transcript_dir, f"{safe_name}.pdf")  # 改用 transcript_dir
        
        print(f"準備存檔至：{pdf_path}")
        
        # 等待頁面完全載入
        time.sleep(2)
        
        # 方法1：使用 Chrome DevTools Protocol
        try:
            print("嘗試方法1：Chrome DevTools Protocol...")
            result = driver.execute_cdp_cmd("Page.printToPDF", {
                "landscape": False,
                "printBackground": True,
                "scale": 0.8,
                "paperWidth": 8.27,
                "paperHeight": 11.7,
                "marginTop": 0.4,
                "marginBottom": 0.4,
                "marginLeft": 0.4,
                "marginRight": 0.4
            })
            
            if 'data' in result:
                pdf_data = base64.b64decode(result['data'])
                with open(pdf_path, 'wb') as f:
                    f.write(pdf_data)
                print(f"PDF 存檔成功：{pdf_path}")
                print(f"檔案大小：{len(pdf_data)} bytes")
                return pdf_path
            else:
                print("方法1 沒有返回數據")
                
        except Exception as e:
            print(f"方法1 失敗：{e}")
        
        # 方法2：截圖然後轉PDF
        try:
            print("嘗試方法2：截圖轉PDF...")
            
            total_height = driver.execute_script("return document.body.scrollHeight")
            driver.set_window_size(1024, total_height)
            time.sleep(0.5)
            
            screenshot = driver.get_screenshot_as_png()
            
            try:                
                image = Image.open(io.BytesIO(screenshot))
                image.save(pdf_path, "PDF", resolution=100.0)
                print(f"PDF 存檔成功（方法2）：{pdf_path}")
                return pdf_path
                
            except ImportError:
                png_path = pdf_path.replace('.pdf', '.png')
                with open(png_path, 'wb') as f:
                    f.write(screenshot)
                print(f"截圖存檔成功：{png_path}")
                return png_path
                
        except Exception as e:
            print(f"方法2 失敗：{e}")
        
        print("所有PDF存檔方法都失敗")
        return ""
        
    except Exception as e:
        print(f"PDF存檔過程發生錯誤：{e}")
        return ""

def fill_one_owner_with_sequence(driver, payload, owner, wait, outdir_root):
    """
    使用序號化檔名的所有權人處理流程
    """
    
    # 取用 payload / owner
    district = payload.get("district", "").strip()
    section = payload.get("section_hint", "").strip()
    lot8 = zfill8_digits(payload.get("lot", ""))
    lot_display = payload.get("lot_display", "")
    uid = owner.get("id_no", "").strip().upper()
    bday7 = owner.get("birthday", "").strip()
    name = owner.get("name") or "（未命名）"

    print(f"處理資料：")
    print(f"   行政區：{district}")
    print(f"   段小段：{section}")
    print(f"   地號：{lot_display} ({lot8})")
    print(f"   所有權人：{name}")

    # 基本檢查
    if not district or not (uid or bday7):
        print("資料不足（行政區 或 統編/生日欠缺），跳過。")
        return False

    # 最多重試3次完整流程
    for attempt in range(3):
        try:
            print(f"\n第 {attempt + 1} 次完整流程嘗試...")

            # 回到表單頁
            try:
                driver.get(KAOHSIUNG_URL)
            except Exception as conn_error:
                error_msg = str(conn_error)
                if "ERR_CONNECTION_TIMED_OUT" in error_msg or "ERR_NAME_NOT_RESOLVED" in error_msg or "ERR_CONNECTION_REFUSED" in error_msg:
                    print("\033[91m" + "=" * 70, flush=True)
                    print("⚠️  高雄市土地增值稅網站連線失敗", flush=True)
                    print("=" * 70, flush=True)
                    print(f"網址: {KAOHSIUNG_URL}", flush=True)
                    print("\n可能原因:", flush=True)
                    print("  1. 網站伺服器暫時無法連線或維護中", flush=True)
                    print("  2. 網路連線不穩定", flush=True)
                    print("  3. 網站網址已變更", flush=True)
                    print("\n建議處理:", flush=True)
                    print("  1. 請稍後再試", flush=True)
                    print("  2. 檢查網路連線狀態", flush=True)
                    print("  3. 手動開啟瀏覽器確認網站是否正常運作", flush=True)
                    print("=" * 70 + "\033[0m", flush=True)
                    raise
                else:
                    raise
            WebDriverWait(driver, 12).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, SELECTORS["district"]))
            )
            time.sleep(0.5)

            # 1) 選行政區 + 段小段
            print(f"選行政區/段小段...")
            try:
                select_district_and_section_with_memory(driver, district, section)
            except Exception as e:
                print(f"行政區/段選取失敗：{e}")
                if attempt < 2:  # 不是最後一次，繼續重試
                    continue
                else:
                    return False

            # 2) 填寫其他欄位
            print(f"填寫地號和所有權人資料...")
            ensure_form_filled(driver, lot8, uid, bday7)

            # 3) 驗證碼和提交
            print(f"處理驗證碼和提交...")
            ok = solve_captcha_and_submit_fixed_v2(driver, district, section, lot8, uid, bday7, wait=None, max_try=3)
            
            if ok:
                print("查詢成功！開始處理結果...")
                
                # 4) 解析結果頁面內容
                result_info = extract_result_page_info(driver)
                
                # 顯示結果摘要
                print("\n" + "="*50)
                print("查詢結果摘要")
                print("="*50)
                
                for key, value in result_info.items():
                    if isinstance(value, list):
                        print(f"{key}:")
                        for i, item in enumerate(value, 1):
                            print(f"  {i}. {item}")
                    else:
                        print(f"{key}: {value}")
                
                print("="*50)
                
                # 5) 使用序號化檔名儲存 PDF
                outdir = os.path.join(outdir_root)
                pdf_path = save_current_page_as_pdf_with_sequence(driver, outdir, district, section, lot8, name)
                
                if pdf_path:
                    print(f"處理完成，檔案儲存成功")
                else:
                    print("檔案儲存失敗，但查詢成功")
                
                return True
            
            elif not ok:
                print("驗證碼處理失敗，可能需要重新填寫表單")
                if attempt < 2:
                    print("準備重新開始完整流程...")
                    continue
                else:
                    print("多次嘗試後仍然失敗")
                    return False

        except Exception as e:
            error_msg = str(e)
            # 如果是連線錯誤，不顯示技術性錯誤訊息（因為前面已經顯示友善訊息了）
            is_connection_error = any(err in error_msg for err in ["ERR_CONNECTION_TIMED_OUT", "ERR_NAME_NOT_RESOLVED", "ERR_CONNECTION_REFUSED"])

            if not is_connection_error:
                print(f"第 {attempt + 1} 次嘗試發生錯誤：{e}", flush=True)

            if attempt < 2:
                if not is_connection_error:
                    print("準備重試...", flush=True)
                continue
            else:
                return False

    print("所有嘗試都失敗")
    return False

# ========= Main =========
# def run_with_validation(args):
#     """
#     加入驗證邏輯的主程式
#     """
#     # 找 JSON
#     json_path = args.json
#     if not json_path:
#         base = args.json_dir or "output"
#         if args.json_dir:
#             print(f"使用 JSON 目錄：{args.json_dir}（自動挑最新 .json）")

#         json_path = latest_json_in(base)
#         if not json_path:
#             print("找不到 JSON（請用 --json 指定檔案，或把檔案放在 output\\ 下）")
#             return
#         print(f"使用 JSON：{json_path}")
#     else:
#         print(f"指定 JSON：{json_path}")

#     # 使用帶驗證的解析函數
#     payloads, parse_logs = extract_payloads_with_validation(json_path)
    
#     for line in parse_logs:
#         print(line)

#     if not payloads:
#         print("\n沒有符合條件的資料可處理。")
#         print("請確認：")
#         print("1. 謄本類型是否為「土地謄本 - 第一類」")
#         print("2. 行政區是否為高雄市轄區")
#         return

#     print(f"\n符合條件的標的筆數：{len(payloads)}")


#     print("開始執行自動化查詢...")

#     # 根資料夾：以「第一筆」決定
#     p0 = payloads[0]
#     district0 = p0.get("district", "").strip()
#     section0 = p0.get("section_hint", "").strip()
#     lot80 = zfill8_digits(p0.get("lot", ""))
#     outdir_root = label_for_lot_dir(district0, section0, lot80)
#     os.makedirs(outdir_root, exist_ok=True)

#     # 開瀏覽器
#     show_gui = bool(args.show)
#     driver = start_driver(show_gui=show_gui)
    
#     try:
#         wait = WebDriverWait(driver, 12)

#         # 逐標的 × 逐所有權人
#         for pi, payload in enumerate(payloads, 1):
#             owners = payload.get("owners") or []
#             lot_display = payload.get("lot_display", "")
            
#             print(f"\n標的 {pi}/{len(payloads)}：{payload.get('district', '')}{payload.get('section_hint', '')} 地號 {lot_display}；所有權人 {len(owners)} 位")

#             for oi, owner in enumerate(owners, 1):
#                 name = owner.get("name") or f"所有權人{oi}"
#                 print(f"\n第 {oi} 位：{name}")
                
#                 try:
#                     success = fill_one_owner_with_sequence(driver, payload, owner, wait, outdir_root)
#                     if success:
#                         print(f"第 {oi} 位完成")
#                     else:
#                         print(f"第 {oi} 位處理未完成")
                        
#                 except UnexpectedAlertPresentException:
#                     msg = accept_any_alert(driver)
#                     if msg:
#                         print(f"警示：{msg}")
#                     print(f"第 {oi} 位處理遇到警示，未完成")
#                 except TimeoutException:
#                     print(f"第 {oi} 位處理逾時")
#                 except Exception as e:
#                     print(f"第 {oi} 位處理發生例外：{e}")

#         print("程式執行完成")
        
#     finally:
#         try:
#             driver.quit()
#         except Exception:
#             pass


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="高雄地政查詢 自動化")
    parser.add_argument("--json", help="指定 JSON 檔路徑", default=None)
    parser.add_argument("--json-dir", help="JSON 目錄", default=None)
    parser.add_argument("--show", action="store_true", help="顯示瀏覽器視窗")
    parser.add_argument("--base-dir", help="主程式目錄", default=None)  # 🔥 支援 BASE_DIR 架構
    args = parser.parse_args()
    
    # 直接使用垂直展開Excel輸出
    run_with_complete_features(args)
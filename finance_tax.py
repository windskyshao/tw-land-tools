# -*- coding: utf-8 -*-
"""

"""
import sys
import os

# 🔥 打包環境修正：將 _internal 加入 sys.path
if getattr(sys, 'frozen', False):
    # 如果是從打包的 EXE 執行
    internal_dir = os.path.dirname(os.path.abspath(__file__))
    if internal_dir not in sys.path:
        sys.path.insert(0, internal_dir)
elif '__file__' in globals():
    # 開發環境
    script_dir = os.path.dirname(os.path.abspath(__file__))
    if script_dir not in sys.path:
        sys.path.insert(0, script_dir)


import sys
print("=" * 78, flush=True)
print("===      財政部土地增值稅試算自動化      ===", flush=True)
print("=" * 78, flush=True)
print("模組載入中，請稍候......", flush=True)
sys.stdout.flush()
import os
import re
import sys
import json
import time
import argparse
import ctypes
import keyboard  # 用於系統級鍵盤事件
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import ssl
import urllib3
import traceback
from PyPDF2 import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfutils
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.fonts import addMapping
from PyPDF2 import PdfReader, PdfWriter
import io

# 禁用SSL警告
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_helper import verify_and_fix_chrome_window
import requests
from urllib.parse import urlparse

# 🔥 基準目錄設定（data.json 和工作資料夾的位置）
from base_dir_helper import BASE_DIR, get_data_json_path, get_work_folder

# 消費者物價總指數Excel檔案網址 - 財政部版本
CPI_EXCEL_URLS = [
    # "https://www.dgbas.gov.tw/public/data/dgbas03/bs3/inquire/cpispleym.xls",
    "https://ws.dgbas.gov.tw/001/Upload/463/relfile/10315/2552/cpispleym.xls",  # 舊網址（測試用）
]

# CPI設定檔路徑
CPI_CONFIG_FILE = "cpi_url_config.json"

# 財政部網址 - 修正為財政部網站
FINANCE_URLS = [
    "https://www.etax.nat.gov.tw/etwmain/etw158w/51",  # 財政部土地增值稅試算
]

# 表單選擇器 - 根據實際財政部HTML結構更新
SELECTORS = {
    # 基本資料輸入欄位（使用Angular的formcontrolname）
    "year_select": 'select#year[formcontrolname="year"]',  # 前次移轉年期
    "month_select": 'select#month[formcontrolname="month"]',  # 前次移轉月份
    "current_value": 'input#currVal[formcontrolname="currVal"]',  # 土地公告現值
    "previous_value": 'input#origVal[formcontrolname="origVal"]',  # 原規定地價或前次移轉現值
    "price_index": 'input#priceIdx[formcontrolname="priceIdx"]',  # 消費者物價總指數
    "land_area": 'input#area[formcontrolname="area"]',  # 土地宗地面積
    
    # 移轉權利範圍
    "numerator": 'input#numerator[formcontrolname="numerator"]',  # 分子
    "denominator": 'input#denominator[formcontrolname="denominator"]',  # 分母
    
    # 土地類型（單選按鈕）
    "land_type_urban": 'input#landType0[formcontrolname="landType"][value="0"]',  # 都市土地
    "land_type_rural": 'input#landType1[formcontrolname="landType"][value="1"]',   # 非都市土地
    
    # 按鈕
    "submit_btn": 'button[type="submit"][title="增值稅估算"]',  # 增值稅估算按鈕
    "clear_btn": 'button[type="button"][title="清除資料"]',  # 清除資料按鈕
}

# 🔥 自動偵測 DPI 縮放比例
def get_dpi_scale():
    """偵測系統 DPI 縮放比例"""
    try:
        hdc = ctypes.windll.user32.GetDC(0)
        dpi = ctypes.windll.gdi32.GetDeviceCaps(hdc, 88)  # LOGPIXELSX
        ctypes.windll.user32.ReleaseDC(0, hdc)
        dpi_scale = dpi / 96.0  # 96 DPI = 100%
        return dpi_scale
    except Exception:
        return 1.0

# 🔥 視窗大小設定：從 main.py 生成的設定檔讀取，自動填滿剩餘空間
finance_window_width = 1024
finance_window_height = 1024
finance_window_x = 0
finance_window_y = 0

try:
    config_path = os.path.join(BASE_DIR, 'window_config.json')
    if os.path.exists(config_path):
        with open(config_path, 'r', encoding='utf-8') as f:
            window_config = json.load(f)

        work_area = window_config.get('work_area', {})
        main_window = window_config.get('main_window', {})

        if work_area and main_window:
            # 取得 DPI 縮放比例
            dpi_scale = window_config.get('dpi_scale')
            if dpi_scale is None:
                dpi_scale = get_dpi_scale()

            # 取得主程式視窗資訊
            main_x = main_window.get('x', 1280)
            main_y = main_window.get('y', 0)
            main_w = main_window.get('width', 640)
            main_h = main_window.get('height', 1020)

            # 取得螢幕工作區域
            work_left = work_area.get('left', 0)
            work_top = work_area.get('top', 0)
            work_right = work_area.get('right', 1920)
            work_bottom = work_area.get('bottom', 1080)

            # 計算剩餘空間（根據主程式位置決定）
            screen_center_x = work_left + (work_right - work_left) // 2

            if main_x < screen_center_x:
                # 主程式在左側，Chrome 視窗放右側填滿剩餘空間
                finance_window_x_physical = main_x + main_w
                finance_window_width_physical = work_right - finance_window_x_physical
            else:
                # 主程式在右側，Chrome 視窗放左側填滿剩餘空間
                finance_window_x_physical = work_left
                finance_window_width_physical = main_x - work_left

            # 高度與主程式相同
            finance_window_y_physical = main_y
            finance_window_height_physical = main_h

            # 轉換為邏輯像素
            finance_window_width = int(finance_window_width_physical / dpi_scale)
            finance_window_height = int(finance_window_height_physical / dpi_scale) + 32  # +32 for titlebar
            finance_window_x = int(finance_window_x_physical / dpi_scale)
            finance_window_y = int(finance_window_y_physical / dpi_scale)

except Exception as e:
    # 使用預設值
    pass

# 全域變數控制訊息輸出等級
VERBOSE_MODE = False  # 設為 False 只顯示重要訊息

def flush_print(message, level="info", end='\n'):
    """簡化版輸出函數，支援訊息等級控制"""
    if level == "important" or VERBOSE_MODE:
        print(message, end=end, flush=True)
        sys.stdout.flush()
        time.sleep(0.05)  # 縮短等待時間

def merge_and_clean_pdfs(transcript_dir, section_name, lot_number):
    """合併所有帶標題的PDF並清理重複檔案"""
    try:
        import glob
        
        flush_print("\n=== 開始PDF合併與清理作業 ===")
        
        # 找出所有帶標題的PDF檔案
        pattern = os.path.join(transcript_dir, "*_with_header.pdf")
        header_files = glob.glob(pattern)
        
        if not header_files:
            flush_print("沒有找到帶標題的PDF檔案")
            return
        
        # 按編號排序
        def extract_number(filename):
            match = re.search(r'result_(\d+)_', os.path.basename(filename))
            return int(match.group(1)) if match else 0
        
        header_files.sort(key=extract_number)
        flush_print(f"找到 {len(header_files)} 個帶標題的PDF檔案")
        
        # 合併PDF
        writer = PdfWriter()
        for pdf_file in header_files:
            try:
                reader = PdfReader(pdf_file)
                for page in reader.pages:
                    writer.add_page(page)
                flush_print(f"已合併: {os.path.basename(pdf_file)}")
            except Exception as e:
                flush_print(f"合併失敗: {os.path.basename(pdf_file)} - {e}")
        
        # 保存合併後的PDF
        merged_filename = f"01_土地增值稅_財政部版_{section_name}_{lot_number}_{now_ymd()}.pdf"
        merged_path = os.path.join(transcript_dir, merged_filename)
        
        with open(merged_path, 'wb') as output_file:
            writer.write(output_file)
        
        flush_print(f"✓ PDF合併完成: {merged_filename}")
        
        # 刪除所有個別PDF檔案（包括帶標題和不帶標題的）
        pattern_all = os.path.join(transcript_dir, "result_*.pdf")
        all_result_files = glob.glob(pattern_all)
        
        deleted_count = 0
        for file_path in all_result_files:
            try:
                os.remove(file_path)
                deleted_count += 1
                flush_print(f"已刪除: {os.path.basename(file_path)}")
            except Exception as e:
                flush_print(f"刪除失敗: {os.path.basename(file_path)} - {e}")

        flush_print(f"✓ 清理完成，共刪除 {deleted_count} 個個別PDF檔案")
        flush_print(f"\n\033[93m✓ 最終保留檔案：{merged_path} 和 Excel統計表\033[0m", "important")
        
    except Exception as e:
        flush_print(f"PDF合併清理過程發生錯誤: {e}")


def safe_input(prompt):
    """安全的輸入函數，確保提示訊息能顯示"""
    flush_print(prompt, end='')
    sys.stdout.flush()
    time.sleep(0.2)
    return input()

def now_ymd():
    return datetime.now().strftime("%Y%m%d")

def now_hms():
    return datetime.now().strftime("%H%M%S")

def load_user_cpi_urls():
    """讀取用戶保存的 CPI 網址"""
    try:
        if os.path.exists(CPI_CONFIG_FILE):
            with open(CPI_CONFIG_FILE, 'r', encoding='utf-8') as f:
                config = json.load(f)
                urls = config.get('user_urls', [])
                if urls:
                    flush_print(f"✓ 已載入 {len(urls)} 個使用者提供的網址", "important")
                return urls
    except Exception as e:
        flush_print(f"讀取CPI設定檔失敗：{e}")
    return []

def save_user_cpi_url(url):
    """保存用戶提供的新 CPI 網址，並記錄版本號以供推理"""
    try:
        # 讀取現有配置
        config = {}
        if os.path.exists(CPI_CONFIG_FILE):
            with open(CPI_CONFIG_FILE, 'r', encoding='utf-8') as f:
                config = json.load(f)

        # 取得現有網址列表和歷史記錄
        user_urls = config.get('user_urls', [])
        history = config.get('history', [])

        # 從網址中提取版本號（如果有的話）
        import re
        version_match = re.search(r'/(\d{4})/cpispleym', url)
        version_num = version_match.group(1) if version_match else "未知"

        # 如果網址不在列表中，加到最前面
        if url not in user_urls:
            user_urls.insert(0, url)
            # 只保留最近5個網址
            user_urls = user_urls[:5]

            # 記錄歷史（供推理用）
            history.append({
                "date": now_ymd(),
                "url": url,
                "version": version_num
            })

        # 更新配置
        config['user_urls'] = user_urls
        config['history'] = history
        config['last_updated'] = now_ymd()

        # 保存配置
        with open(CPI_CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)

        flush_print(f"✓ 已保存新網址到設定檔（版本：{version_num}）", "important")
        return True
    except Exception as e:
        flush_print(f"保存CPI設定檔失敗：{e}", "important")
        return False

def save_success_log(url):
    """記錄成功下載的網址到日誌檔"""
    try:
        with open("cpi_download_log.txt", "a", encoding="utf-8") as log:
            log.write(f"{now_ymd()},{url},成功\n")
    except:
        pass

def ask_user_for_cpi_url(session, headers):
    """詢問使用者提供新的 CPI 下載網址"""
    flush_print("\n" + "="*39, "important")
    flush_print("📋 請提供新的 CPI 檔案下載網址", "important")
    flush_print("="*39, "important")

    flush_print("\n\033[96m✓請用瀏覽器開啟：\033[0m", "important")
    flush_print("\033[96m✓  https://www.etax.nat.gov.tw/etwmain/etw158w/51\033[0m", "important")

    flush_print("\n\033[96m✓找到「各年月為基期之消費者物價總指數－稅務專用」\033[0m", "important")
    flush_print("\033[96m✓在 EXCEL 連結上按右鍵 → 複製連結網址\033[0m", "important")

    flush_print("\n請貼上完整的下載網址（或按 Enter 跳過）：\033[0m", "important")
    flush_print("\033[96m✓範例: https://ws.dgbas.gov.tw/001/Upload/463/relfile/10315/2570/cpispleym.xls\n\033[0m", "important")
                            
    user_url = input("網址: ").strip()

    if not user_url:
        flush_print("已跳過", "important")
        return None

    # 驗證網址格式
    if not user_url.startswith('http'):
        flush_print("✗ 無效的網址格式", "important")
        return None

    # 測試網址是否有效
    flush_print(f"\n正在驗證網址...", "important")
    try:
        response = session.get(user_url, headers=headers, timeout=15)
        if response.status_code == 200 and len(response.content) > 10000:
            # 下載成功
            filename = f"cpispleym_{now_ymd()}.xls"
            with open(filename, 'wb') as f:
                f.write(response.content)

            flush_print(f"✓ 網址有效！檔案下載成功", "important")
            flush_print(f"  檔案：{filename}", "important")
            flush_print(f"  大小：{len(response.content):,} bytes", "important")

            # 保存這個網址
            save_user_cpi_url(user_url)

            return filename
        else:
            flush_print(f"✗ 網址無效：HTTP {response.status_code} 或檔案太小", "important")
            return None
    except Exception as e:
        flush_print(f"✗ 網址測試失敗：{e}", "important")
        return None

def download_cpi_excel():
    """下載CPI檔案 - 優先爬蟲取得最新網址，確保 CPI 資料是最新的"""
    try:
        flush_print("下載最新CPI資料...", "important")

        session = requests.Session()
        session.verify = False

        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7',
            'Connection': 'keep-alive',
        }

        # 🔥 方法1：優先爬蟲主計總處統計處網站（官方 CPI 來源）
        import re
        try:
            flush_print("正在從主計總處網站取得最新 CPI 連結...", "important")
            main_url = "https://www.stat.gov.tw/cp.aspx?n=2665"
            
            response = session.get(main_url, headers=headers, timeout=30)
            
            if response.status_code == 200:
                html_content = response.text
                flush_print(f"✓ 網頁載入成功（{len(html_content)} bytes）", "important")

                # 尋找包含「稅務專用」和「cpispleym.xls」的連結
                # 模式：href="(https://[^"]*cpispleym\.xls)"
                pattern = r'href="(https://[^"]*cpispleym\.xls)"'
                matches = re.findall(pattern, html_content)
                
                if matches:
                    # 找到多個連結，選擇第一個（通常是最新的）
                    download_url = matches[0]
                    flush_print(f"✓ 找到下載連結：{download_url}", "important")
                    
                    # 🔥 下載檔案
                    flush_print("開始下載...", "important")
                    file_response = session.get(download_url, headers=headers, timeout=30)
                    
                    if file_response.status_code == 200 and len(file_response.content) > 10000:
                        filename = f"cpispleym_{now_ymd()}.xls"
                        with open(filename, 'wb') as f:
                            f.write(file_response.content)
                        
                        flush_print(f"✓ 下載成功！", "important")
                        flush_print(f"  檔案：{filename}", "important")
                        flush_print(f"  大小：{len(file_response.content):,} bytes", "important")
                        
                        # 記錄成功
                        save_success_log(download_url)
                        return filename
                    else:
                        flush_print(f"✗ 下載失敗：檔案太小或連結無效", "important")
                else:
                    flush_print("✗ 網頁中未找到 cpispleym.xls 連結", "important")
                    
                    # 🔥 備用：試試 ODS 格式
                    flush_print("嘗試尋找 ODS 格式...", "important")
                    pattern_ods = r'href="(https://[^"]*cpispleym\.ods)"'
                    matches_ods = re.findall(pattern_ods, html_content)
                    
                    if matches_ods:
                        download_url = matches_ods[0]
                        flush_print(f"✓ 找到 ODS 連結：{download_url}", "important")
                        
                        file_response = session.get(download_url, headers=headers, timeout=30)
                        if file_response.status_code == 200 and len(file_response.content) > 10000:
                            filename = f"cpispleym_{now_ymd()}.ods"
                            with open(filename, 'wb') as f:
                                f.write(file_response.content)
                            flush_print(f"✓ ODS 下載成功：{filename}", "important")
                            return filename
            else:
                flush_print(f"✗ 網頁訪問失敗：HTTP {response.status_code}", "important")
                
        except Exception as e:
            flush_print(f"✗ 爬蟲失敗：{e}", "important")

        # 🔥 方法2：爬蟲失敗時，使用用戶之前保存的網址（備用方案）
        user_urls = load_user_cpi_urls()
        if user_urls:
            latest_url = user_urls[0]  # 只試最新的
            flush_print(f"嘗試使用之前保存的網址...", "important")
            try:
                response = session.get(latest_url, headers=headers, timeout=15)
                if response.status_code == 200 and len(response.content) > 10000:
                    filename = f"cpispleym_{now_ymd()}.xls"
                    with open(filename, 'wb') as f:
                        f.write(response.content)
                    flush_print(f"✓ 下載成功！（注意：可能非最新版本）", "important")
                    save_success_log(latest_url)
                    return filename
                else:
                    flush_print(f"  ✗ 失敗：該網址可能已過期", "important")
            except Exception as e:
                flush_print(f"  ✗ 失敗：{e}", "important")

        # 🔥 方法3：請用戶提供新網址（必須是最新資料）
        flush_print("\n" + "="*39, "important")
        flush_print("\033[96m⚠️  所有自動下載方式都失敗了\033[0m", "important")

        flush_print("="*39, "important")

        result = ask_user_for_cpi_url(session, headers)
        if result:
            return result

        # 如果用戶沒有提供網址，無法繼續（CPI 必須是最新資料）
        flush_print("\n" + "="*39, "important")
        flush_print("✗ 無法取得最新 CPI 資料", "important")
        flush_print("="*39, "important")
        flush_print("由於 CPI 資料必須使用最新版本，程式無法繼續執行。", "important")
        flush_print("請稍後重新執行程式，或聯繫管理員更新 CPI 下載機制。", "important")
        return None
        
    except Exception as e:
        flush_print(f"下載過程錯誤：{e}", "important")
        return create_default_cpi_data()

def create_default_cpi_data():
    """創建預設的CPI數據檔案"""
    try:
        flush_print("正在創建預設CPI數據...")
        
        # 創建基本的CPI數據
        default_cpi = {
            "110年01月": "100.50",
            "110年02月": "100.75",
            "110年03月": "101.00",
            "110年04月": "101.25",
            "110年05月": "101.50",
            "110年06月": "101.75",
            "110年07月": "102.00",
            "110年08月": "102.25",
            "110年09月": "102.50",
            "110年10月": "102.75",
            "110年11月": "103.00",
            "110年12月": "103.25",
            "111年01月": "103.50",
            "111年02月": "103.75",
            "111年03月": "104.00",
            "111年04月": "104.25",
            "111年05月": "104.50",
            "111年06月": "104.75",
            "111年07月": "105.00",
            "111年08月": "105.25",
            "111年09月": "105.50",
            "111年10月": "105.75",
            "111年11月": "106.00",
            "111年12月": "106.25",
        }
        
        # 將預設數據保存為JSON檔案
        filename = f"default_cpi_{now_ymd()}.json"
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(default_cpi, f, ensure_ascii=False, indent=2)
        
        flush_print(f"✓ 預設CPI數據已創建：{filename}")
        return filename
        
    except Exception as e:
        flush_print(f"創建預設CPI數據失敗：{e}")
        return None

def parse_cpi_data(file_path):
    """解析CPI數據 - 修正版，對應實際稅務專用工作表格式"""
    try:
        flush_print("正在解析消費者物價總指數資料...", "important")
        
        if file_path.endswith('.json'):
            with open(file_path, 'r', encoding='utf-8') as f:
                cpi_dict = json.load(f)
            flush_print(f"從JSON檔案載入 {len(cpi_dict)} 筆CPI資料", "important")
            return cpi_dict
            
        elif file_path.endswith(('.xls', '.xlsx')):
            try:
                # 讀取稅務專用工作表，不設定header，直接讀取原始資料
                df = pd.read_excel(file_path, sheet_name="稅務專用", header=None)
            except:
                df = pd.read_excel(file_path, sheet_name=0, header=None)
            
            cpi_dict = {}
            
            # 從第5行開始讀取（index=4），跳過標題
            for index in range(4, len(df)):
                row = df.iloc[index]
                
                # 檢查第一欄是否為年份
                if pd.isna(row.iloc[0]) or not isinstance(row.iloc[0], (int, float)):
                    continue
                
                try:
                    year_num = int(row.iloc[0])
                    # 確保是合理的民國年份
                    if year_num < 48 or year_num > 120:
                        continue
                except (ValueError, TypeError):
                    continue
                
                # 讀取12個月的資料（第2-13欄，index 1-12）
                for month in range(1, 13):
                    col_index = month  # 第2欄=1月(index=1), 第3欄=2月(index=2)...
                    
                    if col_index < len(row) and pd.notna(row.iloc[col_index]):
                        try:
                            cpi_value = float(row.iloc[col_index])
                            
                            # 建立多種查詢格式的key
                            key1 = f"{year_num:03d}年{month:02d}月"  # 089年01月
                            key2 = f"{year_num:03d}{month:02d}"      # 08901
                            key3 = f"{year_num:02d}{month:02d}"      # 8901 (年份<100時)
                            
                            cpi_dict[key1] = str(cpi_value)
                            cpi_dict[key2] = str(cpi_value)
                            if year_num < 100:
                                cpi_dict[key3] = str(cpi_value)
                            
                            # 驗證關鍵年份
                            if year_num == 89 and month == 1:
                                flush_print(f"驗證民國89年1月CPI: {cpi_value}", "important")
                                
                        except (ValueError, TypeError):
                            continue
            
            if cpi_dict:
                flush_print(f"CPI解析完成，共載入 {len(cpi_dict)} 筆資料", "important")
                return cpi_dict
            else:
                flush_print("Excel檔案中未找到有效的CPI數據", "important")
                return create_backup_cpi_dict()
            
    except Exception as e:
        flush_print(f"解析CPI檔案失敗：{e}", "important")
        return create_backup_cpi_dict()

def create_backup_cpi_dict():
    """創建備用的CPI數據字典"""
    flush_print("正在創建備用CPI數據...")
    
    # 基於財政部實際CPI數據的備用字典
    backup_cpi = {
        "087年01月": "137.5", "087年02月": "137.8", "087年03月": "137.3", 
        "087年04月": "136.8", "087年05月": "137.2", "087年06月": "135.3",
        "087年07月": "136.2", "087年08月": "136.9", "087年09月": "136.0",
        "087年10月": "135.0", "087年11月": "133.6", "087年12月": "135.8",
        "088年01月": "135.2", "088年02月": "136.1", "088年03月": "135.8",
        # 更多年份的數據...
        "108年01月": "104.64", "108年02月": "104.69", "108年03月": "104.90",
        "109年01月": "105.30", "109年02月": "105.31", "109年03月": "104.79",
        "110年01月": "103.85", "110年02月": "104.30", "110年03月": "104.62",
        "111年01月": "107.88", "111年02月": "108.22", "111年03月": "108.80",
        "112年01月": "111.52", "112年02月": "110.95", "112年03月": "110.49",
        "113年01月": "109.35", "113年02月": "109.12", "113年03月": "108.95",
        "114年01月": "108.75", "114年02月": "108.60", "114年03月": "108.45",
    }
    
    flush_print(f"✓ 備用CPI數據創建完成，共 {len(backup_cpi)} 筆數據")
    return backup_cpi

def start_driver(show_gui: bool = True):
    """啟動瀏覽器 - 財政部版本"""
    from selenium.webdriver.chrome.options import Options

    flush_print("正在啟動Chrome瀏覽器...")
    sys.stdout.flush()

    service = Service(ChromeDriverManager().install())
    opts = Options()

    if not show_gui:
        opts.add_argument("--headless")
        flush_print("使用無頭模式 - 瀏覽器在背景執行")
    else:
        flush_print("使用GUI模式 - 瀏覽器視窗將會顯示")

    # 加強防檢測設定
    opts.add_argument("--disable-gpu")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-extensions")
    opts.add_argument("--disable-plugins")
    opts.add_argument("--disable-images")
    # 🔥 不在啟動參數設定視窗大小，改為啟動後再調整（避免 DPI 改變時崩潰）
    # opts.add_argument("--window-position=0,0")
    # opts.add_argument("--window-size=1024,900")

    # 財政部網站適用的User-Agent
    opts.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")

    # 忽略SSL錯誤
    opts.add_argument("--ignore-ssl-errors")
    opts.add_argument("--ignore-certificate-errors")
    opts.add_argument("--allow-running-insecure-content")

    # 防檢測設定
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option('useAutomationExtension', False)

    # 🔥 強制使用淺色模式（避免暗黑模式影響截圖品質）
    opts.add_argument('--disable-features=WebUIDarkMode,ForcedColors')
    opts.add_argument('--force-color-profile=srgb')
    opts.add_argument('--blink-settings=preferredColorScheme=1')  # 1=light

    # 設定 Chrome 偏好，強制淺色主題
    prefs = {
        'profile.default_content_setting_values.prefers_color_scheme': 1,  # 1=light, 2=dark
        'profile.managed_default_content_settings.images': 1,
        'browser.theme.color_scheme': 'light'
    }
    opts.add_experimental_option('prefs', prefs)

    driver = webdriver.Chrome(service=service, options=opts)

    # 🔥 啟動後立即設定視窗大小和位置
    try:
        driver.set_window_size(finance_window_width, finance_window_height)
        driver.set_window_position(finance_window_x, finance_window_y)
    except Exception as e:
        print(f"[WARNING] 視窗設定失敗: {e}，繼續執行")

    # 移除webdriver屬性，避免被檢測
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

    driver.set_page_load_timeout(60)
    flush_print("Chrome瀏覽器已啟動")
    sys.stdout.flush()

    return driver

def try_open_finance_website(driver):
    """嘗試開啟財政部土地增值稅試算網站"""
    try:
        flush_print("正在嘗試訪問財政部土地增值稅試算網站...")

        zoom_applied = False  # 🔥 標記是否已經執行過縮放，避免重複縮放

        for i, url in enumerate(FINANCE_URLS, 1):
            try:
                flush_print(f"嘗試網址 {i}/{len(FINANCE_URLS)}: {url}")
                
                driver.get(url)

                # 網頁載入後重新設定視窗大小（避免被網站重置）
                try:
                    driver.set_window_size(finance_window_width, finance_window_height)
                    driver.set_window_position(finance_window_x, finance_window_y)
                except:
                    pass

                # 🔥 DPI 自動修正（只在 DPI 不同步時動作，正常 PC 無影響）
                verify_and_fix_chrome_window(driver)

                # 🔥 自動判斷是否需要縮放頁面（解決高 DPI 或小螢幕下的排版問題）
                # 根據 DPI 決定縮放次數：150% → 2次(80%)，175% → 4次(67%)
                # 🔥 只在第一次嘗試時執行縮放，避免重試時重複縮放
                if not zoom_applied:
                    zoom_count = 0
                    try:
                        config_path = os.path.join(BASE_DIR, 'window_config.json')
                        if os.path.exists(config_path):
                            with open(config_path, 'r', encoding='utf-8') as f:
                                config = json.load(f)

                            current_dpi = config.get('dpi_scale', 1.0)
                            screen_width = config.get('screen_width', 1920)
                            logical_width = screen_width / current_dpi

                            # 🔥 根據 DPI 和螢幕尺寸決定縮放次數
                            if current_dpi >= 1.75:
                                zoom_count = 4  # 175% 或更高：縮放到 67%（按 4 次）
                                target_zoom = "67%"
                            elif current_dpi >= 1.5:
                                zoom_count = 2  # 150%：縮放到 80%
                                target_zoom = "80%"
                            elif current_dpi >= 1.25 and logical_width < 1200:
                                zoom_count = 2  # 小螢幕 + 125%：縮放到 80%
                                target_zoom = "80%"
                            else:
                                zoom_count = 0  # 100%、大螢幕125%：不縮放
                                target_zoom = "100%"

                            if zoom_count > 0:
                                flush_print(f"[頁面縮放] 偵測到 DPI={current_dpi}x (即 {int(current_dpi * 100)}%), 邏輯寬度={logical_width:.0f}px → 需要縮放至 {target_zoom}")
                            else:
                                flush_print(f"[頁面縮放] 偵測到 DPI={current_dpi}x (即 {int(current_dpi * 100)}%), 邏輯寬度={logical_width:.0f}px → 不需要縮放")
                    except Exception as e:
                        flush_print(f"[頁面縮放] 偵測配置時發生錯誤: {e}，使用預設不縮放")

                    if zoom_count > 0:
                        try:
                            # 🔥 使用 keyboard 庫發送系統級的 Ctrl + - 按鍵（真正的瀏覽器縮放）
                            # 先點擊頁面確保 Chrome 視窗有焦點
                            body = driver.find_element(By.TAG_NAME, 'body')
                            body.click()
                            time.sleep(0.5)

                            # 🔥 根據 DPI 決定按幾次 Ctrl + -
                            flush_print(f"[頁面縮放] 正在使用系統鍵盤縮放（按 {zoom_count} 次）...")
                            for i in range(zoom_count):
                                keyboard.press_and_release('ctrl+-')  # 系統級按鍵
                                time.sleep(0.5)

                            flush_print(f"[頁面縮放] ✓ 已使用系統鍵盤縮放至 {target_zoom}")
                        except Exception as e:
                            flush_print(f"[頁面縮放] 設定縮放時發生錯誤: {e}")
                    else:
                        flush_print("[頁面縮放] 螢幕配置正常，不需要縮放")

                    zoom_applied = True  # 🔥 標記已執行過縮放

                time.sleep(0.5)

                page_title = driver.title
                page_source = driver.page_source.lower()
                
                flush_print(f"  頁面標題：{page_title}")
                
                # 檢查是否成功進入財政部試算頁面
                if any(keyword in page_source for keyword in [
                    "土地增值稅", "試算", "計算", "前次移轉", "公告現值", 
                    "input", "select", "form"
                ]):
                    flush_print("✓ 成功進入財政部土地增值稅試算頁面")
                    return True
                else:
                    flush_print(f"  網址 {i} 未找到試算相關內容...")
                    continue
                        
            except Exception as e:
                flush_print(f"  網址 {i} 訪問失敗：{e}")
                continue
        
        flush_print("✗ 無法自動開啟財政部網站")
        flush_print("請手動開啟 https://www.etax.nat.gov.tw/etwmain/etw158w/51")
        input("\n手動開啟試算頁面後，請按Enter鍵繼續...")
        return True
        
    except Exception as e:
        flush_print(f"訪問試算網站時發生錯誤：{e}")
        return False

def wait_for_form_elements(driver, timeout=30):
    """等待表單元素載入 - 財政部版本"""
    try:
        flush_print("等待表單元素載入...")
        
        # 等待頁面基本載入
        time.sleep(0.2)
        
        # 嘗試等待多種可能的表單元素
        selectors_to_wait = [
            "input[type='text']",
            "input[type='number']", 
            "input[name*='area']",
            "input[name*='value']",
            "input[name*='date']",
            "input[type='submit']",
            "button[type='submit']",
        ]
        
        elements_found = []
        
        for selector in selectors_to_wait:
            try:
                elements = driver.find_elements(By.CSS_SELECTOR, selector)
                if elements:
                    elements_found.append(f"{selector} ({len(elements)}個)")
                    flush_print(f"✓ 找到元素: {selector} - {len(elements)}個")
            except Exception as e:
                continue
        
        if elements_found:
            flush_print(f"✓ 表單元素載入完成，找到: {', '.join(elements_found)}")
            time.sleep(0.5)
            return True
        else:
            flush_print("⚠ 未找到預期的表單元素，但繼續執行...")
            return True
        
    except Exception as e:
        flush_print(f"等待表單元素時發生錯誤: {e}")
        return True

def latest_json_in(folder: str):
    """找到資料夾中最新的JSON檔案"""
    if not os.path.isdir(folder):
        return None
    cands = []
    for fn in os.listdir(folder):
        if not fn.lower().endswith(".json"): 
            continue
        p = os.path.join(folder, fn)
        try:
            cands.append((os.path.getmtime(p), p))
        except Exception:
            pass
    if not cands:
        return None
    return sorted(cands)[-1][1]

def validate_transcript_type(block):
    """驗證謄本類型：接受土地謄本（一類 / 二類 / 電傳）"""
    transcript_type = (block or {}).get("謄本類型", "").strip()

    if ("土地謄本" in transcript_type and
        ("第一類" in transcript_type or "第二類" in transcript_type or "電傳" in transcript_type)):
        return True, transcript_type

    return False, transcript_type

def parse_geo_from_title(title: str):
    """解析地理資訊

    支援格式舉例（\\w+[縣市] 為可選前綴）：
      - 高雄市鳳山區過埤段 0012-0000      → district=鳳山區
      - 屏東縣枋山鄉莿桐段 0157-0000      → district=枋山鄉
      - 屏東縣屏東市新和段 0094          → district=屏東市（縣轄市）
      - 鳳山區過埤段 0012                 → district=鳳山區（無縣市前綴）
    """
    title = (title or "").strip()
    district = ""
    section = ""
    lot_front = ""
    lot_tail = ""

    # 可選的 縣/直轄市 前綴，接著 1~3 字 + (鄉|鎮|市|區) 為 district
    pattern = r"(?:[^\s]+?[縣市])?(?P<district>[^\s]{1,3}(?:鄉|鎮|市|區))(?P<section>.+?段(?:.+?小段)?)\s*(?P<lot_front>\d{3,4})(?:-(?P<lot_tail>\d{4}))?地?號?"
    match = re.search(pattern, title)
    
    if match:
        district = match.group("district")
        section = match.group("section")
        lot_front = match.group("lot_front")
        lot_tail = match.group("lot_tail") or "0000"
        
        return {
            "district": district,
            "section_hint": section,
            "lot_front": lot_front,
            "lot_tail": lot_tail,
            "lot_number": (lot_front + lot_tail).zfill(8)[:8],
            "area": ""
        }
    
    return {"district": "", "section_hint": "", "lot_front": "", "lot_tail": "", "lot_number": "", "area": ""}

def extract_year_month_from_ownership(ownership_data):
    """從所有權部擷取所有的年月資料"""
    year_months = []
    
    if isinstance(ownership_data, list):
        for owner in ownership_data:
            if isinstance(owner, dict):
                for key, value in owner.items():
                    if "年月" in key and value:
                        year_months.append(value.strip())
    
    return year_months

def get_cpi_for_year_month(year_month_str, cpi_dict):
    """根據年月字串查找對應的CPI值 - 財政部版本"""
    # 標準化年月格式
    year_month_str = year_month_str.strip()
    
    # 如果直接匹配
    if year_month_str in cpi_dict:
        return cpi_dict[year_month_str]
    
    # 嘗試解析不同格式
    import re
    
    # 匹配格式如：098年06月（民國年）
    match = re.search(r'(\d{3})年(\d{1,2})月', year_month_str)
    if match:
        year = match.group(1)
        month = int(match.group(2))
        
        # 嘗試多種可能的格式
        possible_keys = [
            f"{year}年{month:02d}月",  # 098年06月
            f"{int(year):03d}年{month:02d}月",  # 確保3位數年份
        ]
        
        for key in possible_keys:
            if key in cpi_dict:
                return cpi_dict[key]
    
    # 如果都找不到，顯示可用的年份範圍供調試
    available_years = set()
    for key in list(cpi_dict.keys())[:10]:
        year_match = re.search(r'(\d{3})年', key)
        if year_match:
            available_years.add(year_match.group(1))
    
    if available_years:
        available_years_list = sorted(list(available_years))[:5]
        flush_print(f"  可用的CPI年份範例: {available_years_list}")
    
    return None

def extract_payloads_from_json(json_path: str, cpi_dict=None):
    """從JSON擷取資料，包含CPI查詢 - 財政部版本"""
    with open(json_path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    blocks = raw if isinstance(raw, list) else [raw]
    payloads = []

    # 🔥 自動偵測 DPI 縮放比例
    def get_dpi_scale():
        """偵測系統 DPI 縮放比例"""
        try:
            hdc = ctypes.windll.user32.GetDC(0)
            dpi = ctypes.windll.gdi32.GetDeviceCaps(hdc, 88)  # LOGPIXELSX
            ctypes.windll.user32.ReleaseDC(0, hdc)
            dpi_scale = dpi / 96.0  # 96 DPI = 100%
            return dpi_scale
        except Exception:
            return 1.0

    # 🔥 視窗大小設定：從 main.py 生成的設定檔讀取，自動填滿剩餘空間
    finance_window_width = 1024
    finance_window_height = 1024
    finance_window_x = 0
    finance_window_y = 0

    try:
        from base_dir_helper import get_base_dir
        BASE_DIR = get_base_dir()
        config_path = os.path.join(BASE_DIR, 'window_config.json')
        if os.path.exists(config_path):
            with open(config_path, 'r', encoding='utf-8') as f:
                window_config = json.load(f)

            work_area = window_config.get('work_area', {})
            main_window = window_config.get('main_window', {})

            if work_area and main_window:
                # 取得 DPI 縮放比例
                dpi_scale = window_config.get('dpi_scale')
                if dpi_scale is None:
                    dpi_scale = get_dpi_scale()

                # 取得主程式視窗資訊
                main_x = main_window.get('x', 1280)
                main_y = main_window.get('y', 0)
                main_w = main_window.get('width', 640)
                main_h = main_window.get('height', 1020)

                # 根據主程式位置，智能計算 Chrome 視窗的位置和大小
                work_left = work_area.get('left', 0)
                work_top = work_area.get('top', 0)
                work_right = work_area.get('right', 1920)
                work_bottom = work_area.get('bottom', 1080)

                # 計算 Chrome 視窗大小：固定佔螢幕 2/3 寬度
                screen_width_physical = work_right - work_left
                chrome_width = (screen_width_physical * 2) // 3

                # 判斷主程式在螢幕的哪一側，決定 Chrome 位置
                main_window_center_x = main_x + main_w / 2
                screen_center_x = (work_left + work_right) / 2

                if main_window_center_x > screen_center_x:
                    # 主程式在右側，Chrome 放左側
                    chrome_x = work_left
                else:
                    # 主程式在左側，Chrome 放右側
                    chrome_x = work_right - chrome_width

                # 高度使用完整工作區域高度
                chrome_y = work_top
                chrome_height = work_bottom - work_top - 20

                # 確保視窗大小合理
                if chrome_width < 800:
                    chrome_width = 800
                if chrome_height < 600:
                    chrome_height = 600

                # 轉換為 Chrome 的邏輯像素
                chrome_width_logical = int(chrome_width / dpi_scale)
                chrome_x_logical = int(chrome_x / dpi_scale)
                chrome_y_logical = int(chrome_y / dpi_scale)

                # 🔥 Chrome 視窗高度需要加上標題欄高度（約 32px 邏輯像素）
                # 這樣整個視窗（包含標題欄）才會和主程式一樣高
                CHROME_TITLEBAR_HEIGHT = 32
                chrome_height_logical = int(chrome_height / dpi_scale) + CHROME_TITLEBAR_HEIGHT

                finance_window_width = chrome_width_logical
                finance_window_height = chrome_height_logical
                finance_window_x = chrome_x_logical
                finance_window_y = chrome_y_logical
    except Exception as e:
        pass

    # 🔥 嘗試讀取 data.json 以補充行政區資訊
    data_json_info = None
    try:
        from base_dir_helper import get_data_json_path
        data_json_path = get_data_json_path()
        if os.path.exists(data_json_path):
            with open(data_json_path, 'r', encoding='utf-8') as df:
                data_list = json.load(df)
                if data_list and len(data_list) > 0:
                    data_json_info = data_list[0]
                    flush_print(f"[資訊] 從 data.json 讀取輔助資訊：{data_json_info.get('city', '')}{data_json_info.get('area', '')}")
    except Exception as e:
        flush_print(f"[警告] 無法讀取 data.json：{e}")

    flush_print(f"開始解析JSON，共 {len(blocks)} 個區塊...")

    for idx, block in enumerate(blocks, 1):
        flush_print(f"--- 處理第 {idx} 筆資料 ---")
        
        # 驗證謄本類型
        is_valid_type, transcript_type = validate_transcript_type(block)
        if not is_valid_type:
            flush_print(f"X 第 {idx} 筆跳過：謄本類型 '{transcript_type}' 不是土地謄本第一類或第二類")
            continue
        
        flush_print(f"✓ 謄本類型符合：{transcript_type}")
        
        # 解析基本資訊
        basic = (block or {}).get("基本資訊") or {}
        title = (basic.get("地號建號") or basic.get("地號") or "").strip()
        geo = parse_geo_from_title(title)

        # 🔥 如果 geo 中缺少行政區或地段資訊，從 data.json 補充
        if data_json_info:
            if not geo.get("district") and data_json_info.get("area"):
                geo["district"] = data_json_info.get("area", "")
                flush_print(f"  [補充] 從 data.json 補充行政區：{geo['district']}")
            if not geo.get("section_hint") and data_json_info.get("section"):
                geo["section_hint"] = data_json_info.get("section", "")
                flush_print(f"  [補充] 從 data.json 補充地段：{geo['section_hint']}")
            if not geo.get("lot_number") and data_json_info.get("lot_number"):
                lot_from_data = data_json_info.get("lot_number", "")
                flush_print(f"  [補充] 從 data.json 補充地號：{lot_from_data}")

                # 🔥 同時補充 lot_front 和 lot_tail 用於 lot_display
                # 處理不同的地號格式
                if len(lot_from_data) == 8:
                    # 8位數格式：前4位母號，後4位子號
                    geo["lot_front"] = lot_from_data[:4].lstrip('0') or '0'
                    geo["lot_tail"] = lot_from_data[4:]
                    # lot_number 保持原始格式（去除前導零）
                    if geo["lot_tail"] == '0000':
                        geo["lot_number"] = geo["lot_front"]
                    else:
                        geo["lot_number"] = f"{geo['lot_front']}-{geo['lot_tail'].lstrip('0')}"
                    flush_print(f"  [補充] 解析8位地號：母號{geo['lot_front']} 子號{geo['lot_tail']} -> lot_number={geo['lot_number']}")
                elif '-' in lot_from_data:
                    # 格式如 "408-1"
                    parts = lot_from_data.split('-')
                    geo["lot_front"] = parts[0]
                    geo["lot_tail"] = parts[1].zfill(4) if len(parts) > 1 else '0000'
                    geo["lot_number"] = lot_from_data  # 保持原始格式
                    flush_print(f"  [補充] 解析分隔地號：母號{geo['lot_front']} 子號{geo['lot_tail']} -> lot_number={geo['lot_number']}")
                else:
                    # 只有母號，如 "408"
                    geo["lot_front"] = lot_from_data
                    geo["lot_tail"] = '0000'
                    geo["lot_number"] = lot_from_data  # 保持原始格式
                    flush_print(f"  [補充] 解析單一地號：母號{geo['lot_front']} -> lot_number={geo['lot_number']}")

        # 擷取土地面積
        land_info = (block or {}).get("標示部") or {}
        area = ""
        if isinstance(land_info, dict):
            area = land_info.get("面積", "")
        elif isinstance(land_info, list) and land_info:
            area = land_info[0].get("面積", "") if isinstance(land_info[0], dict) else ""
        
        # 解析所有權人和年月資料
        owners_raw = (block or {}).get("所有權部") or []
        owners = []
        
        # 擷取所有年月資料
        year_months = extract_year_month_from_ownership(owners_raw)
        cpi_values = []
        
        if cpi_dict and year_months:
            flush_print(f"  找到年月資料：{year_months}")
            for ym in year_months:
                cpi_value = get_cpi_for_year_month(ym, cpi_dict)
                if cpi_value:
                    cpi_values.append(cpi_value)
                    flush_print(f"  {ym} 對應CPI：{cpi_value}")
                else:
                    flush_print(f"  {ym} 找不到對應CPI")
        
        # 🔥 識別公同共有群組
        co_ownership_groups = {}  # {group_id: [owner1, owner2, ...]}

        for ow in owners_raw:
            if not isinstance(ow, dict):
                continue
            name = (ow.get("所有權人") or ow.get("姓名") or "").strip() or "（未命名）"
            rights_range = ow.get("權利範圍", "")
            other_notes = ow.get("其他登記事項", "")
            login_sequence = ow.get("登記次序", "")

            # 檢查是否為公同共有
            is_co_ownership = "公同共有" in rights_range
            group_id = None

            # 從「其他登記事項」中解析公同共有群組
            if is_co_ownership and other_notes:
                import re
                # 例如：「登記次序１３至１８為公同共有。」
                match = re.search(r'登記次序[０-９0-9]+至[０-９0-9]+為公同共有', other_notes)
                if match:
                    group_id = match.group(0)  # 使用完整註記作為群組ID

            # 如果是公同共有但沒有明確群組註記，使用權利範圍作為群組ID
            if is_co_ownership and not group_id:
                group_id = f"公同共有_{rights_range}"

            owner_info = {
                "name": name,
                "is_co_ownership": is_co_ownership,
                "co_ownership_group": group_id,
                "rights_range": rights_range,
                "login_sequence": login_sequence
            }

            # 如果是公同共有，加入群組
            if is_co_ownership and group_id:
                if group_id not in co_ownership_groups:
                    co_ownership_groups[group_id] = []
                co_ownership_groups[group_id].append(owner_info)

            owners.append(owner_info)

        if not owners:
            owners = [{"name": "（未提供）"}]

        # 保留完整的原始數據
        # 🔥 處理 lot_display：如果子號是 0000，只顯示母號
        lot_display_value = ""
        if geo.get('lot_front'):
            lot_tail = geo.get('lot_tail', '0000')
            if lot_tail == '0000':
                lot_display_value = geo.get('lot_front', '')
            else:
                lot_display_value = f"{geo.get('lot_front', '')}-{lot_tail.lstrip('0')}"

        payload = {
            "district": geo.get("district", ""),
            "section_hint": geo.get("section_hint", ""),
            "lot_number": geo.get("lot_number", ""),
            "lot_display": lot_display_value,
            "area": area,
            "owners": owners,
            "transcript_type": transcript_type,
            "year_months": year_months,
            "cpi_values": cpi_values,

            # 🔥 公同共有群組資訊
            "co_ownership_groups": co_ownership_groups,

            # 保留完整的原始JSON數據
            "標示部": block.get("標示部", {}),
            "所有權部": block.get("所有權部", []),
            "基本資訊": block.get("基本資訊", {}),
            "他項權利部": block.get("他項權利部", [])
        }
        
        payloads.append(payload)
        flush_print(f"✓ 第 {idx} 筆解析完成：{geo.get('district', '')}{geo.get('section_hint', '')} 地號 {geo.get('lot_number', '')}")

    flush_print(f"\n解析完成，共處理 {len(payloads)} 筆有效資料", "important")
    return payloads

def clear_form(driver):
    """財政部表單不需要清除，跳過此步驟"""
    flush_print("財政部表單不需要清除，跳過此步驟")
    return True

def determine_land_type(payload):
    """根據JSON標示部判斷土地類型"""
    land_info = payload.get("標示部") or {}
    
    # 處理標示部可能是字典或列表的情況
    if isinstance(land_info, list) and land_info:
        land_info = land_info[0] if isinstance(land_info[0], dict) else {}
    elif not isinstance(land_info, dict):
        land_info = {}
    
    use_zone = land_info.get("使用分區", "").strip()
    use_type = land_info.get("使用地類別", "").strip()
    
    flush_print(f"使用分區: '{use_zone}', 使用地類別: '{use_type}'")
    
    # 如果都是"（空白）"就是都市土地
    if use_zone == "（空白）" and use_type == "（空白）":
        return "urban", "都市土地"
    # 如果有其他資料就是非都市土地
    elif use_zone or use_type:
        return "rural", "非都市土地"
    else:
        # 預設為都市土地
        return "urban", "都市土地"

def extract_form_data_from_json(payload, owner_index=None):
    """從JSON解析表單所需的實際資料 - 支援多組歷次移轉

    Args:
        payload: 包含謄本資料的字典
        owner_index: 指定要處理的所有權人索引（0-based），若為 None 則處理第一個所有權人
    """
    try:
        flush_print(f"=== 開始解析JSON資料（多組模式）所有權人索引: {owner_index if owner_index is not None else 0} ===")

        # 取得標示部資料
        land_info = payload.get("標示部") or {}
        if isinstance(land_info, list) and land_info:
            land_info = land_info[0] if isinstance(land_info[0], dict) else {}
        elif not isinstance(land_info, dict):
            land_info = {}

        # 取得所有權部資料
        ownership_info = payload.get("所有權部") or []
        if not isinstance(ownership_info, list):
            ownership_info = [ownership_info] if ownership_info else []

        # 儲存所有組的資料
        all_transfer_records = []

        # 🔥 根據 owner_index 找到對應的所有權人資料
        target_owner = None
        if owner_index is not None and 0 <= owner_index < len(ownership_info):
            target_owner = ownership_info[owner_index]
        elif ownership_info:
            target_owner = ownership_info[0]

        if target_owner:
            
            # 解析公告土地現值（電傳格式數字與「元」間可能有空格，如「59,900 元/平方公尺」）
            current_land_value = land_info.get("公告土地現值", "")
            current_value = ""
            if current_land_value:
                import re
                match = re.search(r'([\d,]+)\s*元/平方公尺', current_land_value)
                if match:
                    current_value = match.group(1).replace(',', '')
            
            # 解析整體權利範圍（作為 fallback，當歷次取得權利範圍無法解析時使用）
            rights_range = target_owner.get("權利範圍", "")
            numerator = "1"
            denominator = "1"

            if rights_range:
                import re
                patterns = [
                    r'公同共有\s*(\d+)分之(\d+)',
                    r'(\d+)分之(\d+)',
                    r'全部\s+(\d+)分之(\d+)',
                ]
                for pattern in patterns:
                    match = re.search(pattern, rights_range)
                    if match:
                        denominator = match.group(1)
                        numerator = match.group(2)
                        break
            
            # 🔥 動態偵測組數：掃描 target_owner 鍵中所有「年月N」的 N，取最大值
            #    原本硬編碼 20 會把超過 20 組的資料默默丟掉導致試算錯誤
            import re as _re_gc
            _group_indices = []
            for _k in target_owner.keys():
                _m = _re_gc.match(r'^年月(\d+)$', str(_k))
                if _m:
                    _group_indices.append(int(_m.group(1)))
            max_groups = max(_group_indices) if _group_indices else 0

            if max_groups > 0:
                flush_print(f"  偵測到 {max_groups} 組歷次移轉資料")

            for i in range(1, max_groups + 1):
                year_month_key = f"年月{i}"
                land_value_key = f"地價{i}"
                rights_range_key = f"歷次取得權利範圍{i}"
                
                # 檢查這一組是否存在
                if year_month_key not in target_owner:
                    break  # 沒有更多組了

                year_month_str = target_owner.get(year_month_key, "")
                land_value_str = target_owner.get(land_value_key, "")
                rights_range_str = target_owner.get(rights_range_key, "")
                
                if not year_month_str:
                    continue  # 跳過空的組
                
                # 轉換年月格式
                year_month_code = ""
                import re
                patterns = [
                    r'(\d{3})年(\d{2})月',
                    r'(\d{3})年(\d{1})月',
                ]
                
                for pattern in patterns:
                    match = re.search(pattern, year_month_str)
                    if match:
                        year = match.group(1)
                        month = match.group(2).zfill(2)
                        year_month_code = f"{year}{month}"
                        break
                
                # 清理地價
                previous_value = ""
                if land_value_str:
                    cleaned = str(land_value_str).replace('元/平方公尺', '').replace(',', '').strip()
                    match = re.search(r'([\d.]+)', cleaned)
                    if match:
                        previous_value = match.group(1).split('.')[0]
                
                # 從「歷次取得權利範圍」解析該組的分子/分母
                # 優先用各組自己的歷次取得權利範圍，fallback 用整體 權利範圍
                rec_numerator = numerator
                rec_denominator = denominator
                if rights_range_str:
                    _m = re.search(r'(\d+)分之(\d+)', rights_range_str)
                    if _m:
                        rec_denominator = _m.group(1)
                        rec_numerator = _m.group(2)

                # 儲存這一組資料
                record = {
                    "group_number": i,
                    "current_value": current_value,
                    "numerator": rec_numerator,
                    "denominator": rec_denominator,
                    "year_month": year_month_code,
                    "year_month_display": year_month_str,
                    "previous_value": previous_value,
                    "rights_range": rights_range_str
                }
                
                all_transfer_records.append(record)
                flush_print(f"✓ 第{i}組: 年月={year_month_str}, 地價={previous_value}, 權利範圍={rights_range_str}")
        
        flush_print(f"=== 共解析 {len(all_transfer_records)} 組資料 ===")
        
        return all_transfer_records
        
    except Exception as e:
        flush_print(f"解析JSON資料時發生錯誤：{e}")
        flush_print(f"錯誤詳情: {traceback.format_exc()}")
        return []

def handle_price_index_query_finance(driver, cpi_values, form_data):
    """財政部版本物價指數處理 - 使用正確選擇器"""
    try:
        flush_print("處理物價指數...")
        
        # 財政部沒有自動查詢按鈕，直接填入計算好的CPI值
        if form_data.get('price_index'):
            try:
                price_field = driver.find_element(By.CSS_SELECTOR, SELECTORS["price_index"])
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", price_field)
                time.sleep(0.2)
                price_field.clear()
                price_field.send_keys(form_data['price_index'])
                driver.execute_script("arguments[0].dispatchEvent(new Event('input', {bubbles: true}));", price_field)
                driver.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles: true}));", price_field)
                flush_print(f"✓ 已填入物價指數：{form_data['price_index']}")
                return True
            except Exception as e:
                flush_print(f"填寫物價指數失敗：{e}")
                return False
        else:
            # 如果沒有預先計算的CPI值，使用備用邏輯
            cpi_value = "105.0"  # 預設值
            if cpi_values and len(cpi_values) > 0:
                cpi_value = str(cpi_values[0])
            
            try:
                price_field = driver.find_element(By.CSS_SELECTOR, SELECTORS["price_index"])
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", price_field)
                time.sleep(0.2)
                price_field.clear()
                price_field.send_keys(cpi_value)
                driver.execute_script("arguments[0].dispatchEvent(new Event('input', {bubbles: true}));", price_field)
                driver.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles: true}));", price_field)
                flush_print(f"✓ 已填入備用物價指數：{cpi_value}")
                return True
            except Exception as e:
                flush_print(f"填寫備用物價指數失敗：{e}")
                return False
            
    except Exception as e:
        flush_print(f"處理物價指數時發生錯誤：{e}")
        return False

def scroll_to_element_and_fill(driver, selector, value, field_name):
    """改進的捲動和填寫函數"""
    max_attempts = 3
    
    for attempt in range(max_attempts):
        try:
            # 先捲動到頁面頂部
            driver.execute_script("window.scrollTo(0, 0);")
            time.sleep(0.2)
            
            # 尋找元素
            element = driver.find_element(By.CSS_SELECTOR, selector)
            
            # 捲動到元素位置，確保在畫面中央
            driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", element)
            time.sleep(0.2)
            
            # 等待元素可互動
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, selector)))
            
            # 清除並填入數值
            element.clear()
            time.sleep(0.3)
            element.send_keys(str(value))
            
            flush_print(f"✓ 已填入{field_name}：{value}")
            return True
            
        except Exception as e:
            flush_print(f"第{attempt+1}次嘗試填寫{field_name}失敗：{str(e)}")
            
            if attempt < max_attempts - 1:
                time.sleep(0.5)
                # 嘗試不同的捲動位置
                scroll_positions = [0, 300, 600, 900, 1200, 1500]
                for pos in scroll_positions:
                    try:
                        driver.execute_script(f"window.scrollTo(0, {pos});")
                        time.sleep(0.2)
                        element = driver.find_element(By.CSS_SELECTOR, selector)
                        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", element)
                        time.sleep(0.2)
                        break
                    except:
                        continue
            else:
                flush_print(f"✗ 填寫{field_name}最終失敗：{selector}")
                return False
    
    return False

def select_radio_with_scroll(driver, selector, option_name):
    """改進的單選按鈕選擇函數"""
    try:
        # 捲動到頁面不同位置尋找單選按鈕
        scroll_positions = [0, 400, 800, 1200, 1600]
        
        for pos in scroll_positions:
            try:
                driver.execute_script(f"window.scrollTo(0, {pos});")
                time.sleep(0.2)
                
                radio_element = driver.find_element(By.CSS_SELECTOR, selector)
                driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", radio_element)
                time.sleep(0.2)
                
                # 使用JavaScript點擊確保成功
                driver.execute_script("arguments[0].click();", radio_element)
                flush_print(f"✓ 已選擇：{option_name}")
                return True
                
            except Exception as e:
                continue
        
        flush_print(f"✗ 無法找到或選擇：{option_name}")
        return False
        
    except Exception as e:
        flush_print(f"選擇{option_name}失敗：{e}")
        return False

# def fill_form_interactive_finance(driver, payload, owner, transcript_dir, sequence_num):
    """改進的財政部版本表單填寫函數"""
    district = payload.get("district", "")
    section = payload.get("section_hint", "")
    lot_display = payload.get("lot_display", "")
    area = payload.get("area", "").replace("平方公尺", "").replace(",", "").strip()
    owner_name = owner.get("name", "")
    cpi_values = payload.get("cpi_values", [])
    
    flush_print(f"\n開始填寫財政部試算表單：{district}{section} 地號 {lot_display} - {owner_name}")
    
    try:
        time.sleep(0.5)
        
        # 從JSON解析實際資料
        form_data = extract_form_data_from_json(payload)
        
        flush_print(f"準備填入的數據:")
        flush_print(f"- 土地面積: {area}")
        flush_print(f"- 申報移轉現值: {form_data['current_value']}")
        flush_print(f"- 權利範圍: {form_data['numerator']}/{form_data['denominator']}")
        flush_print(f"- 前次移轉年月: {form_data['year_month']}")
        flush_print(f"- 前次移轉現值: {form_data['previous_value']}")
        
        # 等待頁面完全載入
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        time.sleep(0.5)
        
        # 1. 填寫土地面積 - 使用改進的函數
        scroll_to_element_and_fill(driver, 'input[name="landArea"]', area, "土地面積")
        
        # 2. 填寫申報移轉現值
        if form_data['current_value']:
            scroll_to_element_and_fill(driver, 'input[name="currentValue"]', form_data['current_value'], "申報移轉現值")
        
        # 3. 填寫權利範圍分子
        if form_data['numerator']:
            scroll_to_element_and_fill(driver, 'input[name="numerator"]', form_data['numerator'], "權利範圍分子")
        
        # 4. 填寫權利範圍分母
        if form_data['denominator']:
            scroll_to_element_and_fill(driver, 'input[name="denominator"]', form_data['denominator'], "權利範圍分母")
        
        # 5. 填寫前次移轉年月
        if form_data['year_month']:
            scroll_to_element_and_fill(driver, 'input[name="transferDate"]', form_data['year_month'], "前次移轉年月")
        
        # 6. 填寫前次移轉現值
        if form_data['previous_value']:
            prev_value = str(form_data['previous_value']).split('.')[0]
            scroll_to_element_and_fill(driver, 'input[name="previousValue"]', prev_value, "前次移轉現值")
        
        # 7. 關鍵：處理物價指數
        price_success = handle_price_index_query(driver, cpi_values, form_data)
        if not price_success:
            flush_print("✗ 物價指數處理失敗，這可能會導致試算錯誤")
        
        # 8. 選擇土地類型 - 使用改進的單選按鈕函數
        try:
            land_type, land_type_name = determine_land_type(payload)
            if land_type == "urban":
                select_radio_with_scroll(driver, 'input[type="radio"][value="urban"]', land_type_name)
            else:
                select_radio_with_scroll(driver, 'input[type="radio"][value="rural"]', land_type_name)
        except Exception as e:
            flush_print(f"選擇土地類型失敗：{e}")
        
        # 9. 保存填寫完成的PDF
        header_text = f"編號{sequence_num:03d} | {owner_name} | {district}{section} {lot_display}"
        driver.execute_script("window.scrollTo(0, 0);")
        time.sleep(0.5)
        
   
        if filled_pdf_path:
            final_filled_pdf = os.path.join(transcript_dir, f"filled_{sequence_num:03d}_{district}_{lot_display}_{owner_name}_with_header.pdf")
            if add_header_to_pdf_centered_middle(filled_pdf_path, header_text, final_filled_pdf):
                if os.path.exists(filled_pdf_path):
                    os.remove(filled_pdf_path)
                filled_pdf_path = final_filled_pdf
            
            flush_print(f"✓ 已保存帶標題的填寫PDF")
            return "filled", filled_pdf_path
        else:
            return "filled", None
                    
    except Exception as e:
        flush_print(f"填寫表單時發生錯誤：{e}")
        return False, None


def submit_form_and_save_result_finance(driver, transcript_dir, filename_base, sequence_num, owner_name, district, section, lot_display):
    try:
        # 提交試算
        flush_print("提交增值稅估算...")
        
        submit_btn = driver.find_element(By.CSS_SELECTOR, SELECTORS["submit_btn"])
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", submit_btn)
        time.sleep(0.5)
        
        # 點擊增值稅估算按鈕
        driver.execute_script("arguments[0].click();", submit_btn)
        flush_print("✓ 已點擊增值稅估算按鈕")

        # 🔥 等結果頁面載入（改用 WebDriverWait 偵測 NT$ 字樣，最多 5 秒）
        flush_print("等待試算結果載入...")
        try:
            WebDriverWait(driver, 5).until(
                lambda d: 'NT$' in d.find_element(By.TAG_NAME, 'body').text
                          or 'NT&nbsp;$' in d.find_element(By.TAG_NAME, 'body').text
            )
        except TimeoutException:
            flush_print("  試算結果等待逾時（5 秒），繼續提取")

        # === 重要：只調用一次稅額提取函數 ===
        tax_amounts = extract_tax_amounts_finance(driver)
        flush_print(f"從提取函數得到的稅額數據: {tax_amounts}")
        
        # 保存PDF等其他邏輯...
        result_pdf_path = save_page_as_pdf(driver, transcript_dir, f"result_{sequence_num:03d}_{filename_base}")
        
        # 確保返回正確的稅額數據 (移除重複調用)
        return True, result_pdf_path, tax_amounts
        
    except Exception as e:
        flush_print(f"提交表單時發生錯誤：{e}")
        # 出錯時返回空數據
        return False, None, {"一般稅額": "", "自用稅額": ""}

def extract_tax_amounts_finance(driver):
    """從財政部試算結果頁面提取稅額數據 - 清潔簡化版"""
    try:
        tax_data = {"一般稅額": "", "自用稅額": ""}
        
        # 獲取頁面文字內容
        try:
            page_text = driver.find_element(By.TAG_NAME, "body").text
        except Exception as e:
            flush_print(f"無法獲取頁面文字: {e}")
            return tax_data
        
        import re
        
        # 精確匹配 (一) 和 (二) 的稅額格式
        patterns = [
            r'\(一\).*?NT\$?([\d,]+)元',  # (一) = 自用稅額
            r'\(二\).*?NT\$?([\d,]+)元',  # (二) = 一般稅額
        ]
        
        for i, pattern in enumerate(patterns):
            matches = re.findall(pattern, page_text, re.IGNORECASE | re.DOTALL)
            if matches:
                amount = matches[0].replace(',', '').strip()
                if amount.isdigit() and int(amount) >= 0:  # 允許0值和所有正數
                    if i == 0:
                        tax_data["自用稅額"] = amount
                        flush_print(f"✓ 找到自用稅額: {matches[0]}")
                    elif i == 1:
                        tax_data["一般稅額"] = amount
                        flush_print(f"✓ 找到一般稅額: {matches[0]}")
        
        # 備用方法：如果精確匹配失敗
        if not tax_data["一般稅額"] and not tax_data["自用稅額"]:
            all_amounts = re.findall(r'NT\$?([\d,]+)元', page_text)
            valid_amounts = [amt.replace(',', '') for amt in all_amounts 
                           if amt.replace(',', '').isdigit() and 100000 <= int(amt.replace(',', '')) <= 1000000]
            
            if len(valid_amounts) >= 2:
                tax_data["自用稅額"] = valid_amounts[0]
                tax_data["一般稅額"] = valid_amounts[1]
                flush_print(f"備用方法 - 自用: {valid_amounts[0]}, 一般: {valid_amounts[1]}")
        
        return tax_data
        
    except Exception as e:
        flush_print(f"提取稅額錯誤: {e}")
        return {"一般稅額": "", "自用稅額": ""}

def extract_tax_value_generic(driver, keywords):
    """通用稅額數值提取函數"""
    for keyword in keywords:
        try:
            # 查找包含關鍵字的元素
            elements = driver.find_elements(By.XPATH, f"//*[contains(text(), '{keyword}')]")
            
            for element in elements:
                # 在元素附近查找數字
                parent_text = element.find_element(By.XPATH, "..").text
                
                # 提取數字
                import re
                numbers = re.findall(r'(\d{1,3}(?:,\d{3})*)\s*元', parent_text)
                if numbers:
                    return int(numbers[0].replace(',', ''))
                    
        except:
            continue
    
    return 0

#===================================
# PDF處理和Excel報表函數（與台中版相同）
def setup_chinese_font():
    """設置中文字體"""
    try:
        font_paths = [
            "C:\\Windows\\Fonts\\msjh.ttc",
            "C:\\Windows\\Fonts\\kaiu.ttf",
            "C:\\Windows\\Fonts\\mingliu.ttc",
        ]
        
        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    pdfmetrics.registerFont(TTFont('ChineseFont', font_path))
                    return 'ChineseFont'
                except:
                    continue
        
        return 'Helvetica'
    except:
        return 'Helvetica'

def create_enhanced_excel_summary(results, output_path):
    """創建增強版Excel報告，包含稅額數據和加總 - 支援統一編號和多組資料"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "財政部土地增值稅試算統計"

        # 第一列：標題欄（A1~K1 跨欄合併）+ 製表日期（L1~M1）
        ws.merge_cells('A1:K1')
        title_cell = ws['A1']
        title_cell.value = "財 政 部 土 地 增 值 稅 試 算 列 表"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25  # 設定標題列高度

        # 🔥 製表日期（民國格式）
        from datetime import datetime
        today = datetime.now()
        roc_year = today.year - 1911
        date_str = f"製表日期：民國{roc_year:03d}年{today.month:02d}月{today.day:02d}日"
        ws.merge_cells('L1:M1')
        date_cell = ws['L1']
        date_cell.value = date_str
        date_cell.font = Font(size=10)
        date_cell.alignment = Alignment(horizontal='right', vertical='center')

        # 第二列：表頭
        headers = ["編號", "行政區", "地段", "地號", "所有權人", "統一編號", "面積", "土地公告現值", "前次移轉現值或原規定", "前次移轉年月日", "物價指數", "自用住宅稅額", "一般土地稅額"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

        # 初始化總計變數
        total_general = 0
        total_once = 0
        owner_totals = {}

        # 🔥 只有一個 for 迴圈（從第 3 列開始）
        for row_idx, result in enumerate(results, 3):
            # 取得基本資訊
            owner_name = result.get("owner_name", "")
            group_number = result.get("group_number", 1)
            total_groups = result.get("total_groups", 1)
            owner_id = result.get("owner_id", "")
            
            # 建立唯一識別 key
            owner_key = f"{owner_name}|{owner_id}"
            
            # 基本信息
            # 🔥 編號優先用 result.sequence_num（程式已維護好的流水號），
            #    否則退回 row_idx-2（因為資料從第 3 列開始，第一筆應為 1）
            ws.cell(row=row_idx, column=1, value=result.get("sequence_num", row_idx-2))
            ws.cell(row=row_idx, column=2, value=result.get("district", ""))
            ws.cell(row=row_idx, column=3, value=result.get("section", ""))
            ws.cell(row=row_idx, column=4, value=result.get("lot_display", ""))
            
            # 所有權人（顯示組別資訊）
            if total_groups > 1:
                ws.cell(row=row_idx, column=5, value=f"{owner_name} (第{group_number}/{total_groups}組)")
            else:
                ws.cell(row=row_idx, column=5, value=owner_name)
            
            # 統一編號
            ws.cell(row=row_idx, column=6, value=owner_id)

            # 取得 payload_data
            payload_data = result.get("payload_data", {})

            # 其他欄位
            ws.cell(row=row_idx, column=7, value=result.get("area", ""))

            # 數值欄位（靠右對齊）
            cell_8 = ws.cell(row=row_idx, column=8, value=payload_data.get("current_value", ""))
            cell_8.alignment = Alignment(horizontal='right')

            cell_9 = ws.cell(row=row_idx, column=9, value=payload_data.get("previous_value", ""))
            cell_9.alignment = Alignment(horizontal='right')

            cell_10 = ws.cell(row=row_idx, column=10, value=payload_data.get("year_month", ""))
            cell_10.alignment = Alignment(horizontal='right')

            cell_11 = ws.cell(row=row_idx, column=11, value=payload_data.get("cpi_value", ""))
            cell_11.alignment = Alignment(horizontal='right')
            
            # 稅額數據處理
            tax_data = result.get("tax_amounts", {})
            general_tax = tax_data.get("一般稅額", "") if isinstance(tax_data, dict) else ""
            once_tax = tax_data.get("自用稅額", "") if isinstance(tax_data, dict) else ""
            
            # 寫入稅額
            if general_tax is not None and str(general_tax).strip() != "":
                try:
                    clean_value = str(general_tax).replace(',', '').strip()
                    if clean_value.isdigit() or (clean_value.startswith('-') and clean_value[1:].isdigit()):
                        numeric_value = int(clean_value)
                        ws.cell(row=row_idx, column=13, value=numeric_value)
                        ws.cell(row=row_idx, column=13).number_format = '#,##0'
                        total_general += numeric_value
                except (ValueError, TypeError):
                    pass

            if once_tax and str(once_tax).strip():
                try:
                    clean_value = str(once_tax).replace(',', '').strip()
                    numeric_value = int(clean_value)
                    ws.cell(row=row_idx, column=12, value=numeric_value)
                    ws.cell(row=row_idx, column=12).number_format = '#,##0'
                    total_once += numeric_value
                except (ValueError, TypeError):
                    pass
            
            # 所有權人統計
            if owner_key not in owner_totals:
                owner_totals[owner_key] = {
                    "name": owner_name,
                    "id": owner_id,
                    "一般稅額": 0,
                    "自用稅額": 0,
                    "組數": 0
                }
            
            if general_tax:
                try:
                    amount = int(str(general_tax).replace(',', ''))
                    owner_totals[owner_key]["一般稅額"] += amount
                except:
                    pass
            
            if once_tax:
                try:
                    amount = int(str(once_tax).replace(',', ''))
                    owner_totals[owner_key]["自用稅額"] += amount
                except:
                    pass
            
            owner_totals[owner_key]["組數"] += 1
        
        # 添加總計行（因為增加了標題列，所以要 +4）
        total_row = len(results) + 4
        ws.cell(row=total_row, column=5, value="總計").font = Font(bold=True)
        ws.cell(row=total_row, column=12, value=total_once).font = Font(bold=True)
        ws.cell(row=total_row, column=12).number_format = '#,##0'
        ws.cell(row=total_row, column=13, value=total_general).font = Font(bold=True)
        ws.cell(row=total_row, column=13).number_format = '#,##0'

        # 添加按所有權人的加總
        if owner_totals:
            owner_start_row = total_row + 3
            ws.cell(row=owner_start_row, column=5, value="各所有權人小計").font = Font(bold=True)

            current_row = owner_start_row + 1
            for owner_key, totals in owner_totals.items():
                ws.cell(row=current_row, column=5, value=f"{totals['name']} ({totals['組數']}組)")
                ws.cell(row=current_row, column=6, value=totals['id'])
                ws.cell(row=current_row, column=12, value=totals['自用稅額'])
                ws.cell(row=current_row, column=12).number_format = '#,##0'
                ws.cell(row=current_row, column=13, value=totals['一般稅額'])
                ws.cell(row=current_row, column=13).number_format = '#,##0'
                current_row += 1
        
        # 自動調整欄位寬度
        from openpyxl.cell.cell import MergedCell
        for column_cells in ws.columns:
            length = 0
            column_letter = None

            # 找到第一個非合併儲存格來取得 column_letter
            for cell in column_cells:
                if not isinstance(cell, MergedCell):
                    column_letter = cell.column_letter
                    break

            # 如果整欄都是合併儲存格，跳過
            if column_letter is None:
                continue

            for cell in column_cells:
                try:
                    # 跳過合併儲存格
                    if isinstance(cell, MergedCell):
                        continue

                    cell_value = str(cell.value)
                    if cell_value and cell_value != 'None':
                        # 計算中文字元數量（更精確的判斷）
                        chinese_chars = 0
                        for char in cell_value:
                            # 中文字元、全形符號的 Unicode 範圍
                            if '\u4e00' <= char <= '\u9fff' or '\u3000' <= char <= '\u303f' or '\uff00' <= char <= '\uffef':
                                chinese_chars += 1

                        english_chars = len(cell_value) - chinese_chars
                        # 中文字元寬度 = 2.2，英文字元寬度 = 1.1
                        cell_length = chinese_chars * 2.2 + english_chars * 1.1
                        if cell_length > length:
                            length = cell_length
                except:
                    pass

            if length > 0:
                ws.column_dimensions[column_letter].width = length + 2
            else:
                ws.column_dimensions[column_letter].width = 10

        # 特別設定編號欄位寬度（A欄）
        ws.column_dimensions['A'].width = 6

        # 🔥 設定 L 和 M 欄寬度相同（稅額欄位）
        ws.column_dimensions['L'].width = 14
        ws.column_dimensions['M'].width = 14

        # 設定頁面為 A4 橫式列印，所有欄位在同一頁寬中
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE  # 橫式
        ws.page_setup.paperSize = ws.PAPERSIZE_A4  # A4
        ws.page_setup.fitToPage = True  # 縮放以適應頁面
        ws.page_setup.fitToWidth = 1  # 所有欄位適應一頁寬
        ws.page_setup.fitToHeight = 0  # 高度不限制（自動分頁）

        # 設定列印邊界（單位：英吋）
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3

        wb.save(output_path)
        # flush_print(f"✓ 增強版Excel報告已保存：{output_path}", "important")  # 🔥 訊息簡化
        return True
        
    except Exception as e:
        flush_print(f"創建增強版Excel報告失敗：{e}", "important")
        flush_print(f"詳細錯誤追蹤：\n{traceback.format_exc()}", "important")
        return False

def zfill8_digits(s: str) -> str:
    return re.sub(r"\D", "", s or "").zfill(8)[:8]

def label_for_lot_dir(district: str, section: str, lot8: str) -> str:
    """回傳資料夾名稱：{行政區}{段名}-{地號}

    地號格式處理：
    - 如果包含 "-"，例如 "408-1"，保持原樣
    - 如果不包含 "-"，例如 "408"，視為母號（不加子號）
    """
    # 🔥 直接使用原始地號，不要過度處理
    # 移除非數字字符，但保留 "-"
    lot_clean = re.sub(r'[^\d-]', '', lot8 or "")

    # 如果是純數字（沒有 "-"），轉為整數再轉回字串以去除前導零
    if '-' not in lot_clean:
        lot_clean = str(int(lot_clean)) if lot_clean and lot_clean.isdigit() else lot_clean
    else:
        # 如果有 "-"，分別處理母號和子號的前導零
        parts = lot_clean.split('-')
        if len(parts) == 2:
            mother = str(int(parts[0])) if parts[0] and parts[0].isdigit() else parts[0]
            child = str(int(parts[1])) if parts[1] and parts[1].isdigit() else parts[1]
            lot_clean = f"{mother}-{child}" if child != '0' else mother

    label = f"{district}{section}-{lot_clean}"
    return label

def save_page_as_pdf(driver, transcript_dir, filename):
    """保存頁面為PDF"""
    try:
        os.makedirs(transcript_dir, exist_ok=True)
        pdf_path = os.path.join(transcript_dir, f"{filename}.pdf")
        
        result = driver.execute_cdp_cmd("Page.printToPDF", {
            "landscape": False,
            "printBackground": True,
            "scale": 0.8
        })
        
        if 'data' in result:
            import base64
            pdf_data = base64.b64decode(result['data'])
            with open(pdf_path, 'wb') as f:
                f.write(pdf_data)
            flush_print(f"PDF已保存：{pdf_path}")
            return pdf_path
        
    except Exception as e:
        flush_print(f"保存PDF失敗：{e}")
        return None

#==================================

def handle_year_month_dropdown(driver, year_month_str):
    """處理前次移轉年月的下拉選單（年份和月份分開）- 修正版"""
    try:
        flush_print(f"\n處理前次移轉年月下拉選單：{year_month_str}")
        
        # 先確保在頁面最頂部
        driver.execute_script("window.scrollTo(0, 0);")
        time.sleep(0.2)
        
        # 解析年月字串 (例如: "11112" -> 民國111年12月)
        if len(year_month_str) == 5:
            republic_year = int(year_month_str[:3])  # 111
            month = int(year_month_str[3:])          # 12
        elif len(year_month_str) == 4:
            republic_year = int(year_month_str[:2])   # 87  
            month = int(year_month_str[2:])           # 06
        else:
            flush_print(f"✗ 無法解析年月格式：{year_month_str}")
            return False
        
        # 🎯 關鍵修正：轉換為西元年（根據DevTools發現）
        western_year = republic_year + 1911  # 民國111年 = 西元2022年
        
        flush_print(f"解析結果：民國{republic_year}年{month}月 → 西元{western_year}年")
        
        # 處理年份下拉選單
        try:
            # 確保年份下拉選單可見
            year_select_element = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.ID, "year"))
            )
            year_select = Select(year_select_element)
            
            # 先查看所有可用的年份選項（調試用）
            year_options = [option.get_attribute('value') for option in year_select.options]
            flush_print(f"可用年份選項：{year_options[:5]}...{year_options[-5:]} (共{len(year_options)}個)")
            
            # 🔥 直接使用西元年作為選擇值
            target_year = str(western_year)
            
            if target_year in year_options:
                year_select.select_by_value(target_year)
                flush_print(f"✓ 成功選擇年份：{target_year} (西元年)")
                year_selected = True
            else:
                flush_print(f"✗ 找不到年份選項：{target_year}")
                flush_print(f"  最接近的選項：")
                # 顯示最接近的年份選項
                numeric_options = [int(opt) for opt in year_options if opt.isdigit()]
                if numeric_options:
                    closest_years = sorted(numeric_options, key=lambda x: abs(x - western_year))[:3]
                    for cy in closest_years:
                        flush_print(f"    {cy} (民國{cy-1911}年)")
                return False
            
        except Exception as e:
            flush_print(f"✗ 選擇年份失敗：{e}")
            return False
        
        # 處理月份下拉選單  
        try:
            time.sleep(0.2)  # 等待年份選擇完成
            
            month_select_element = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.ID, "month"))
            )
            month_select = Select(month_select_element)
            
            # 查看月份選項
            month_options = [option.get_attribute('value') for option in month_select.options]
            flush_print(f"可用月份選項：{month_options}")
            
            # 嘗試不同的月份格式
            month_formats_to_try = [
                str(month),        # 12
                f"{month:02d}",    # 12
            ]
            
            month_selected = False
            for month_format in month_formats_to_try:
                try:
                    if month_format in month_options:
                        month_select.select_by_value(month_format)
                        flush_print(f"✓ 已選擇月份：{month_format}")
                        month_selected = True
                        break
                except:
                    continue
            
            if not month_selected:
                flush_print(f"✗ 無法找到匹配的月份選項")
                flush_print(f"  期望月份：{month}, 可用選項：{month_options}")
                return False
            
            return True
            
        except Exception as e:
            flush_print(f"✗ 選擇月份失敗：{e}")
            return False
            
    except Exception as e:
        flush_print(f"處理年月下拉選單時發生錯誤：{e}")
        return False

import requests

def download_and_parse_cpi_data():
    """下載並解析消費者物價指數數據 - 使用爬蟲取得最新網址"""
    cpi_data = {}
    cpi_file = None

    try:
        # 🔥 使用 download_cpi_excel 函數（會爬蟲取得最新網址）
        cpi_file = download_cpi_excel()

        if not cpi_file or not os.path.exists(cpi_file):
            flush_print("✗ 無法下載 CPI 檔案")
            return {}
        
        # 解析Excel檔案
        try:
            import pandas as pd

            # 🔥 顯示正在讀取的檔案路徑
            flush_print(f"正在解析 CPI 檔案：{cpi_file}")

            # 🎯 根據實際結構：跳過標題行，讀取資料
            df = pd.read_excel(cpi_file, sheet_name="稅務專用", header=None)
            # 然後從第5行開始處理（index=4）
            for index in range(4, len(df)):
                flush_print(f"✓ 成功讀取CPI資料，共{len(df)}行")
            
            # 🔍 調試：檢查實際的資料結構
            flush_print(f"CPI資料前3行樣本：")
            for i in range(min(3, len(df))):
                row_data = [str(df.iloc[i, j]) for j in range(min(13, len(df.columns)))]
                flush_print(f"  第{i+1}行：{row_data}")
            
            # 解析資料：A欄=民國年，B到M欄=1月到12月
            for index, row in df.iterrows():
                try:
                    # A欄：民國年份
                    republic_year_value = row.iloc[0]
                    
                    if pd.isna(republic_year_value):
                        continue
                    
                    # 提取民國年份數字
                    if isinstance(republic_year_value, str):
                        import re
                        year_match = re.search(r'(\d+)', str(republic_year_value))
                        if year_match:
                            republic_year = int(year_match.group(1))
                        else:
                            continue
                    else:
                        republic_year = int(float(republic_year_value))
                    
                    # 檢查年份合理性 (民國80-120年)
                    if republic_year < 48 or republic_year > 120:
                        continue
                    
                    # B到M欄：1月到12月的CPI值
                    for month in range(1, 13):
                        col_index = month  # B欄=1月(index=1), C欄=2月(index=2), ..., M欄=12月(index=12)
                        
                        if col_index < len(row) and pd.notna(row.iloc[col_index]):
                            try:
                                cpi_value = float(row.iloc[col_index])
                                
                                # 建立多種查詢格式的索引
                                key1 = f"{republic_year:03d}{month:02d}"        # 08901
                                key2 = f"{republic_year:03d}年{month:02d}月"    # 089年01月  
                                key3 = f"{republic_year:02d}{month:02d}"        # 8901

                                cpi_data[key1] = str(cpi_value)
                                cpi_data[key2] = str(cpi_value)
                                if republic_year < 100:
                                    cpi_data[key3] = str(cpi_value)
                                
                                # # 調試：顯示關鍵範例
                                # if republic_year in [110, 111, 112] and month == 12:
                                #     flush_print(f"  CPI重要資料：民國{republic_year}年{month}月 → {key1} = {cpi_value}")
                                    
                            except (ValueError, TypeError):
                                continue
                                
                except (ValueError, IndexError, TypeError) as e:
                    continue
                    
            flush_print(f"✓ 解析完成，共{len(cpi_data)}筆CPI資料")
            
            # # 🔍 重點調試：檢查111年12月的資料
            # target_keys = ['11112', '111012']  # 可能的格式
            # for key in target_keys:
            #     if key in cpi_data:
            #         flush_print(f"✓ 找到目標資料：{key} = {cpi_data[key]}")
            #     else:
            #         flush_print(f"✗ 找不到：{key}")
            
            # # 顯示111年可用的月份
            # year_111_keys = [k for k in cpi_data.keys() if k.startswith('111')]
            # if year_111_keys:
            #     flush_print(f"民國111年可用月份：{sorted(year_111_keys)}")
            
        except Exception as e:
            flush_print(f"解析CPI檔案失敗：{e}")

            flush_print(f"詳細錯誤：{traceback.format_exc()}")
            return {}
        
    except Exception as e:
        flush_print(f"下載CPI資料時發生錯誤：{e}")
        return {}
    
    finally:
        # 清理下載的 CPI 檔案（可選，保留供調試）
        pass
    
    # 🔥 檢查 CPI 資料的最新年月（確認是否為最新資料）
    if cpi_data:
        # 找出所有年份和月份
        max_year = 0
        max_month = 0
        latest_key = ""
        latest_value = ""

        import re
        for key in cpi_data:
            year = None
            month = None

            # 嘗試解析格式1：114年11月
            match1 = re.search(r'(\d{2,3})年(\d{1,2})月', key)
            if match1:
                year = int(match1.group(1))
                month = int(match1.group(2))
            else:
                # 嘗試解析格式2：11411 或 11401（5位數：3位年+2位月）
                match2 = re.match(r'^(\d{3})(\d{2})$', key)
                if match2:
                    year = int(match2.group(1))
                    month = int(match2.group(2))
                else:
                    # 嘗試解析格式3：8901（4位數：2位年+2位月，適用於民國89年以前）
                    match3 = re.match(r'^(\d{2})(\d{2})$', key)
                    if match3:
                        year = int(match3.group(1))
                        month = int(match3.group(2))

            # 驗證月份合理性（1-12）
            if year is not None and month is not None and 1 <= month <= 12:
                if year > max_year or (year == max_year and month > max_month):
                    max_year = year
                    max_month = month
                    latest_key = key
                    latest_value = cpi_data[key]

        if latest_key:
            # 🔥 使用亮黃色顯示最新年月，讓使用者容易確認是否為最新資料
            yellow = "\033[93m"  # 亮黃色
            reset = "\033[0m"    # 重置顏色
            print(f"\n{yellow}✓ CPI 資料最新年月：民國 {max_year} 年 {max_month} 月 = {latest_value}{reset}", flush=True)

            # 檢查是否為最近的資料（至少要有今年或去年的資料）
            from datetime import datetime
            current_year = datetime.now().year - 1911  # 轉換為民國年

            if max_year >= current_year - 1:
                flush_print(f"✓ CPI 資料為最新版本（包含近期資料）", "success")
            else:
                flush_print(f"⚠️ 警告：CPI 資料可能過舊（最新為民國 {max_year} 年）", "warning")
        else:
            flush_print("⚠️ 無法判斷 CPI 資料的最新年月", "warning")

    return cpi_data


def parse_cpi_data(file_path):
    """解析CPI數據 - 加入詳細調試"""
    try:
        flush_print("=== CPI解析調試開始 ===", "important")
        flush_print(f"讀取檔案: {file_path}", "important")
        
        if file_path.endswith('.json'):
            with open(file_path, 'r', encoding='utf-8') as f:
                cpi_dict = json.load(f)
            flush_print(f"從JSON檔案載入 {len(cpi_dict)} 筆CPI資料", "important")
            return cpi_dict
            
        elif file_path.endswith('.xls'):
            # 🔥 對於舊的 .xls 檔案，直接使用 xlrd 1.2.0 讀取（繞過 pandas）
            import xlrd
            try:
                workbook = xlrd.open_workbook(file_path)
                # 嘗試找「稅務專用」工作表
                if "稅務專用" in workbook.sheet_names():
                    sheet = workbook.sheet_by_name("稅務專用")
                    flush_print(f"Excel檔案讀取成功（工作表：稅務專用），總行數: {sheet.nrows}", "important")
                else:
                    sheet = workbook.sheet_by_index(0)
                    flush_print(f"Excel檔案讀取成功（第一個工作表），總行數: {sheet.nrows}", "important")

                # 轉換為類似 pandas DataFrame 的結構（list of lists）
                df_data = []
                for row_idx in range(sheet.nrows):
                    row_values = [sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols)]
                    df_data.append(row_values)

                # 調試：顯示前幾行資料
                flush_print("Excel第1行資料預覽:", "important")
                if len(df_data) > 0:
                    row_data = [str(df_data[0][j]) if j < len(df_data[0]) else "" for j in range(min(5, len(df_data[0])))]
                    flush_print(f"  第1行: {row_data}", "important")

                cpi_dict = {}
                processed_count = 0

                # 從第5行開始讀取（索引4）
                # flush_print("開始解析CPI資料（從第5行開始）...", "important")  # 🔥 訊息簡化
                for index in range(4, len(df_data)):
                    row = df_data[index]

                    # 調試第一欄的值
                    first_col = row[0] if len(row) > 0 else None
                    # if index < 10:  # 只顯示前幾行的調試  # 🔥 訊息簡化
                    #     flush_print(f"  第{index+1}行第1欄: '{first_col}' (類型: {type(first_col)})", "important")

                    # xlrd 讀取的數字通常是 float 型態
                    if first_col is None or first_col == '' or not isinstance(first_col, (int, float)):
                        continue

                    try:
                        year_num = int(first_col)
                        if year_num < 48 or year_num > 120:
                            continue

                        # 特別調試民國89年  # 🔥 訊息簡化
                        # if year_num == 89:
                        #     flush_print(f"*** 發現民國89年資料 ***", "important")
                        #     month_data = [str(row[j]) if j < len(row) else "N/A" for j in range(1, 13)]
                        #     flush_print(f"民國89年12個月資料: {month_data}", "important")

                        # 處理12個月的資料
                        for month in range(1, 13):
                            col_index = month

                            if col_index < len(row) and row[col_index] not in (None, ''):
                                try:
                                    cpi_value = float(row[col_index])

                                    # 建立查詢key
                                    key1 = f"{year_num:03d}年{month:02d}月"
                                    key2 = f"{year_num:03d}{month:02d}"

                                    cpi_dict[key1] = str(cpi_value)
                                    cpi_dict[key2] = str(cpi_value)

                                    if year_num < 100:
                                        key3 = f"{year_num:02d}{month:02d}"
                                        cpi_dict[key3] = str(cpi_value)

                                    # 特別記錄民國89年1月  # 🔥 訊息簡化
                                    # if year_num == 89 and month == 1:
                                    #     flush_print(f"*** 民國89年1月CPI: {cpi_value} ***", "important")
                                    #     flush_print(f"產生的key: {key1}, {key2}", "important")

                                    processed_count += 1

                                except (ValueError, TypeError) as e:
                                    # if year_num == 89:  # 只對89年顯示錯誤  # 🔥 訊息簡化
                                    #     flush_print(f"民國89年{month}月資料轉換失敗: {row[col_index]} - {e}", "important")
                                    pass
                    except Exception as e:
                        # if index < 10:  # 只顯示前幾行的錯誤  # 🔥 訊息簡化
                        #     flush_print(f"  第{index+1}行處理失敗: {e}", "important")
                        pass

                flush_print(f"CPI解析完成，共處理 {processed_count} 筆資料", "important")
                flush_print(f"CPI字典總key數: {len(cpi_dict)}", "important")

                # 🔥 顯示最新年月（亮黃色）
                max_year, max_month = 0, 0
                for key in cpi_dict.keys():
                    if len(key) == 5 and key.isdigit():  # 格式如 "11412"
                        year = int(key[:3])
                        month = int(key[3:])
                        if year > max_year or (year == max_year and month > max_month):
                            max_year, max_month = year, month
                if max_year > 0:
                    yellow = "\033[93m"
                    reset = "\033[0m"
                    print(f"{yellow}✓ CPI 資料最新年月：民國 {max_year} 年 {max_month} 月{reset}", flush=True)

                return cpi_dict

            except Exception as e:
                flush_print(f"使用xlrd讀取失敗: {e}，嘗試使用pandas", "important")
                # 如果 xlrd 失敗，回退到 pandas
                try:
                    df = pd.read_excel(file_path, sheet_name="稅務專用", header=None)
                    flush_print(f"Excel檔案讀取成功（使用pandas），總行數: {len(df)}", "important")
                except Exception as e2:
                    flush_print(f"讀取稅務專用工作表失敗: {e2}，嘗試第一個工作表", "important")
                    df = pd.read_excel(file_path, sheet_name=0, header=None)

        elif file_path.endswith('.xlsx'):
            # 🔥 對於新的 .xlsx 檔案，使用 pandas + openpyxl
            try:
                df = pd.read_excel(file_path, sheet_name="稅務專用", header=None, engine='openpyxl')
                flush_print(f"Excel檔案讀取成功，總行數: {len(df)}", "important")
            except Exception as e:
                flush_print(f"讀取稅務專用工作表失敗: {e}，嘗試第一個工作表", "important")
                df = pd.read_excel(file_path, sheet_name=0, header=None, engine='openpyxl')
            
            # 調試：顯示前幾行資料
            flush_print("Excel第1行資料預覽:", "important")
            if len(df) > 0:
                row_data = [str(df.iloc[0, j]) if j < len(df.columns) else "" for j in range(min(5, len(df.columns)))]
                flush_print(f"  第1行: {row_data}", "important")
            
            cpi_dict = {}
            processed_count = 0

            # 從第5行開始讀取
            # flush_print("開始解析CPI資料（從第5行開始）...", "important")  # 🔥 訊息簡化
            for index in range(4, len(df)):
                row = df.iloc[index]

                # 調試第一欄的值
                first_col = row.iloc[0] if len(row) > 0 else None
                # if index < 10:  # 只顯示前幾行的調試  # 🔥 訊息簡化
                #     flush_print(f"  第{index+1}行第1欄: '{first_col}' (類型: {type(first_col)})", "important")

                if pd.isna(first_col) or not isinstance(first_col, (int, float)):
                    continue
                
                try:
                    year_num = int(first_col)
                    if year_num < 48 or year_num > 120:
                        continue

                    # 特別調試民國89年  # 🔥 訊息簡化
                    # if year_num == 89:
                    #     flush_print(f"*** 發現民國89年資料 ***", "important")
                    #     month_data = [str(row.iloc[j]) if j < len(row) else "N/A" for j in range(1, 13)]
                    #     flush_print(f"民國89年12個月資料: {month_data}", "important")

                    # 處理12個月的資料
                    for month in range(1, 13):
                        col_index = month
                        
                        if col_index < len(row) and pd.notna(row.iloc[col_index]):
                            try:
                                cpi_value = float(row.iloc[col_index])
                                
                                # 建立查詢key
                                key1 = f"{year_num:03d}年{month:02d}月"
                                key2 = f"{year_num:03d}{month:02d}"
                                
                                cpi_dict[key1] = str(cpi_value)
                                cpi_dict[key2] = str(cpi_value)
                                
                                if year_num < 100:
                                    key3 = f"{year_num:02d}{month:02d}"
                                    cpi_dict[key3] = str(cpi_value)

                                # 特別記錄民國89年1月  # 🔥 訊息簡化
                                # if year_num == 89 and month == 1:
                                #     flush_print(f"*** 民國89年1月CPI: {cpi_value} ***", "important")
                                #     flush_print(f"產生的key: {key1}, {key2}", "important")

                                processed_count += 1
                                
                            except (ValueError, TypeError) as e:
                                # if year_num == 89:  # 只對89年顯示錯誤  # 🔥 訊息簡化
                                #     flush_print(f"民國89年{month}月資料轉換失敗: {row.iloc[col_index]} - {e}", "important")
                                pass
                                
                except (ValueError, TypeError):
                    continue
            
            flush_print(f"CPI解析完成，共處理 {processed_count} 筆資料", "important")
            flush_print(f"字典總數: {len(cpi_dict)}", "important")

            # 🔥 顯示最新年月（亮黃色）
            max_year, max_month = 0, 0
            for key in cpi_dict.keys():
                if len(key) == 5 and key.isdigit():  # 格式如 "11412"
                    year = int(key[:3])
                    month = int(key[3:])
                    if year > max_year or (year == max_year and month > max_month):
                        max_year, max_month = year, month
            if max_year > 0:
                yellow = "\033[93m"
                reset = "\033[0m"
                print(f"{yellow}✓ CPI 資料最新年月：民國 {max_year} 年 {max_month} 月{reset}", flush=True)

            return cpi_dict
            
    except Exception as e:
        flush_print(f"CPI解析失敗: {e}", "important")
        import traceback
        flush_print(f"錯誤詳情: {traceback.format_exc()}", "important")
        return {}

def find_cpi_for_year_month(year_month_code, cpi_data):
    """查詢CPI值 - 加入詳細調試"""
    flush_print(f"=== CPI查詢調試 ===", "important")
    flush_print(f"查詢年月代碼: '{year_month_code}'", "important")
    flush_print(f"CPI資料庫大小: {len(cpi_data)}", "important")
    
    # 直接查詢
    if year_month_code in cpi_data:
        result = cpi_data[year_month_code]
        flush_print(f"直接查詢成功: {year_month_code} = {result}", "important")
        return result
    
    # 顯示相似的key
    similar_keys = [k for k in cpi_data.keys() if "089" in k or "89" in k]
    flush_print(f"相似的89年key: {similar_keys[:10]}", "important")
    
    # 嘗試不同格式
    if len(year_month_code) == 5:  # 08901
        year = year_month_code[:3]
        month = year_month_code[3:]
        
        alt_keys = [
            f"{year}年{month}月",  # 089年01月
            f"{int(year):02d}年{month}月",  # 89年01月  
            f"{int(year):02d}{month}",  # 8901
        ]
        
        flush_print(f"嘗試替代格式: {alt_keys}", "important")
        for alt_key in alt_keys:
            if alt_key in cpi_data:
                result = cpi_data[alt_key]
                flush_print(f"替代格式查詢成功: {alt_key} = {result}", "important")
                return result
    
    flush_print(f"所有查詢都失敗", "important")
    return None

def extract_year_month_from_json(payload):
    """從JSON資料中提取年月資訊 - 修正版"""
    year_months = []
    
    # 🔍 調試：顯示JSON中所有可能包含年月的欄位
    flush_print(f"\n檢查JSON中的年月相關欄位：")
    
    # 查找所有可能的年月欄位
    potential_fields = []
    for key, value in payload.items():
        # 檢查所有可能包含年月資料的欄位
        if any(keyword in key for keyword in ['年月', '日期', '移轉', 'date', '時間']):
            potential_fields.append((key, value))
            flush_print(f"  發現相關欄位：{key} = {value}")
    
    # 檢查所有有權部的資料
    if "所有權部" in payload:
        owners_data = payload["所有權部"]
        if isinstance(owners_data, list):
            for i, owner in enumerate(owners_data):
                if isinstance(owner, dict):
                    for key, value in owner.items():
                        if any(keyword in key for keyword in ['年月', '日期', '移轉']):
                            potential_fields.append((f"所有權部[{i}].{key}", value))
                            flush_print(f"  所有權部資料：{key} = {value}")
    
    # 解析找到的年月資料
    for field_name, field_value in potential_fields:
        if field_value:
            # 解析年月格式，例如："111年12月"、"民國111年12月"
            import re
            
            # 多種年月格式的正則表達式
            patterns = [
                r'(\d+)年(\d+)月',           # 111年12月
                r'民國(\d+)年(\d+)月',        # 民國111年12月  
                r'(\d+)/(\d+)',             # 111/12
                r'(\d{3,4})(\d{2})',        # 11112
            ]
            
            for pattern in patterns:
                match = re.search(pattern, str(field_value))
                if match:
                    republic_year = int(match.group(1))
                    month = int(match.group(2))
                    
                    # 處理不同的年份長度
                    if republic_year > 1000:  # 如果是4位數，可能是西元年
                        if republic_year > 1911:
                            republic_year = republic_year - 1911
                    
                    year_month_code = f"{republic_year:03d}{month:02d}"
                    year_months.append({
                        'field': field_name,
                        'display': field_value,
                        'code': year_month_code,
                        'republic_year': republic_year,
                        'month': month
                    })
                    flush_print(f"✓ 解析年月：{field_name} = {field_value} → 民國{republic_year}年{month}月 (代碼:{year_month_code})")
                    break
    
    if not year_months:
        flush_print(f"⚠ 未找到任何年月資料")
        # 顯示所有JSON欄位供調試
        flush_print(f"JSON所有欄位：")
        for key, value in payload.items():
            flush_print(f"  {key}: {str(value)[:50]}...")
    
    return year_months

def fill_element_by_id_no_scroll(driver, element_id, value, field_name):
    """使用ID填寫元素（不強制捲動到頂部）"""
    try:
        element = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.ID, element_id))
        )
        
        # 溫和的捲動到元素位置（不強制到頂部）
        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", element)
        time.sleep(0.2)
        
        # 確保元素可見且可互動
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, element_id)))
        
        # 點擊並清除
        element.click()
        time.sleep(0.3)
        element.clear()
        time.sleep(0.3)
        
        # 填入數值
        element.send_keys(str(value))
        time.sleep(0.2)
        
        # 觸發Angular事件
        driver.execute_script("""
            var element = arguments[0];
            element.dispatchEvent(new Event('input', { bubbles: true }));
            element.dispatchEvent(new Event('change', { bubbles: true }));
        """, element)
        
        # 驗證填寫結果
        current_value = element.get_attribute('value')
        if current_value == str(value):
            flush_print(f"✓ 成功填寫 {field_name}：{value}")
            return True
        else:
            flush_print(f"✗ 填寫驗證失敗 {field_name}：期望'{value}' 實際'{current_value}'")
            return False
        
    except Exception as e:
        flush_print(f"✗ 填寫 {field_name} 失敗：{e}")
        return False

def add_header_to_pdf_at_specific_location(input_pdf_path, header_text, output_pdf_path):
    """將標題加到PDF的指定位置（根據圖片中的紅框位置）"""
    try:
        # 讀取原始PDF
        reader = PdfReader(input_pdf_path)
        writer = PdfWriter()
        
        for page_num, page in enumerate(reader.pages):
            if page_num == 0:  # 只在第一頁加標題
                # 創建標題overlay
                packet = io.BytesIO()
                can = canvas.Canvas(packet, pagesize=A4)
                
                # 設定字型
                try:
                    # 使用系統中文字型
                    font_paths = [
                        "C:\\Windows\\Fonts\\msjh.ttc",
                        "C:\\Windows\\Fonts\\kaiu.ttf",
                        "C:\\Windows\\Fonts\\mingliu.ttc",
                    ]
                    
                    font_registered = False
                    for font_path in font_paths:
                        if os.path.exists(font_path):
                            try:
                                pdfmetrics.registerFont(TTFont('ChineseFont', font_path))
                                can.setFont("ChineseFont", 12)
                                font_registered = True
                                break
                            except:
                                continue
                    
                    if not font_registered:
                        can.setFont("Helvetica", 12)
                except:
                    can.setFont("Helvetica", 12)
                
                # 修改位置：中間的上面，再低一點
                page_width = A4[0]  # 595.27
                text_width = can.stringWidth(header_text, "ChineseFont", 12)
                x_position = (page_width - text_width) / 2  # 水平居中
                y_position = 650  # 垂直位置：上面一點但比中央低
                
                # 設定文字顏色為紅色
                can.setFillColorRGB(1, 0, 0)  # 紅色
                
                # 繪製標題文字
                can.drawString(x_position, y_position, header_text)
                
                # 完成標題層
                can.save()
                packet.seek(0)
                
                # 創建標題PDF
                overlay_pdf = PdfReader(packet)
                overlay_page = overlay_pdf.pages[0]
                
                # 將標題合併到原始頁面
                page.merge_page(overlay_page)
            
            writer.add_page(page)
        
        # 保存新PDF
        with open(output_pdf_path, 'wb') as output_file:
            writer.write(output_file)
        
        flush_print(f"✓ 已將標題加到指定位置：{output_pdf_path}")
        return True
        
    except Exception as e:
        flush_print(f"✗ 添加標題失敗：{e}")
        return False

def select_land_type_radio(driver, land_type="urban"):
    """選擇土地類型單選按鈕"""
    try:
        # 根據調試結果：landType0=都市土地, landType1=非都市土地
        if land_type == "urban":
            radio_id = "landType0" 
            type_name = "都市土地"
        else:
            radio_id = "landType1"
            type_name = "非都市土地"
        
        # 捲動到表單區域
        driver.execute_script("window.scrollTo(0, 600);")
        time.sleep(0.2)
        
        radio_element = driver.find_element(By.ID, radio_id)
        
        # 使用JavaScript點擊，因為單選按鈕可能被CSS隱藏
        driver.execute_script("arguments[0].click();", radio_element)
        flush_print(f"✓ 已選擇土地類型：{type_name}")
        
        return True
        
    except Exception as e:
        flush_print(f"✗ 選擇土地類型失敗：{e}")
        return False


def fill_form_interactive_finance_fixed(driver, payload, owner, transcript_dir, sequence_num, cpi_data=None, form_data=None):
    """修正版財政部表單填寫函數 - 使用預載CPI資料"""
    district = payload.get("district", "")
    section = payload.get("section_hint", "")
    lot_display = payload.get("lot_display", "")
    area = payload.get("area", "").replace("平方公尺", "").replace(",", "").strip()
    owner_name = owner.get("name", "")
    
    flush_print(f"開始填寫: {district}{section} {lot_display} - {owner_name}", "important")
    
    try:
        time.sleep(0.2)
        
        # 🔥 修改：如果外部已提供 form_data 就用它，否則才解析
        if form_data is None:
            all_records = extract_form_data_from_json(payload)
            if not all_records:
                return "failed", None, {}
            form_data = all_records[0]  # 使用第一組
        
        # 檢查 form_data 是否為列表（多組模式）
        if isinstance(form_data, list):
            if not form_data:
                return "failed", None, {}
            form_data = form_data[0]  # 取第一組
        
        # 等待頁面完全載入
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        time.sleep(0.5)
        
        total_fields = 7
        success_count = 0
        
        # 1. 處理前次移轉年月
        if form_data['year_month']:
            if handle_year_month_dropdown(driver, form_data['year_month']):
                success_count += 1
                flush_print(f"前次移轉年月完成", "important")
        
        # 2. 填寫土地公告現值
        if form_data['current_value']:
            if fill_element_by_id_no_scroll(driver, "currVal", form_data['current_value'], "土地公告現值"):
                success_count += 1
        
        # 3. 填寫前次移轉現值
        if form_data['previous_value']:
            prev_value = str(form_data['previous_value']).split('.')[0]
            if fill_element_by_id_no_scroll(driver, "origVal", prev_value, "前次移轉現值"):
                success_count += 1
        
        # 4. 處理消費者物價指數（依該所有權人該組的前次移轉年月查 CPI）
        flush_print("CPI處理開始", "important")
        cpi_filled = False

        try:
            year_month_code = form_data.get('year_month', '')
            year_month_display = form_data.get('year_month_display', '')
            flush_print(f"使用當前組年月查 CPI：{year_month_display} (代碼:{year_month_code})", "important")

            if year_month_code and cpi_data:
                cpi_value = find_cpi_for_year_month(year_month_code, cpi_data)

                if cpi_value:
                    if fill_element_by_id_no_scroll(driver, "priceIdx", cpi_value, "消費者物價指數"):
                        cpi_filled = True
                        form_data['cpi_value'] = cpi_value
                        flush_print(f"CPI填入成功: {cpi_value}", "important")
                else:
                    flush_print(f"CPI查詢失敗: {year_month_code}", "important")
            else:
                flush_print("無CPI資料或年月資料", "important")

        except Exception as e:
            flush_print(f"CPI處理錯誤: {e}", "important")
        
        # CPI成功才計入統計
        if cpi_filled:
            success_count += 1
        
        # 5. 填寫土地宗地面積
        if fill_element_by_id_no_scroll(driver, "area", area, "土地宗地面積"):
            success_count += 1
        
        # 6. 選擇土地類型（依謄本「使用分區/使用地類別」動態判斷）
        # 都市土地：兩欄位皆為「（空白）」
        # 非都市土地：任一欄位有實際內容（如「一般農業區」「丁種建築用地」）
        _land_type, _land_type_name = determine_land_type(payload)
        if select_land_type_radio(driver, _land_type):
            success_count += 1
            flush_print(f"  → 已選擇：{_land_type_name}", "important")
        
        # 7. 填寫權利範圍
        range_success = 0
        if form_data['numerator']:
            if fill_element_by_id_no_scroll(driver, "numerator", form_data['numerator'], "權利範圍分子"):
                range_success += 1
        
        if form_data['denominator']:
            if fill_element_by_id_no_scroll(driver, "denominator", form_data['denominator'], "權利範圍分母"):
                range_success += 1
        
        if range_success == 2:
            success_count += 1

        # 🔥 檢查資料完整性
        if success_count == total_fields:
            flush_print(f"填寫統計: {success_count}/{total_fields} 項，資料完整", "important")
        else:
            flush_print(f"填寫統計: {success_count}/{total_fields} 項，資料不完整", "important")
            flush_print(f"\n⚠️  警告：表單資料不完整，請檢查以下可能問題：", "important")
            flush_print(f"  - 土地登記謄本是否包含完整資料", "important")
            flush_print(f"  - 是否有欄位填寫失敗", "important")
            flush_print(f"\n程式將暫停執行，請按 Enter 確認後繼續或 Ctrl+C 取消...", "important")
            input()  # 等待使用者確認

        # 8. 提交試算
        try:
            submit_btn = driver.find_element(By.CSS_SELECTOR, 'button[type="submit"]')
            if submit_btn.is_displayed() and submit_btn.is_enabled():
                driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", submit_btn)
                time.sleep(0.2)
                submit_btn.click()
                flush_print(f"試算按鈕已點擊", "important")
            else:
                flush_print(f"試算按鈕不可用", "important")
        except Exception as e:
            flush_print(f"點擊試算按鈕失敗: {e}", "important")

        # 🔥 等試算結果出現（原本固定等 2 秒，改成偵測到結果就繼續，最多 5 秒）
        try:
            WebDriverWait(driver, 5).until(
                lambda d: 'NT$' in d.find_element(By.TAG_NAME, 'body').text
                          or 'NT&nbsp;$' in d.find_element(By.TAG_NAME, 'body').text
            )
        except TimeoutException:
            flush_print("  試算結果等待逾時（5 秒），繼續執行", "important")

        # 9. 保存PDF
        header_text = f"編號{sequence_num:03d} | {owner_name} | {district}{section} {lot_display}"
        result_pdf_path = save_page_as_pdf(driver, transcript_dir, f"result_{sequence_num:03d}_{district}_{lot_display}_{owner_name}")
        
        if result_pdf_path:
            final_result_pdf = os.path.join(transcript_dir, f"result_{sequence_num:03d}_{district}_{lot_display}_{owner_name}_with_header.pdf")
            
            if add_header_to_pdf_at_specific_location(result_pdf_path, header_text, final_result_pdf):
                if os.path.exists(result_pdf_path):
                    os.remove(result_pdf_path)
                result_pdf_path = final_result_pdf
            
            flush_print(f"PDF已保存", "important")
            
            if success_count >= 5:
                return "filled", result_pdf_path, form_data
            else:
                flush_print(f"填寫成功率過低: {success_count}/7", "important")
                return "partial", result_pdf_path, form_data
        else:
            return "failed", None, {}
                    
    except Exception as e:
        flush_print(f"填寫表單錯誤: {e}", "important")
        return "failed", None, {}

#=========================================================

def enhanced_main():
    """增強版主程式，包含PDF合併和稅額統計功能"""
    parser = argparse.ArgumentParser(description="財政部土地增值稅試算自動化 - 增強版")
    parser.add_argument("--json", help="指定JSON檔路徑", default=None)
    parser.add_argument("--json-dir", help="JSON目錄", default="output")
    parser.add_argument("--show", action="store_true", help="顯示瀏覽器視窗")
    parser.add_argument("--headless", action="store_true", help="使用無頭模式")
    parser.add_argument("--base-dir", help="主程式目錄", default=None)  # 🔥 支援 BASE_DIR 架構
    args = parser.parse_args()
    
    json_path = args.json
    if not json_path:
        base = args.json_dir or "output"
        print(f"\n【JSON 來源】:自動搜尋目錄")
        print(f"  目錄路徑：{os.path.abspath(base)}")
        json_path = latest_json_in(base)
        if not json_path:
            print("\n❌ 錯誤：找不到 JSON 檔案")
            print(f"  請確認 {base} 資料夾中有 .json 檔案")
            return
        print(f"  找到最新檔案：{os.path.basename(json_path)}")
        print(f"  完整路徑：{json_path}")
    else:
        print(f"\n【JSON 來源】：指定檔案")
        print(f"  完整路徑：{json_path}")
        if not os.path.exists(json_path):
            print(f"\n❌ 錯誤：檔案不存在")
            print(f"  路徑：{json_path}")
            return
    
    print("-"*39 + "\n")
    
    # 步驟1: 下載最新CPI資料
    flush_print("下載最新CPI資料...", "important")
    global_cpi_data = download_and_parse_cpi_data()
    flush_print(f"CPI載入完成: {len(global_cpi_data)} 筆", "important")

    # 🔥 檢查 CPI 資料是否成功下載
    if not global_cpi_data or len(global_cpi_data) == 0:
        flush_print("\n⚠️ 警告：無法自動下載 CPI 資料", "important")
        flush_print("嘗試使用新的下載機制...", "important")

        # 使用新的下載機制（包含用戶輸入）
        cpi_file = download_cpi_excel()
        if cpi_file:
            global_cpi_data = parse_cpi_data(cpi_file)
            flush_print(f"✓ CPI重新載入完成: {len(global_cpi_data)} 筆", "important")

            # 🔥 檢查第二次解析是否成功
            if not global_cpi_data or len(global_cpi_data) == 0:
                flush_print("\n" + "="*39, "important")
                flush_print("✗ CPI 檔案下載成功，但解析失敗", "important")
                flush_print("="*39, "important")
                flush_print("可能原因：", "important")
                flush_print("1. 檔案格式不正確", "important")
                flush_print("2. 缺少必要的套件（xlrd）", "important")
                flush_print("3. 檔案內容與預期不符", "important")
                flush_print("\n程式無法繼續執行。", "important")
                return
        else:
            flush_print("\n✗ 無法取得 CPI 資料，程式無法繼續", "important")
            return
    
    # 步驟2: 使用已驗證的 JSON 路徑
    flush_print("解析JSON資料...", "important")
    payloads = extract_payloads_from_json(json_path, global_cpi_data)
    if not payloads:
        flush_print("沒有有效的資料可處理")
        return
    
    # 步驟3: 建立輸出目錄
    flush_print("\n步驟3: 建立輸出目錄")
    p0 = payloads[0]
    district0 = p0.get("district", "").strip()
    section0 = p0.get("section_hint", "").strip()
    # 🔥 使用 lot_display 避免地號格式錯誤（例如 "1375-0000" 被處理成 "13750000"）
    lot_display = p0.get("lot_display", "") or p0.get("lot_number", "")

    # 🔥 優先用 data.json 第一筆當資料夾（避免謄本與 data.json 順序不同導致檔案散落）
    # 規則：只要任一謄本地號在 data.json 中存在 → 用 data.json[0] 建資料夾
    # 完全無交集（不同地段）→ 退回謄本第一筆
    def _norm_lot(lot):
        s = str(lot or '').strip().replace('地號', '')
        try:
            if '-' in s:
                a, b = s.split('-', 1)
                return f"{int(a)}-{int(b)}"
            if len(s) >= 8:
                return f"{int(s[:4])}-{int(s[4:])}"
            return f"{int(s)}-0"
        except (ValueError, IndexError):
            return s

    folder_name = None
    try:
        from base_dir_helper import get_data_json_path
        data_json_path = get_data_json_path()
        if os.path.exists(data_json_path):
            with open(data_json_path, 'r', encoding='utf-8') as f:
                data_list = json.load(f)
            if data_list:
                data_set = {(
                    d.get('area', '').strip(),
                    d.get('section', '').strip(),
                    _norm_lot(d.get('lot_number', ''))
                ) for d in data_list}
                has_overlap = any(
                    (p.get('district', '').strip(),
                     p.get('section_hint', '').strip(),
                     _norm_lot(p.get('lot_number', ''))) in data_set
                    for p in payloads
                )
                if has_overlap:
                    first = data_list[0]
                    folder_name = label_for_lot_dir(
                        first.get('area', '').strip(),
                        first.get('section', '').strip(),
                        first.get('lot_number', '')
                    )
                    flush_print(f"✓ 偵測到地號與 data.json 重疊，依 data.json 第一筆建立資料夾：{folder_name}")
    except Exception as _e:
        flush_print(f"[警告] 讀取 data.json 比對失敗：{_e}")

    if not folder_name:
        folder_name = label_for_lot_dir(district0, section0, lot_display)
        flush_print(f"✓ 與 data.json 無交集，依謄本第一筆建立資料夾：{folder_name}")

    # 🔥 使用 get_work_folder 確保在主程式目錄下建立資料夾
    from base_dir_helper import get_work_folder
    outdir_root = get_work_folder(folder_name)

    transcript_dir = os.path.join(outdir_root, "1.基本資料")
    os.makedirs(transcript_dir, exist_ok=True)
    flush_print(f"輸出目錄：{transcript_dir}")

    # 步驟4: 啟動瀏覽器
    flush_print("\n啟動瀏覽器，請稍侯…", "important")
    show_gui_mode = not args.headless
    driver = start_driver(show_gui=show_gui_mode)
    
    results = []
    all_pdfs = []
    sequence_num = 1
    
    try:
        # 步驟5: 嘗試開啟財政部網站
        flush_print("\n步驟5: 開啟財政部網站")
        success = try_open_finance_website(driver)
        
        if not success:
            flush_print("✗ 無法開啟財政部網站，程式結束")
            return
        
        # 步驟6: 等待表單載入
        flush_print("\n步驟6: 等待表單載入")
        if not wait_for_form_elements(driver):
            flush_print("✗ 表單元素載入失敗，程式結束")
            return
        
        # 步驟7: 處理每筆土地資料
        flush_print(f"\n 開始處理土地資料，共 {len(payloads)} 筆", "important")

        # 🔥 進度回報給主程式（格式：__PROGRESS__:current/total:elapsed_sec:eta_sec）
        #   main.py 會攔截這個訊號並在 UI 顯示進度條
        _progress_start_time = time.time()
        _total_payloads = len(payloads)
        print(f"__PROGRESS__:0/{_total_payloads}:0:0", flush=True)

        for pi, payload in enumerate(payloads, 1):
            owners = payload.get("owners", [])
            co_ownership_groups = payload.get("co_ownership_groups", {})

            flush_print(f"\n=== 處理第 {pi}/{len(payloads)} 筆土地資料 ===", "important")
            flush_print(f"地點：{payload.get('district', '')}{payload.get('section_hint', '')} 地號 {payload.get('lot_display', '')}")
            flush_print(f"所有權人數量：{len(owners)}")

            # 🔥 顯示公同共有群組資訊
            if co_ownership_groups:
                flush_print(f"公同共有群組數：{len(co_ownership_groups)}")
                for group_id, group_members in co_ownership_groups.items():
                    member_names = [m.get('name', '') for m in group_members]
                    rights_range = group_members[0].get('rights_range', '') if group_members else ''
                    flush_print(f"  - {group_id}: {len(group_members)}人 ({', '.join(member_names[:3])}{'等' if len(member_names) > 3 else ''}) {rights_range}")

            # 🔥 標記已處理的公同共有群組
            processed_co_ownership_groups = set()
            
            for oi, owner in enumerate(owners, 1):
                # 🔥 檢查是否為公同共有成員
                is_co_ownership = owner.get("is_co_ownership", False)
                co_ownership_group = owner.get("co_ownership_group")

                # 如果是公同共有成員且該群組已處理過，則跳過
                if is_co_ownership and co_ownership_group in processed_co_ownership_groups:
                    flush_print(f"\n--- 跳過第 {oi} 位所有權人：{owner.get('name', '')} (已處理公同共有群組: {co_ownership_group}) ---", "important")
                    continue

                # 🔥 建立顯示名稱（如果是公同共有群組的第一人，顯示群組資訊）
                owner_display_name = owner.get('name', '')
                if is_co_ownership and co_ownership_group:
                    group_members = co_ownership_groups.get(co_ownership_group, [])
                    if len(group_members) > 1:
                        member_names = [m.get('name', '') for m in group_members]
                        rights_range = owner.get('rights_range', '')
                        owner_display_name = f"{owner.get('name', '')}等{len(group_members)}人{rights_range}"
                    # 標記群組已處理
                    processed_co_ownership_groups.add(co_ownership_group)

                flush_print(f"\n--- 處理第 {oi} 位所有權人：{owner_display_name} ---", "important")

                try:
                    # 🔥 解析所有組資料（傳入當前所有權人索引，oi 是 1-based，轉為 0-based）
                    all_records = extract_form_data_from_json(payload, oi - 1)
                    
                    if not all_records:
                        flush_print("無法解析任何資料組，跳過此所有權人")

                        # 🔥 取得統一編號（使用索引）
                        owner_id = ""
                        ownership_data = payload.get("所有權部", [])
                        owner_idx = oi - 1  # oi 是 1-based
                        if isinstance(ownership_data, list) and 0 <= owner_idx < len(ownership_data):
                            owner_record = ownership_data[owner_idx]
                            if isinstance(owner_record, dict):
                                owner_id = owner_record.get("統一編號", "")

                        results.append({
                            "sequence_num": sequence_num,
                            "group_number": 0,
                            "total_groups": 0,
                            "district": payload.get("district", ""),
                            "section": payload.get("section_hint", ""),
                            "lot_display": payload.get("lot_display", ""),
                            "owner_name": owner_display_name,  # 🔥 使用含公同共有註記的名稱
                            "owner_id": owner_id,  # 🔥 加入統一編號
                            "area": payload.get("area", ""),
                            "status": "解析失敗",
                            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "tax_amounts": {"一般稅額": "", "自用稅額": ""}
                        })
                        sequence_num += 1
                        continue
                    
                    # 🔥 循環處理每一組資料
                    for record_idx, form_data in enumerate(all_records, 1):
                        flush_print(f"\n=== {owner_display_name} 第 {record_idx}/{len(all_records)} 組 (編號{sequence_num:03d}) ===", "important")

                        # 🔥 取得統一編號（使用索引，避免第二類謄本姓名重複問題）
                        owner_id = ""
                        ownership_data = payload.get("所有權部", [])
                        owner_idx = oi - 1  # oi 是 1-based，轉為 0-based
                        if isinstance(ownership_data, list) and 0 <= owner_idx < len(ownership_data):
                            owner_record = ownership_data[owner_idx]
                            if isinstance(owner_record, dict):
                                owner_id = owner_record.get("統一編號", "")
                        
                        # 填寫表單（使用當前組資料）
                        fill_result, filled_pdf_path, _ = fill_form_interactive_finance_fixed(
                            driver, payload, owner, transcript_dir, sequence_num, 
                            global_cpi_data, form_data
                        )
                        
                        if fill_result == "filled":
                            if filled_pdf_path:
                                all_pdfs.append(filled_pdf_path)
                            
                            # 提交試算
                            submit_success, result_pdf_path, tax_amounts = submit_form_and_save_result_finance(
                                driver, transcript_dir,
                                f"財政部試算_{payload.get('district', '')}_{payload.get('lot_display', '')}_{owner.get('name', '')}_第{record_idx}組",
                                sequence_num, owner.get('name', ''), 
                                payload.get('district', ''), payload.get('section_hint', ''), 
                                payload.get('lot_display', '')
                            )
                            
                            if result_pdf_path:
                                all_pdfs.append(result_pdf_path)
                            
                            # 🔥 記錄結果（加入統一編號和公同共有註記）
                            result = {
                                "sequence_num": sequence_num,
                                "group_number": record_idx,
                                "total_groups": len(all_records),
                                "district": payload.get("district", ""),
                                "section": payload.get("section_hint", ""),
                                "lot_display": payload.get("lot_display", ""),
                                "owner_name": owner_display_name,  # 🔥 使用含公同共有註記的名稱
                                "owner_id": owner_id,  # 🔥 統一編號
                                "area": payload.get("area", ""),
                                "status": "成功" if submit_success else "填寫成功但提交失敗",
                                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                "tax_amounts": tax_amounts,
                                "payload_data": form_data
                            }
                            
                            results.append(result)
                            
                            # 重新載入準備下一組
                            if record_idx < len(all_records) or oi < len(owners):
                                driver.refresh()
                                time.sleep(0.5)
                        else:
                            # 🔥 失敗情況也加入統一編號
                            result = {
                                "sequence_num": sequence_num,
                                "group_number": record_idx,
                                "total_groups": len(all_records),
                                "district": payload.get("district", ""),
                                "section": payload.get("section_hint", ""),
                                "lot_display": payload.get("lot_display", ""),
                                "owner_name": owner.get("name", ""),
                                "owner_id": owner_id,  # 🔥 統一編號
                                "area": payload.get("area", ""),
                                "status": "跳過",
                                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                "tax_amounts": {"一般稅額": "", "自用稅額": ""}
                            }
                            results.append(result)
                        
                        sequence_num += 1
                    
                except Exception as e:
                    flush_print(f"處理時發生錯誤：{e}")
                    traceback.print_exc()
                    
                    # 🔥 錯誤情況也取得統一編號
                    owner_id = ""
                    try:
                        ownership_data = payload.get("所有權部", [])
                        owner_idx = oi - 1  # oi 是 1-based
                        if isinstance(ownership_data, list) and 0 <= owner_idx < len(ownership_data):
                            owner_record = ownership_data[owner_idx]
                            if isinstance(owner_record, dict):
                                owner_id = owner_record.get("統一編號", "")
                    except:
                        pass
                    
                    results.append({
                        "sequence_num": sequence_num,
                        "group_number": 0,
                        "total_groups": 0,
                        "district": payload.get("district", ""),
                        "section": payload.get("section_hint", ""),
                        "lot_display": payload.get("lot_display", ""),
                        "owner_name": owner.get("name", ""),
                        "owner_id": owner_id,  # 🔥 統一編號
                        "area": payload.get("area", ""),
                        "status": f"錯誤: {str(e)}",
                        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "tax_amounts": {"一般稅額": "", "自用稅額": ""}
                    })
                    sequence_num += 1

            # 🔥 本筆土地處理完，回報進度給主程式
            _elapsed = int(time.time() - _progress_start_time)
            _avg = _elapsed / pi if pi > 0 else 0
            _eta = int(_avg * (_total_payloads - pi))
            print(f"__PROGRESS__:{pi}/{_total_payloads}:{_elapsed}:{_eta}", flush=True)

        # 步驟10: 產生增強版統計報告
        flush_print("\n步驟10: 產生增強版統計報告", "important")
        try:
            # 🔥 先檢查 results 內容
            flush_print(f"準備生成 Excel，共 {len(results)} 筆資料", "important")
            
            # 🔥 檢查是否有統一編號
            has_owner_id = any(r.get('owner_id') for r in results)
            flush_print(f"統一編號檢查：{'有' if has_owner_id else '無'}", "important")
            
            section_name = section0
            lot_number = p0.get("lot_display", "") or lot80[:4].lstrip("0")
            excel_path = os.path.join(transcript_dir, f"01.1_土地增值稅_財政部版_{section_name}_{lot_number}_{now_ymd()}.xlsx")

            # flush_print(f"Excel 輸出路徑：{excel_path}", "important")  # 🔥 訊息簡化

            success = create_enhanced_excel_summary(results, excel_path)

            # if success:  # 🔥 訊息簡化（已在 create_enhanced_excel_summary 內顯示）
            #     flush_print("✓ Excel 報表生成成功", "important")
            # else:
            #     flush_print("✗ Excel 報表生成失敗", "important")
                
        except Exception as e:
            flush_print(f"✗ 生成 Excel 時發生錯誤：{e}", "important")
            flush_print(f"錯誤詳情：{traceback.format_exc()}", "important")
        
        flush_print(f"\n✓ 財政部版處理完成！")
        flush_print(f"輸出目錄：{transcript_dir}")
        flush_print(f"處理筆數：{len(results)}")
        merge_and_clean_pdfs(transcript_dir, section_name, lot_number)
        
        # 顯示處理結果摘要
        success_count = len([r for r in results if r.get('status') == '成功'])
        skip_count = len([r for r in results if r.get('status') == '跳過'])
        error_count = len([r for r in results if '錯誤' in r.get('status', '')])
        
        flush_print(f"成功處理：{success_count} 筆", "important")
        flush_print(f"跳過處理：{skip_count} 筆", "important") 
        flush_print(f"發生錯誤：{error_count} 筆", "important")
        
        total_general = sum([int(str(r.get('tax_amounts', {}).get('一般稅額', 0)).replace(',', '') or 0) for r in results])
        total_once = sum([int(str(r.get('tax_amounts', {}).get('自用稅額', 0)).replace(',', '') or 0) for r in results])
        
        if total_general or total_once:
            flush_print(f"\n=== 稅額統計摘要 ===")
            flush_print(f"自用稅額總計：{total_once:,} 元", "important")
            flush_print(f"一般稅額總計：{total_general:,} 元", "important")
        
        # 🔥 將稅額結果寫入 JSON 供 main.py 使用
        try:
            # 建立 "4.其他相關" 資料夾
            other_dir = os.path.join(outdir_root, "4.其他相關")
            os.makedirs(other_dir, exist_ok=True)

            tax_result_file = os.path.join(other_dir, "財政部_tax_result.json")
            
            tax_summary = {
                "自用住宅稅額": total_once,
                "一般土地稅額": total_general,
                "計算日期": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "地段": section_name,
                "地號": lot_number,
                "詳細資料": []
            }
            
            # 加入每位所有權人的明細
            for result in results:
                if result.get('status') == '成功':
                    owner_detail = {
                        "所有權人": result.get('owner_name', ''),
                        "統一編號": result.get('owner_id', ''),
                        "組別": f"第{result.get('group_number', 0)}組",
                        "自用稅額": result.get('tax_amounts', {}).get('自用稅額', 0),
                        "一般稅額": result.get('tax_amounts', {}).get('一般稅額', 0)
                    }
                    tax_summary["詳細資料"].append(owner_detail)
            
            with open(tax_result_file, 'w', encoding='utf-8') as f:
                json.dump(tax_summary, f, ensure_ascii=False, indent=2)
            
            flush_print(f"\n\033[96m✓ 稅額結果已儲存：{tax_result_file}\033[0m", "important")
            flush_print(f"  → 【物件資料編輯器】可讀取此檔案自動帶入稅額", "important")
            
        except Exception as e:
            flush_print(f"儲存稅額結果失敗：{e}", "important")    
    except Exception as e:
        flush_print(f"程式執行時發生錯誤：{e}")
    finally:
        flush_print("\n自動關閉瀏覽器...")
        try:
            driver.quit()
            flush_print("瀏覽器已關閉")
        except Exception as e:
            flush_print(f"關閉瀏覽器時發生錯誤：{e}")
        flush_print("財政部版程式執行完成")

if __name__ == "__main__":
    enhanced_main()